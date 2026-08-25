import io
import base64
import re
import traceback
from dataclasses import dataclass
from datetime import datetime
from typing import Optional
from pathlib import Path

import pandas as pd
import altair as alt
import streamlit as st
from openpyxl import load_workbook
from textwrap import dedent

from ui.styles import apply_styles


APP_VERSION = "V61.1 - Maritex Inventory Control"

st.set_page_config(
    page_title="Centralizador Stock Marketplaces",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="expanded",
)


# ============================================================
# PLANTILLAS PERSISTENTES
# ============================================================
BASE_DIR = Path(__file__).resolve().parent
TEMPLATES_DIR = BASE_DIR / "templates"
TEMPLATES_DIR.mkdir(parents=True, exist_ok=True)

DATA_DIR = BASE_DIR / "data"
DATA_DIR.mkdir(parents=True, exist_ok=True)
METRICS_FILE = DATA_DIR / "metricas.csv"

SAVED_TEMPLATE_FILES = {
    "Paris Marketplace": "plantilla_paris.xlsx",
    "Mercado Libre": "plantilla_meli.xlsx",
}


def template_path_for(marketplace_name: str) -> Path:
    filename = SAVED_TEMPLATE_FILES.get(marketplace_name)
    if not filename:
        raise ValueError(f"No existe plantilla persistente configurada para {marketplace_name}.")
    return TEMPLATES_DIR / filename


@st.cache_data(show_spinner=False)
def read_saved_template_bytes(path_text: str, modified_ns: int) -> bytes:
    return Path(path_text).read_bytes()


def get_saved_template(marketplace_name: str):
    path = template_path_for(marketplace_name)
    if not path.exists():
        return None, None
    stat = path.stat()
    return read_saved_template_bytes(str(path), stat.st_mtime_ns), path.name


def save_marketplace_template(marketplace_name: str, uploaded_file) -> Path:
    path = template_path_for(marketplace_name)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(uploaded_file.getvalue())
    read_saved_template_bytes.clear()
    return path



def read_excel_flexline(raw: bytes, filename: str) -> pd.DataFrame:
    """
    Lee exportaciones Excel de Flexline.
    Soporta:
    - .xlsx mediante openpyxl
    - .xls mediante xlrd
    - .xls que en realidad contienen una tabla HTML (algunas exportaciones ERP)
    """
    name = str(filename).lower()
    head = raw[:4096].lstrip().lower()

    # Algunas exportaciones antiguas usan extensión XLS pero contienen HTML.
    looks_like_html = (
        head.startswith(b"<")
        and (
            b"<html" in head
            or b"<table" in head
            or b"<!doctype" in head
        )
    )

    if looks_like_html:
        try:
            html_text = raw.decode("utf-8-sig")
        except UnicodeDecodeError:
            html_text = raw.decode("cp1252", errors="replace")

        tables = pd.read_html(io.StringIO(html_text))
        if not tables:
            raise ValueError("El archivo XLS/HTML no contiene tablas legibles.")
        return tables[0].astype(str)

    if name.endswith(".xlsx"):
        return pd.read_excel(
            io.BytesIO(raw),
            dtype=str,
            engine="openpyxl",
        )

    if name.endswith(".xls"):
        try:
            return pd.read_excel(
                io.BytesIO(raw),
                dtype=str,
                engine="xlrd",
            )
        except ImportError as exc:
            raise ValueError(
                "Para leer archivos Excel .xls es necesario instalar la dependencia "
                "'xlrd>=2.0.1'. Agrégala al archivo requirements.txt del proyecto."
            ) from exc
        except Exception as exc:
            raise ValueError(
                f"No fue posible leer el archivo XLS: {exc}"
            ) from exc

    raise ValueError("Formato Excel no soportado.")


def parse_erp_sales_dates(series: pd.Series) -> pd.Series:
    """
    Parser robusto para fechas de ERP Ventas.

    Problema que resuelve:
    algunas exportaciones XLS de Flexline entregan fechas con "/" en formato
    MM/DD/YYYY, mientras otras exportaciones usan DD/MM/YYYY o DD-MM-YYYY.
    Pandas puede invertir mes/día en fechas ambiguas como 08/12/2026.

    Estrategia:
    - reconoce seriales de fecha Excel;
    - reconoce YYYY-MM-DD / YYYY/MM/DD;
    - para fechas con "/" compara interpretación day-first vs month-first;
    - usa evidencia del propio archivo y evita elegir una interpretación
      que genere ventas futuras;
    - para fechas con "-" mantiene preferencia DD-MM-YYYY, usada por Flexline.
    """
    raw = series.copy()

    # Resultado inicialmente vacío.
    result = pd.Series(pd.NaT, index=raw.index, dtype="datetime64[ns]")

    # Trabajar con texto normalizado.
    txt = raw.fillna("").astype(str).str.strip()
    txt = txt.str.replace(r"\s+", " ", regex=True)

    # --------------------------------------------------------
    # 1) Seriales Excel (ej. 45897, 45897.0)
    # --------------------------------------------------------
    numeric = pd.to_numeric(txt.str.replace(",", ".", regex=False), errors="coerce")
    excel_mask = numeric.between(20000, 80000, inclusive="both")
    if excel_mask.any():
        result.loc[excel_mask] = pd.to_datetime(
            numeric.loc[excel_mask],
            unit="D",
            origin="1899-12-30",
            errors="coerce",
        )

    remaining = result.isna() & txt.ne("")

    # --------------------------------------------------------
    # 2) Formato ISO / año primero
    # --------------------------------------------------------
    iso_mask = remaining & txt.str.match(r"^\d{4}[-/]\d{1,2}[-/]\d{1,2}")
    if iso_mask.any():
        result.loc[iso_mask] = pd.to_datetime(
            txt.loc[iso_mask],
            errors="coerce",
            yearfirst=True,
        )

    remaining = result.isna() & txt.ne("")

    # --------------------------------------------------------
    # 3) Fechas con guión: preferencia DD-MM-YYYY
    # --------------------------------------------------------
    dash_mask = remaining & txt.str.match(r"^\d{1,2}-\d{1,2}-\d{4}")
    if dash_mask.any():
        dash_values = txt.loc[dash_mask]
        parsed_dash = pd.to_datetime(
            dash_values,
            errors="coerce",
            dayfirst=True,
        )
        result.loc[dash_mask] = parsed_dash

    remaining = result.isna() & txt.ne("")

    # --------------------------------------------------------
    # 4) Fechas con slash: detectar orientación del archivo
    # --------------------------------------------------------
    slash_mask = remaining & txt.str.match(r"^\d{1,2}/\d{1,2}/\d{4}")
    if slash_mask.any():
        slash_values = txt.loc[slash_mask]

        parts = slash_values.str.extract(
            r"^(?P<a>\d{1,2})/(?P<b>\d{1,2})/(?P<y>\d{4})"
        )
        a = pd.to_numeric(parts["a"], errors="coerce")
        b = pd.to_numeric(parts["b"], errors="coerce")

        # Evidencia no ambigua del archivo.
        evidence_dayfirst = int((a > 12).sum())   # 24/08 => DD/MM
        evidence_monthfirst = int((b > 12).sum()) # 08/24 => MM/DD

        candidate_dayfirst = pd.to_datetime(
            slash_values,
            errors="coerce",
            dayfirst=True,
        )
        candidate_monthfirst = pd.to_datetime(
            slash_values,
            errors="coerce",
            dayfirst=False,
        )

        today = pd.Timestamp(datetime.now().date())
        future_limit = today + pd.Timedelta(days=1)

        # Penalizar interpretaciones que generan ventas futuras.
        future_dayfirst = int(
            (candidate_dayfirst.dropna().dt.normalize() > future_limit).sum()
        )
        future_monthfirst = int(
            (candidate_monthfirst.dropna().dt.normalize() > future_limit).sum()
        )

        invalid_dayfirst = int(candidate_dayfirst.isna().sum())
        invalid_monthfirst = int(candidate_monthfirst.isna().sum())

        if evidence_dayfirst > evidence_monthfirst:
            use_dayfirst = True
        elif evidence_monthfirst > evidence_dayfirst:
            use_dayfirst = False
        elif future_dayfirst != future_monthfirst:
            # En archivos de ventas, elegir la interpretación que no invente
            # meses futuros es una señal muy fuerte.
            use_dayfirst = future_dayfirst < future_monthfirst
        elif invalid_dayfirst != invalid_monthfirst:
            use_dayfirst = invalid_dayfirst < invalid_monthfirst
        else:
            # Último fallback: formato regional histórico de Flexline.
            use_dayfirst = True

        result.loc[slash_mask] = (
            candidate_dayfirst if use_dayfirst else candidate_monthfirst
        )

    remaining = result.isna() & txt.ne("")

    # --------------------------------------------------------
    # 5) Fallback final para timestamps ya textualizados por Excel
    # --------------------------------------------------------
    if remaining.any():
        # Probar primero year-first (ej. 2026-08-24 00:00:00).
        fallback = pd.to_datetime(
            txt.loc[remaining],
            errors="coerce",
            yearfirst=True,
        )
        result.loc[remaining] = fallback

    return result


@st.cache_data(show_spinner=False)
def read_metrics_data(raw: bytes, filename: str) -> pd.DataFrame:
    """Lee la exportación documental de Flexline usada por la página Métricas."""
    name = str(filename).lower()

    if name.endswith((".xlsx", ".xls")):
        df = read_excel_flexline(raw, filename)
    elif name.endswith(".csv"):
        last_error = None
        df = None
        for encoding in ("utf-8-sig", "cp1252", "latin1"):
            try:
                candidate = pd.read_csv(
                    io.BytesIO(raw),
                    sep=None,
                    engine="python",
                    encoding=encoding,
                    dtype=str,
                )
                if len(candidate.columns) > 1:
                    df = candidate
                    break
            except Exception as exc:
                last_error = exc

        if df is None:
            raise ValueError(f"No fue posible leer el archivo ERP Ventas: {last_error}")
    else:
        raise ValueError("ERP Ventas debe ser un archivo CSV, XLS o XLSX.")

    # Eliminar columnas completamente vacías / auxiliares.
    df = df.dropna(axis=1, how="all").copy()

    if "Fecha" in df.columns:
        # V58: no inferir mes/día a ciegas. Flexline puede exportar
        # XLS con fechas DD/MM o MM/DD según el formato del archivo.
        df["Fecha_dt"] = parse_erp_sales_dates(df["Fecha"])

    for col in ("Total", "TotalIngreso", "Monto", "Importe", "Valor", "Peso Total", "Capacidad"):
        if col in df.columns:
            df[f"{col}_num"] = df[col].apply(parse_number)

    # V55: seleccionar automáticamente la fuente monetaria correcta.
    df = ensure_sales_amount_column(df)

    return df


def save_metrics_file(uploaded_file) -> Path:
    """Guarda la última exportación de métricas para reutilizarla."""
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    target = DATA_DIR / "metricas.csv"
    target.write_bytes(uploaded_file.getvalue())
    read_metrics_data.clear()
    return target


def load_saved_metrics():
    if not METRICS_FILE.exists():
        return None, None

    raw = METRICS_FILE.read_bytes()
    return raw, METRICS_FILE.name


ERP_SOURCE_PREFIX = {
    "stock": "erp_stock",
    "ventas": "erp_ventas",
}


def _erp_source_candidates(kind: str):
    prefix = ERP_SOURCE_PREFIX[kind]
    return sorted(DATA_DIR.glob(f"{prefix}.*"), key=lambda p: p.stat().st_mtime, reverse=True)


def save_erp_source(kind: str, uploaded_file) -> Path:
    """Guarda una fuente ERP persistente conservando su extensión."""
    if kind not in ERP_SOURCE_PREFIX:
        raise ValueError(f"Fuente ERP no soportada: {kind}")

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    suffix = Path(uploaded_file.name).suffix.lower() or ".csv"
    prefix = ERP_SOURCE_PREFIX[kind]

    # Mantener una sola versión activa por tipo de fuente.
    for old in DATA_DIR.glob(f"{prefix}.*"):
        try:
            old.unlink()
        except Exception:
            pass

    target = DATA_DIR / f"{prefix}{suffix}"
    target.write_bytes(uploaded_file.getvalue())

    if kind == "stock":
        read_flexline.clear()
    else:
        read_metrics_data.clear()

    return target


def load_erp_source(kind: str):
    """Devuelve bytes, nombre y fecha de la última fuente ERP guardada."""
    candidates = _erp_source_candidates(kind)
    if not candidates:
        return None, None, None
    path = candidates[0]
    return path.read_bytes(), path.name, datetime.fromtimestamp(path.stat().st_mtime)


def ensure_stock_source_loaded():
    if st.session_state.get("shared_stock_bytes") is None:
        raw, name, loaded_at = load_erp_source("stock")
        if raw is not None:
            st.session_state.shared_stock_bytes = raw
            st.session_state.shared_stock_name = name
            st.session_state.shared_stock_loaded_at = loaded_at


def ensure_sales_source_loaded():
    if st.session_state.get("metrics_bytes") is None:
        raw, name, loaded_at = load_erp_source("ventas")
        if raw is not None:
            st.session_state["metrics_bytes"] = raw
            st.session_state["metrics_name"] = name
            st.session_state["metrics_loaded_at"] = loaded_at



MONTH_NAMES_ES = {
    1: "Enero",
    2: "Febrero",
    3: "Marzo",
    4: "Abril",
    5: "Mayo",
    6: "Junio",
    7: "Julio",
    8: "Agosto",
    9: "Septiembre",
    10: "Octubre",
    11: "Noviembre",
    12: "Diciembre",
}


def month_label_es(period_value) -> str:
    """Convierte Period('2026-08', 'M') en 'Agosto 2026'."""
    period = pd.Period(period_value, freq="M")
    return f"{MONTH_NAMES_ES.get(period.month, period.month)} {period.year}"


def available_sales_months(df: pd.DataFrame):
    """Lista de meses con al menos una fecha válida, de más reciente a más antiguo."""
    if "Fecha_dt" not in df.columns:
        return []
    dates = df["Fecha_dt"].dropna()
    if dates.empty:
        return []

    periods = (
        dates.dt.to_period("M")
        .drop_duplicates()
        .sort_values(ascending=False)
        .tolist()
    )
    return periods


def month_bounds(period_value):
    """Inicio y fin calendario de un período mensual."""
    period = pd.Period(period_value, freq="M")
    start = period.start_time.date()
    end = period.end_time.date()
    return start, end


def format_clp(value) -> str:
    try:
        number = int(round(float(value)))
    except Exception:
        number = 0
    return "$" + f"{number:,}".replace(",", ".")


def metric_status(value, true_value="S") -> bool:
    return str(value).strip().upper() == str(true_value).strip().upper()


def resolve_sales_amount_column(df: pd.DataFrame):
    """
    Detecta el campo monetario correcto para ERP Ventas.

    Prioridad:
    1. Total_num, si contiene monto comercial distinto de cero.
    2. TotalIngreso_num, si Total está vacío o suma cero.
    3. Otras variantes conocidas, si aparecen en futuras exportaciones.

    Devuelve (nombre_columna_num, etiqueta_origen, suma_absoluta_detectada).
    """
    candidates = [
        ("Total_num", "Total"),
        ("TotalIngreso_num", "TotalIngreso"),
        ("Monto_num", "Monto"),
        ("Importe_num", "Importe"),
        ("Valor_num", "Valor"),
    ]

    # Si existen grupos comerciales, evaluar solamente documentos relevantes.
    work = df.copy()
    if "TipoDocto" in work.columns:
        try:
            work["_grupo_tmp"] = work["TipoDocto"].apply(classify_commercial_document)
            commercial = work[
                work["_grupo_tmp"].isin(["Factura", "Boleta", "Nota de crédito"])
            ]
            if not commercial.empty:
                work = commercial
        except Exception:
            pass

    diagnostics = []
    for col, label in candidates:
        if col not in work.columns:
            continue

        numeric = pd.to_numeric(work[col], errors="coerce").fillna(0.0)
        abs_sum = float(numeric.abs().sum())
        non_zero = int((numeric.abs() > 0.000001).sum())
        diagnostics.append((col, label, abs_sum, non_zero))

    # Primero: candidato con datos reales, respetando prioridad de negocio.
    for col, label, abs_sum, non_zero in diagnostics:
        if non_zero > 0 and abs_sum > 0:
            return col, label, abs_sum

    # Si todos existen pero están en cero, devolver el primero para diagnóstico.
    if diagnostics:
        col, label, abs_sum, _ = diagnostics[0]
        return col, label, abs_sum

    return None, None, 0.0


def ensure_sales_amount_column(df: pd.DataFrame) -> pd.DataFrame:
    """
    Agrega columnas comerciales normalizadas:
    - VentaMonto_num
    - VentaMontoCampo

    De esta forma todo Métricas Vendedores usa una sola columna,
    independiente de si Flexline exporta Total o TotalIngreso.
    """
    work = df.copy()

    # Crear variantes numéricas para posibles columnas monetarias.
    for col in ("Total", "TotalIngreso", "Monto", "Importe", "Valor"):
        if col in work.columns and f"{col}_num" not in work.columns:
            work[f"{col}_num"] = work[col].apply(parse_number)

    amount_col, amount_label, _ = resolve_sales_amount_column(work)

    if amount_col is None:
        work["VentaMonto_num"] = 0.0
        work["VentaMontoCampo"] = "No detectado"
    else:
        work["VentaMonto_num"] = pd.to_numeric(
            work[amount_col], errors="coerce"
        ).fillna(0.0)
        work["VentaMontoCampo"] = amount_label

    return work


def render_html(content: str) -> None:
    """Renderiza HTML sin que Markdown convierta etiquetas indentadas en código."""
    clean = dedent(content).strip()
    clean = re.sub(r">\s+<", "><", clean)
    st.markdown(clean, unsafe_allow_html=True)

def is_sales_document_type(value: str) -> bool:
    """
    Regla comercial V43:
    Solo suman como venta facturas y boletas.

    Se consideran los nombres reales observados en Flexline:
    - FACTURA...
    - F.VTA...
    - FV...
    - BOLETA...
    - BV...

    Se excluyen cierres, notas de crédito/débito, notas de venta,
    devoluciones, picking, cotizaciones y otros movimientos.
    """
    value = "" if value is None else str(value).strip().upper()

    if not value:
        return False

    # Documentos que nunca deben sumar como venta.
    excluded_tokens = (
        "CIERRE",
        "NC ",
        "NC(",
        "NC(",
        "ND ",
        "ND(",
        "DEVOL",
        "NOTA VENTA",
        "NOTA VTA",
        "NV ",
        "PICKING",
        "COTIZACION",
        "COMPROMISO",
        "GUIA",
    )

    if any(token in value for token in excluded_tokens):
        return False

    # Facturas
    if "FACTURA" in value:
        return True

    if value.startswith("F.VTA"):
        return True

    if value.startswith("FV "):
        return True

    # Boletas
    if "BOLETA" in value:
        return True

    if value.startswith("BV "):
        return True

    return False



def classify_commercial_document(value: str) -> str:
    """
    Clasifica documentos para la gestión comercial:
    - Factura
    - Boleta
    - Nota de crédito
    - Otro
    """
    value = "" if value is None else str(value).strip().upper()

    if not value:
        return "Otro"

    # Notas de crédito: se restan del monto final.
    if (
        value.startswith("NC ")
        or value.startswith("NC(")
        or value.startswith("NC(E")
        or "NC DEVOL" in value
        or "NC REFACT" in value
        or "NC BOLVTA" in value
        or "NC CORPORATIVO" in value
        or "NC VENTA" in value
        or "NOTA CREDITO" in value
        or "NOTA DE CREDITO" in value
    ):
        return "Nota de crédito"

    # Excluir cierres y documentos no comerciales.
    excluded_tokens = (
        "CIERRE",
        "NOTA VENTA",
        "NOTA VTA",
        "NV ",
        "PICKING",
        "COTIZACION",
        "COMPROMISO",
        "GUIA",
        "DEVOL.",
        "DEVOLUCION",
        "ND ",
        "ND(",
    )

    if any(token in value for token in excluded_tokens):
        return "Otro"

    if "FACTURA" in value or value.startswith("F.VTA") or value.startswith("FV "):
        return "Factura"

    if "BOLETA" in value or value.startswith("BV "):
        return "Boleta"

    return "Otro"


def filter_sales_documents(df: pd.DataFrame) -> pd.DataFrame:
    """Solo Facturas y Boletas."""
    if "TipoDocto" not in df.columns:
        return df.iloc[0:0].copy()

    work = df.copy()
    work["Grupo comercial"] = work["TipoDocto"].apply(classify_commercial_document)
    return work[work["Grupo comercial"].isin(["Factura", "Boleta"])].copy()


def filter_commercial_documents(df: pd.DataFrame) -> pd.DataFrame:
    """Facturas + Boletas + Notas de crédito realizadas."""
    if "TipoDocto" not in df.columns:
        return df.iloc[0:0].copy()

    work = df.copy()
    work["Grupo comercial"] = work["TipoDocto"].apply(classify_commercial_document)
    work = work[
        work["Grupo comercial"].isin(
            ["Factura", "Boleta", "Nota de crédito"]
        )
    ].copy()

    # V58: una venta realizada no debe quedar fechada en el futuro.
    # Esto además protege contra inversiones DD/MM <-> MM/DD.
    if "Fecha_dt" in work.columns and work["Fecha_dt"].notna().any():
        tomorrow_end = (
            pd.Timestamp(datetime.now().date())
            + pd.Timedelta(days=2)
            - pd.Timedelta(seconds=1)
        )
        work = work[
            work["Fecha_dt"].isna()
            | (work["Fecha_dt"] <= tomorrow_end)
        ].copy()

    return work


def calculate_commercial_totals(
    df: pd.DataFrame,
    vat_rate: float = 0.19,
):
    """
    Facturas y Boletas suman.
    Notas de crédito restan.
    Total viene con IVA; sin IVA se calcula dividiendo por (1 + tasa IVA).
    """
    if df.empty:
        return {
            "ventas_brutas_con_iva": 0.0,
            "notas_credito_con_iva": 0.0,
            "venta_neta_con_iva": 0.0,
            "ventas_brutas_sin_iva": 0.0,
            "notas_credito_sin_iva": 0.0,
            "venta_neta_sin_iva": 0.0,
            "iva_neto": 0.0,
        }

    work = df.copy()

    if "Grupo comercial" not in work.columns:
        work["Grupo comercial"] = work["TipoDocto"].apply(
            classify_commercial_document
        )

    work = ensure_sales_amount_column(work)
    total_col = "VentaMonto_num"

    sales = pd.to_numeric(
        work.loc[
            work["Grupo comercial"].isin(["Factura", "Boleta"]),
            total_col,
        ],
        errors="coerce",
    ).fillna(0.0).sum()

    # Las NC pueden venir positivas o negativas según la exportación.
    # Siempre las tratamos como monto a descontar.
    credits = pd.to_numeric(
        work.loc[
            work["Grupo comercial"].eq("Nota de crédito"),
            total_col,
        ],
        errors="coerce",
    ).fillna(0.0).abs().sum()

    net_with_vat = sales - credits

    divisor = 1 + float(vat_rate)
    sales_without_vat = sales / divisor
    credits_without_vat = credits / divisor
    net_without_vat = net_with_vat / divisor
    net_vat = net_with_vat - net_without_vat

    return {
        "ventas_brutas_con_iva": float(sales),
        "notas_credito_con_iva": float(credits),
        "venta_neta_con_iva": float(net_with_vat),
        "ventas_brutas_sin_iva": float(sales_without_vat),
        "notas_credito_sin_iva": float(credits_without_vat),
        "venta_neta_sin_iva": float(net_without_vat),
        "iva_neto": float(net_vat),
    }



def seller_period_comparison(
    sales_df: pd.DataFrame,
    start_date,
    end_date,
) -> pd.DataFrame:
    """
    Compara ventas por vendedor con el período inmediatamente anterior
    de igual duración.
    """
    if (
        "Fecha_dt" not in sales_df.columns
        or sales_df["Fecha_dt"].isna().all()
        or start_date is None
        or end_date is None
    ):
        return pd.DataFrame()

    start_ts = pd.Timestamp(start_date)
    end_ts = pd.Timestamp(end_date) + pd.Timedelta(days=1) - pd.Timedelta(seconds=1)

    period_days = max((pd.Timestamp(end_date) - pd.Timestamp(start_date)).days + 1, 1)
    prev_end = start_ts - pd.Timedelta(seconds=1)
    prev_start = start_ts - pd.Timedelta(days=period_days)

    current = sales_df[
        (sales_df["Fecha_dt"] >= start_ts)
        & (sales_df["Fecha_dt"] <= end_ts)
    ].copy()

    previous = sales_df[
        (sales_df["Fecha_dt"] >= prev_start)
        & (sales_df["Fecha_dt"] <= prev_end)
    ].copy()

    seller_col = "Vendedor" if "Vendedor" in sales_df.columns else None
    if seller_col is None:
        return pd.DataFrame()

    def _agg(frame):
        if frame.empty:
            return pd.DataFrame(columns=["Vendedor", "Venta", "Documentos"])

        frame = ensure_sales_amount_column(frame)
        frame = frame.copy()
        frame["Vendedor"] = (
            frame["Vendedor"].fillna("Sin vendedor").astype(str).str.strip()
        )
        frame["_VentaFirmada"] = frame.apply(
            lambda r: (
                -abs(float(r.get("VentaMonto_num", 0) or 0))
                if r.get("Grupo comercial") == "Nota de crédito"
                else float(r.get("VentaMonto_num", 0) or 0)
            ),
            axis=1,
        )

        number_col = "Numero" if "Numero" in frame.columns else "Vendedor"

        return (
            frame.groupby("Vendedor", as_index=False)
            .agg(
                Venta=("_VentaFirmada", "sum"),
                Documentos=(number_col, "nunique"),
            )
        )

    cur = _agg(current).rename(
        columns={
            "Venta": "Venta actual",
            "Documentos": "Documentos actuales",
        }
    )
    prev = _agg(previous).rename(
        columns={
            "Venta": "Venta anterior",
            "Documentos": "Documentos anteriores",
        }
    )

    result = cur.merge(prev, on="Vendedor", how="outer").fillna(0)

    result["Variación %"] = result.apply(
        lambda r: (
            ((r["Venta actual"] - r["Venta anterior"]) / r["Venta anterior"]) * 100
            if r["Venta anterior"] != 0
            else (100.0 if r["Venta actual"] > 0 else 0.0)
        ),
        axis=1,
    )

    return result




def render_sidebar_html(content: str) -> None:
    """Renderiza HTML dentro del sidebar (no en el panel principal)."""
    clean = dedent(content).strip()
    clean = re.sub(r">\s+<", "><", clean)
    st.sidebar.markdown(clean, unsafe_allow_html=True)


def image_file_to_base64(path: Path) -> str:
    """Convierte una imagen local a base64 para usarla dentro del sidebar."""
    if not path.exists():
        return ""
    return base64.b64encode(path.read_bytes()).decode("utf-8")


LOGO_PATH = BASE_DIR / "assets" / "logo_grupo_maritex.png"
LOGO_BASE64 = image_file_to_base64(LOGO_PATH)


# ============================================================
# ESTILOS V61 · HOJA EXTERNA
# ============================================================
apply_styles(BASE_DIR / "styles.css")


render_html("""
<style>
/* ============================================================
   V39 · VISUAL PREMIUM — STOCK GENERAL
   ============================================================ */
:root{
  --v39-bg:#f6f7f9;
  --v39-card:#ffffff;
  --v39-border:#e2e6ea;
  --v39-text:#111827;
  --v39-muted:#6b7280;
  --v39-purple:#6536f3;
  --v39-purple-soft:#eee9ff;
  --v39-green:#2dbf72;
  --v39-green-soft:#e7f8ef;
  --v39-blue:#3b82f6;
  --v39-blue-soft:#e8f1ff;
  --v39-orange:#ff7a45;
  --v39-orange-soft:#fff0e9;
  --v39-red:#ef4444;
  --v39-red-soft:#fff0f0;
  --v39-yellow:#f5b81b;
}

[data-testid="stAppViewContainer"]{
  background:var(--v39-bg)!important;
}
.block-container{
  max-width:1540px!important;
  padding:2.15rem 2.25rem 3rem!important;
}

/* Sidebar */
[data-testid="stSidebar"]{
  background:linear-gradient(180deg,#090d12 0%,#0b1118 100%)!important;
  border-right:1px solid #1d2630!important;
}
.mtx-logo{
  font-size:42px!important;
  margin-bottom:22px!important;
}
.sidebar-menu-label{
  margin-top:2px!important;
  color:#7f8b98!important;
}
[data-testid="stSidebar"] .stButton button{
  min-height:58px!important;
  border-radius:10px!important;
  background:transparent!important;
  border:1px solid transparent!important;
  padding:0 16px!important;
  font-size:13px!important;
  font-weight:720!important;
}
[data-testid="stSidebar"] .stButton button:hover{
  background:#121922!important;
  border-color:#202b36!important;
}
[data-testid="stSidebar"] .stButton button[kind="primary"]{
  background:linear-gradient(90deg,#20262e 0%,#171d24 100%)!important;
  border-color:#2d3641!important;
  box-shadow:inset 4px 0 0 #dbea00!important;
}
.sidebar-footer-clean{
  background:#111821!important;
  border:1px solid #202a35!important;
  border-radius:10px!important;
  padding:16px!important;
  margin:25px 12px 0!important;
}

/* Header */
.v39-page-head{
  display:flex;
  align-items:flex-start;
  justify-content:space-between;
  gap:20px;
  margin:0 0 18px;
}
.v39-page-title{
  font-size:2.05rem;
  line-height:1.05;
  font-weight:850;
  letter-spacing:-.045em;
  color:#101318;
}
.v39-page-subtitle{
  margin-top:8px;
  color:#697482;
  font-size:.88rem;
}

/* File card */
.v39-file-card{
  display:flex;
  align-items:center;
  gap:14px;
  background:#fff;
  border:1px solid var(--v39-border);
  border-radius:14px;
  padding:15px 18px;
  margin:10px 0 18px;
  box-shadow:0 1px 2px rgba(16,24,40,.025);
}
.v39-file-icon{
  width:42px;height:42px;border-radius:12px;
  display:flex;align-items:center;justify-content:center;
  background:var(--v39-purple-soft);color:var(--v39-purple);
  font-size:20px;font-weight:900;
}
.v39-file-name{font-weight:800;color:#252b33;font-size:.9rem}
.v39-file-meta{font-size:.7rem;color:#818a95;margin-top:3px}

/* KPI */
.v39-kpi-grid{
  display:grid;
  grid-template-columns:repeat(4,minmax(0,1fr));
  gap:16px;
  margin:0 0 18px;
}
.v39-kpi-card{
  min-height:138px;
  background:#fff;
  border:1px solid var(--v39-border);
  border-radius:15px;
  padding:18px 18px 15px;
  box-shadow:0 1px 3px rgba(16,24,40,.035);
}
.v39-kpi-top{
  display:flex;align-items:center;gap:13px;
}
.v39-kpi-icon{
  width:48px;height:48px;border-radius:50%;
  display:flex;align-items:center;justify-content:center;
  font-size:21px;font-weight:900;
}
.v39-icon-purple{background:#ece5ff;color:#6d3cf4}
.v39-icon-green{background:#e3f8ec;color:#25ad65}
.v39-icon-blue{background:#e5efff;color:#3d7feb}
.v39-icon-orange{background:#ffebe4;color:#ff7040}
.v39-kpi-label{font-size:.72rem;color:#687280;font-weight:700}
.v39-kpi-value{font-size:1.65rem;line-height:1.1;font-weight:850;color:#181d24;margin-top:3px}
.v39-kpi-foot{
  margin-top:14px;padding-top:11px;border-top:1px solid #eef0f2;
  color:#9098a2;font-size:.67rem
}

/* attention */
.v39-alert{
  display:flex;align-items:center;justify-content:space-between;gap:18px;
  background:#fffafa;border:1px solid #f3d6d6;border-left:4px solid #ef4444;
  border-radius:12px;padding:15px 18px;margin:0 0 18px;
}
.v39-alert-left{display:flex;align-items:center;gap:14px}
.v39-alert-icon{
  width:40px;height:40px;border-radius:50%;
  background:#ffe7e7;color:#ef4444;
  display:flex;align-items:center;justify-content:center;font-size:21px;
}
.v39-alert-title{font-weight:820;color:#8b2d2d;font-size:.84rem}
.v39-alert-items{display:flex;flex-wrap:wrap;gap:15px;margin-top:7px;color:#7b6870;font-size:.69rem}
.v39-alert-items strong{color:#2a3036}

/* analytic cards */
.v39-analytics{
  display:grid;
  grid-template-columns:1fr 1.2fr;
  gap:16px;
  margin-bottom:18px;
}
.v39-analytics-card{
  background:#fff;border:1px solid var(--v39-border);
  border-radius:15px;padding:17px 18px;
  min-height:275px;
}
.v39-card-title{font-size:.82rem;font-weight:820;color:#293039;margin-bottom:13px}
.v39-card-title small{font-weight:500;color:#89929d}
.v39-donut-wrap{
  display:grid;grid-template-columns:190px 1fr;align-items:center;gap:10px;
  min-height:205px;
}
.v39-donut{
  width:152px;height:152px;border-radius:50%;
  position:relative;margin:auto;
}
.v39-donut::after{
  content:"";position:absolute;inset:25px;border-radius:50%;background:#fff;
}
.v39-donut-center{
  position:absolute;inset:0;z-index:2;
  display:flex;align-items:center;justify-content:center;flex-direction:column;
}
.v39-donut-center strong{font-size:1.22rem;color:#20252c}
.v39-donut-center span{font-size:.65rem;color:#8b949e}
.v39-legend{display:flex;flex-direction:column;gap:10px}
.v39-legend-row{display:flex;align-items:center;justify-content:space-between;gap:10px;font-size:.7rem}
.v39-legend-name{display:flex;align-items:center;gap:8px;color:#68727e}
.v39-dot{width:8px;height:8px;border-radius:50%;display:inline-block}
.v39-legend-row strong{color:#2b3138}

/* Filters */
.v39-filter-card{
  background:#fff;border:1px solid var(--v39-border);
  border-radius:15px;padding:16px 16px 10px;margin-top:18px;
}
.v39-section-line{
  display:flex;align-items:center;justify-content:space-between;
  margin-bottom:4px;
}
.v39-view-label{font-size:.75rem;font-weight:800;color:#303740}

/* Streamlit fields */
div[data-testid="stTextInput"] input,
div[data-baseweb="select"] > div{
  min-height:42px!important;
  border-radius:9px!important;
  border-color:#e1e5ea!important;
  background:#fbfcfd!important;
}
div[data-testid="stFileUploader"]{
  background:#fff!important;
  border:1px dashed #d7dce2!important;
  border-radius:12px!important;
  padding:4px!important;
}
div[data-testid="stDataFrame"]{
  border-radius:12px!important;border-color:#e1e5e9!important;
}
div[data-testid="stDownloadButton"] button{
  background:linear-gradient(90deg,#6d3cf4,#5b28ed)!important;
  border-color:#5b28ed!important;border-radius:9px!important;
}
div[data-testid="stButton"] button{
  border-radius:9px;
}

/* Radio pills in main content */
[data-testid="stAppViewContainer"] div[role="radiogroup"]{
  gap:8px!important;
}
[data-testid="stAppViewContainer"] div[role="radiogroup"] label{
  border:1px solid #dfe4e9;
  background:#fff;
  border-radius:999px;
  padding:6px 11px!important;
}
[data-testid="stAppViewContainer"] div[role="radiogroup"] label:has(input:checked){
  background:#f0ebff;
  border-color:#cfc0ff;
}
[data-testid="stAppViewContainer"] div[role="radiogroup"] label p{
  font-size:.71rem!important;
}

/* hide old generic cards on stock page when new ones are present */
.v39-stock-page + .inventory-kpis{display:none!important}

@media(max-width:1100px){
  .v39-kpi-grid{grid-template-columns:repeat(2,minmax(0,1fr))}
  .v39-analytics{grid-template-columns:1fr}
}

/* ============================================================
   V40 · LOGO GRUPO MARITEX
   ============================================================ */
.grupo-maritex-logo{
    width:100%;
    display:flex;
    align-items:center;
    justify-content:flex-start;
    margin:0 0 30px 0;
    padding:0 6px;
    box-sizing:border-box;
}
.grupo-maritex-logo img{
    width:230px;
    max-width:100%;
    height:auto;
    display:block;
    object-fit:contain;
    border-radius:0;
    background:#000;
}
.grupo-maritex-logo-fallback{
    display:flex;
    align-items:center;
    gap:9px;
    margin:0 6px 30px;
    color:#fff;
    font-family:Arial,sans-serif;
    font-size:23px;
    font-weight:800;
}
.grupo-maritex-logo-fallback .gm-x{
    color:#e52629;
    font-size:31px;
    font-weight:900;
}
.sidebar-shell-top{
    padding-top:30px!important;
}


/* ============================================================
   V41 · TEMPLATE COMPACTO
   ============================================================ */
.template-compact-wrap{margin-top:0!important}
.template-compact-head{margin-bottom:6px!important}
.template-compact-subtitle{
    margin-top:3px!important;
    font-size:.76rem!important;
    color:#7b8590!important
}
.template-info-strip{
    margin:5px 0 9px;
    padding:8px 11px;
    background:#eef5ff;
    border:1px solid #dce9fb;
    border-radius:8px;
    color:#1f5f9e;
    font-size:10.5px;
    line-height:1.35
}
.template-card-title{
    font-size:1rem!important;
    font-weight:800!important;
    margin-bottom:2px!important;
    color:#1d232b!important
}
.template-card-meta{
    font-size:.64rem!important;
    color:#8c96a1!important;
    margin:2px 0 7px!important
}
div[data-testid="stFileUploader"]{margin-bottom:0!important}
div[data-testid="stFileUploaderDropzone"]{
    min-height:50px!important;
    padding:6px 9px!important;
    border-radius:8px!important
}
div[data-testid="stFileUploaderDropzone"] button{
    min-height:32px!important;
    padding:3px 12px!important;
    font-size:10.5px!important
}
div[data-testid="stFileUploaderDropzone"] small{
    font-size:9px!important
}
div[data-testid="stAlert"]{
    padding:8px 11px!important;
    border-radius:8px!important
}
div[data-testid="stAlert"] p{
    font-size:10.5px!important;
    margin:0!important
}
.template-ops-bar{
    margin-top:9px;
    padding:10px 13px;
    background:#fff;
    border:1px solid #e1e5e9;
    border-radius:9px;
    display:flex;
    align-items:center;
    gap:22px;
    flex-wrap:wrap;
    font-size:10.5px
}
.template-ops-item{
    display:flex;
    align-items:center;
    gap:5px
}
.template-ops-label{color:#949ca6}
.template-ops-value{
    font-weight:800;
    color:#252b33
}
.template-ops-status{
    margin-left:auto;
    color:#2ba96b;
    font-weight:800;
    white-space:nowrap
}


/* ============================================================
   V42 · MÉTRICAS PREMIUM
   ============================================================ */
.metrics-head{
    display:flex;
    align-items:flex-start;
    justify-content:space-between;
    gap:16px;
    margin-bottom:12px;
}
.metrics-title{
    font-size:2rem;
    line-height:1.05;
    font-weight:850;
    letter-spacing:-.04em;
    color:#11151a;
}
.metrics-subtitle{
    margin-top:7px;
    color:#737e89;
    font-size:.84rem;
}
.metrics-source{
    background:#fff;
    border:1px solid #e1e5e9;
    border-radius:11px;
    padding:10px 13px;
    margin:8px 0 14px;
    display:flex;
    align-items:center;
    justify-content:space-between;
    gap:14px;
    color:#69737d;
    font-size:.7rem;
}
.metrics-source strong{color:#252b33}
.metrics-dot{
    width:8px;height:8px;border-radius:50%;
    background:#2dbf72;display:inline-block;margin-right:6px
}
.metrics-kpi-grid{
    display:grid;
    grid-template-columns:repeat(4,minmax(0,1fr));
    gap:13px;
    margin-bottom:14px
}
.metrics-kpi{
    background:#fff;
    border:1px solid #e1e5e9;
    border-radius:14px;
    padding:15px 16px;
    min-height:112px;
    box-shadow:0 1px 3px rgba(16,24,40,.03)
}
.metrics-kpi-label{
    color:#707a85;font-size:.68rem;font-weight:720
}
.metrics-kpi-value{
    color:#151a20;font-size:1.55rem;font-weight:850;
    margin-top:5px;letter-spacing:-.025em
}
.metrics-kpi-note{
    color:#929aa4;font-size:.63rem;margin-top:8px
}
.metrics-card{
    background:#fff;
    border:1px solid #e1e5e9;
    border-radius:14px;
    padding:15px 16px;
    min-height:300px
}
.metrics-card-title{
    color:#252b33;font-size:.78rem;font-weight:820;margin-bottom:8px
}
.metrics-section-title{
    font-size:1rem;font-weight:850;color:#20262d;
    margin:16px 0 9px
}
.metrics-alert{
    border:1px solid #f2ded0;
    background:#fffaf6;
    border-left:4px solid #ff8a4c;
    border-radius:10px;
    padding:10px 13px;
    margin:10px 0 14px;
    font-size:.7rem;
    color:#7a5b49
}
.metrics-inventory-grid{
    display:grid;
    grid-template-columns:repeat(5,minmax(0,1fr));
    gap:11px;
    margin-bottom:14px
}
.metrics-mini{
    background:#fff;border:1px solid #e1e5e9;border-radius:12px;
    padding:12px 13px
}
.metrics-mini-label{font-size:.64rem;color:#7d8792;font-weight:700}
.metrics-mini-value{font-size:1.25rem;color:#20262d;font-weight:850;margin-top:4px}
.metrics-mini-note{font-size:.59rem;color:#949da7;margin-top:5px}
@media(max-width:1150px){
    .metrics-kpi-grid{grid-template-columns:repeat(2,minmax(0,1fr))}
    .metrics-inventory-grid{grid-template-columns:repeat(2,minmax(0,1fr))}
}


/* ============================================================
   V43 · VENDEDORES
   ============================================================ */
.seller-head{
    display:flex;
    align-items:flex-start;
    justify-content:space-between;
    margin-bottom:12px
}
.seller-title{
    font-size:2rem;
    font-weight:850;
    letter-spacing:-.04em;
    color:#11151a
}
.seller-subtitle{
    margin-top:7px;
    color:#737e89;
    font-size:.84rem
}
.seller-rule{
    display:inline-flex;
    align-items:center;
    gap:7px;
    background:#edf8f2;
    border:1px solid #d5eee0;
    border-radius:999px;
    padding:6px 10px;
    color:#287b50;
    font-size:.66rem;
    font-weight:750
}
.seller-grid{
    display:grid;
    grid-template-columns:repeat(5,minmax(0,1fr));
    gap:13px;
    margin:12px 0 14px
}
.seller-kpi{
    background:#fff;
    border:1px solid #e1e5e9;
    border-radius:14px;
    padding:15px 16px;
    min-height:112px
}
.seller-kpi-label{
    color:#707a85;
    font-size:.67rem;
    font-weight:720
}
.seller-kpi-value{
    color:#151a20;
    font-size:1.45rem;
    font-weight:850;
    margin-top:5px;
    letter-spacing:-.025em
}
.seller-kpi-note{
    color:#929aa4;
    font-size:.61rem;
    margin-top:7px
}
.seller-card{
    background:#fff;
    border:1px solid #e1e5e9;
    border-radius:14px;
    padding:15px 16px;
    min-height:320px
}
.seller-card-title{
    color:#252b33;
    font-size:.78rem;
    font-weight:820;
    margin-bottom:9px
}
.seller-insights{
    display:grid;
    grid-template-columns:repeat(3,minmax(0,1fr));
    gap:11px;
    margin:13px 0
}
.seller-insight{
    border:1px solid #e1e5e9;
    background:#fff;
    border-radius:12px;
    padding:12px 13px
}
.seller-insight-title{
    font-size:.67rem;
    font-weight:800;
    color:#303740
}
.seller-insight-text{
    margin-top:6px;
    font-size:.64rem;
    color:#79838e;
    line-height:1.45
}
.seller-section-title{
    font-size:1rem;
    font-weight:850;
    color:#20262d;
    margin:16px 0 9px
}
@media(max-width:1150px){
    .seller-grid{grid-template-columns:repeat(2,minmax(0,1fr))}
    .seller-insights{grid-template-columns:1fr}
}


/* ============================================================
   V44 · BÚSQUEDA CLIENTE → DOCUMENTOS
   ============================================================ */
.client-search-summary{
    background:#fff;
    border:1px solid #e1e5e9;
    border-left:4px solid #6536f3;
    border-radius:11px;
    padding:11px 13px;
    margin:10px 0 12px;
    display:flex;
    align-items:center;
    justify-content:space-between;
    gap:14px;
    flex-wrap:wrap;
}
.client-search-title{
    font-size:.72rem;
    font-weight:820;
    color:#272e36;
}
.client-search-meta{
    margin-top:4px;
    color:#818b96;
    font-size:.63rem;
}
.client-search-kpis{
    display:flex;
    gap:18px;
    flex-wrap:wrap;
}
.client-search-kpi{
    text-align:right;
}
.client-search-kpi span{
    display:block;
    font-size:.58rem;
    color:#929ba5;
}
.client-search-kpi strong{
    display:block;
    margin-top:2px;
    font-size:.78rem;
    color:#252b33;
}


/* ============================================================
   V45 · VENTAS + NC + IVA
   ============================================================ */
.sales-final-grid{
    display:grid;
    grid-template-columns:repeat(4,minmax(0,1fr));
    gap:12px;
    margin:12px 0 15px;
}
.sales-final-card{
    background:#fff;
    border:1px solid #e1e5e9;
    border-radius:13px;
    padding:14px 15px;
}
.sales-final-card.main{
    border-color:#cfc2ff;
    background:#fbf9ff;
}
.sales-final-label{
    font-size:.64rem;
    color:#7d8792;
    font-weight:720;
}
.sales-final-value{
    margin-top:5px;
    font-size:1.35rem;
    font-weight:850;
    color:#171c22;
}
.sales-final-note{
    margin-top:6px;
    font-size:.59rem;
    color:#919aa4;
}
.credit-value{color:#c94a4a!important}
.sales-rule-strip{
    margin:8px 0 12px;
    padding:9px 12px;
    border:1px solid #e2e6ea;
    border-radius:9px;
    background:#fff;
    font-size:.65rem;
    color:#707a84;
}
@media(max-width:1100px){
    .sales-final-grid{grid-template-columns:repeat(2,minmax(0,1fr))}
}


/* ============================================================
   V46 · KPIs COMERCIALES CLICKEABLES
   ============================================================ */
.sales-kpi-button-wrap .stButton > button{
    width:100% !important;
    min-height:108px !important;
    height:108px !important;
    display:flex !important;
    align-items:flex-start !important;
    justify-content:flex-start !important;
    text-align:left !important;
    padding:14px 15px !important;
    border-radius:13px !important;
    border:1px solid #e1e5e9 !important;
    background:#fff !important;
    color:#171c22 !important;
    box-shadow:none !important;
    white-space:pre-line !important;
    line-height:1.35 !important;
}
.sales-kpi-button-wrap .stButton > button:hover{
    border-color:#b9a6ff !important;
    background:#fbf9ff !important;
    transform:translateY(-1px);
}
.sales-kpi-button-wrap.main .stButton > button{
    border-color:#cfc2ff !important;
    background:#fbf9ff !important;
}
.sales-kpi-button-wrap.credit .stButton > button{
    color:#c94a4a !important;
}
.sales-detail-box{
    margin:10px 0 14px;
    padding:12px 14px;
    border:1px solid #e2e6ea;
    border-radius:11px;
    background:#fff;
}
.sales-detail-title{
    font-size:.76rem;
    font-weight:850;
    color:#252b33;
    margin-bottom:4px;
}
.sales-detail-subtitle{
    font-size:.63rem;
    color:#818b96;
}


/* ============================================================
   V47 · META VENDEDORES
   ============================================================ */
.seller-goal-wrap{
    margin:12px 0 15px;
    background:#fff;
    border:1px solid #e1e5e9;
    border-radius:13px;
    padding:14px 15px;
}
.seller-goal-head{
    display:flex;
    align-items:flex-start;
    justify-content:space-between;
    gap:16px;
    margin-bottom:10px;
}
.seller-goal-title{
    font-size:.72rem;
    font-weight:850;
    color:#242b33;
}
.seller-goal-sub{
    margin-top:4px;
    font-size:.62rem;
    color:#858f99;
}
.seller-goal-values{
    display:flex;
    gap:22px;
    flex-wrap:wrap;
}
.seller-goal-item{
    min-width:120px;
}
.seller-goal-item span{
    display:block;
    font-size:.58rem;
    color:#929ba5;
}
.seller-goal-item strong{
    display:block;
    margin-top:3px;
    font-size:.84rem;
    color:#20262d;
}
.seller-goal-bar{
    width:100%;
    height:12px;
    background:#eceff3;
    border-radius:999px;
    overflow:hidden;
}
.seller-goal-fill{
    height:100%;
    border-radius:999px;
    background:linear-gradient(90deg,#6d3cf4,#8b5cf6);
}
.seller-goal-foot{
    display:flex;
    justify-content:space-between;
    gap:12px;
    margin-top:7px;
    font-size:.61rem;
    color:#858f99;
}
.seller-goal-foot strong{
    color:#252b33;
}

</style>
""")


# ============================================================
# CONFIGURACIÓN DE MARKETPLACES
# (Antes había un if/elif duplicado por cada marketplace.
#  Ahora agregar uno nuevo es solo sumar una entrada al dict.)
# ============================================================

@dataclass
class TemplateConfig:
    sheet_name: str
    sku_column: str
    destination_stock_column: str
    first_data_row: int
    current_stock_column: Optional[str] = None
    product_column: Optional[str] = None
    info_message: str = ""


MARKETPLACE_CONFIGS = {
    "Paris Marketplace": TemplateConfig(
        sheet_name="stock",
        sku_column="B",
        destination_stock_column="H",
        first_data_row=2,
        current_stock_column="G",
        product_column="C",
        info_message="Paris: sku_seller (B) → nuevo_stock (H)",
    ),
    "Mercado Libre": TemplateConfig(
        sheet_name="Publicaciones",
        # PLANTILLA MELI.xlsx:
        # E = SKU de Mercado Libre (este es el código que cruza con Producto en Flexline)
        # H = QUANTITY / Stock en tu depósito
        sku_column="E",
        destination_stock_column="H",
        first_data_row=6,
        current_stock_column="H",
        # G muestra la variante (talla/color) y evita mostrar fórmulas de TITLE.
        product_column="G",
        info_message=(
            "Mercado Libre: SKU (E) ↔ Producto Flexline → QUANTITY (H). "
            "Stock tomado exclusivamente desde Casa Matriz."
        ),
    ),
}

# Rango razonable de filas vacías consecutivas antes de asumir que
# ya no hay más datos. Evita recorrer miles de filas con formato
# pero sin contenido, que openpyxl a veces incluye en max_row.
MAX_CONSECUTIVE_EMPTY_ROWS = 500


# ============================================================
# UTILIDADES DE ERROR
# ============================================================

def show_error(exc: Exception, context: str = "") -> None:
    """Muestra un error legible al usuario y deja el traceback en un expander."""
    prefix = f"{context}: " if context else ""
    st.error(f"❌ {prefix}{exc}")
    with st.expander("Detalles técnicos"):
        st.code(traceback.format_exc())


# ============================================================
# FUNCIONES DE NORMALIZACIÓN
# ============================================================

def normalize_code(value) -> str:
    """Normaliza un código/SKU a un formato comparable (alfanumérico, mayúsculas).

    Cubre tres casos que antes no se resolvían bien:
    - Floats que vienen de pandas/Excel: 1305120.0 -> 1305120
    - Notación científica de Excel para números largos: 1.23457E+11 -> 123457000000
    - Separadores no alfanuméricos (guiones, espacios): 1305120-5 -> 13051205
    """
    if pd.isna(value):
        return ""

    text = str(value).strip()
    if not text:
        return ""

    if re.fullmatch(r"-?\d+(\.\d+)?[Ee][+-]?\d+", text):
        try:
            as_float = float(text)
            if as_float.is_integer():
                text = str(int(as_float))
        except (ValueError, OverflowError):
            pass

    if re.fullmatch(r"-?\d+\.\d+", text):
        try:
            as_float = float(text)
            if as_float.is_integer():
                text = str(int(as_float))
        except ValueError:
            pass

    text = re.sub(r"[^A-Za-z0-9]", "", text)

    return text.upper()


def parse_number(value) -> float:
    """Convierte un valor de stock a float, soportando formato CL/ES y US.

    Bug corregido: antes "1.500" (mil quinientos, separador de miles chileno)
    se interpretaba como 1.5 porque solo se manejaba la coma como decimal.
    Ahora se detecta si el punto es separador de miles (grupos de 3 dígitos)
    o decimal, y si hay coma Y punto juntos se usa el que aparece último
    como separador decimal.
    """
    if pd.isna(value):
        return 0.0

    text = str(value).strip()
    if not text:
        return 0.0

    text = re.sub(r"[^\d,.\-]", "", text)
    if not text:
        return 0.0

    has_comma = "," in text
    has_dot = "." in text

    if has_comma and has_dot:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif has_comma:
        text = text.replace(".", "").replace(",", ".")
    elif has_dot and re.fullmatch(r"-?\d{1,3}(\.\d{3})+", text):
        text = text.replace(".", "")

    try:
        return float(text)
    except (ValueError, TypeError):
        return 0.0


def excel_column_number(column_letters: str) -> int:
    """Convierte letras de columna Excel (A, B, ..., AA) a número de columna.

    Antes una columna vacía o inválida producía column=0 y un IndexError
    críptico de openpyxl al procesar. Ahora se valida explícitamente.
    """
    text = str(column_letters).strip().upper()

    if not text or not text.isalpha():
        raise ValueError(
            f"Columna Excel inválida: '{column_letters}'. "
            "Debe contener solo letras (A, B, AA, etc.)."
        )

    value = 0
    for char in text:
        value = value * 26 + ord(char) - ord("A") + 1

    return value



def detect_column(df: pd.DataFrame, candidates: list[str]) -> Optional[str]:
    """Busca una columna por coincidencia exacta y luego normalizada."""
    if df is None or df.empty:
        return None

    columns = list(df.columns)

    # Coincidencia exacta
    for candidate in candidates:
        if candidate in columns:
            return candidate

    # Coincidencia ignorando espacios, tildes simples/case y símbolos
    def norm(value: str) -> str:
        value = str(value).strip().lower()
        value = (
            value.replace("á", "a")
            .replace("é", "e")
            .replace("í", "i")
            .replace("ó", "o")
            .replace("ú", "u")
            .replace("ñ", "n")
        )
        return re.sub(r"[^a-z0-9]", "", value)

    normalized = {norm(col): col for col in columns}

    for candidate in candidates:
        key = norm(candidate)
        if key in normalized:
            return normalized[key]

    return None


def build_stock_general_view(
    df: pd.DataFrame,
    code_col: str,
    name_col: Optional[str],
    warehouse_col: Optional[str],
    stock_col: str,
    price_list_col: Optional[str],
    price_col: Optional[str],
) -> pd.DataFrame:
    """Construye la vista estándar de Stock General sin alterar el archivo ERP."""
    view = pd.DataFrame(index=df.index)

    view["Código"] = df[code_col].fillna("").astype(str).str.strip()

    if name_col:
        view["Producto"] = df[name_col].fillna("").astype(str).str.strip()
    else:
        view["Producto"] = ""

    if warehouse_col:
        view["Bodega"] = df[warehouse_col].fillna("").astype(str).str.strip()
    else:
        view["Bodega"] = ""

    view["Stock"] = df[stock_col].apply(parse_number).round().astype(int)

    if price_list_col:
        view["Lista de precio"] = df[price_list_col].fillna("").astype(str).str.strip()
    else:
        view["Lista de precio"] = ""

    if price_col:
        view["Precio"] = df[price_col].apply(parse_number).round().astype(int)
    else:
        view["Precio"] = pd.Series([pd.NA] * len(df), dtype="Int64")

    # Eliminar filas sin código y ordenar para facilitar lectura
    view = view[view["Código"] != ""].copy()

    sort_cols = [col for col in ["Código", "Bodega", "Lista de precio"] if col in view.columns]
    if sort_cols:
        view = view.sort_values(sort_cols, kind="stable")

    return view.reset_index(drop=True)



def build_inventory_source(df: pd.DataFrame) -> pd.DataFrame:
    """
    Normaliza el archivo de stock general exportado desde Flexline.
    Diseñado para columnas como:
    Producto, Descripción, Bodega, Stock, StockDisponible,
    StockPorLlegar, StockPorDespachar, DescFamilia, DescSubFamilia,
    DescTipo, Precio Venta.
    """
    required = ["Producto", "Descripción", "Bodega", "StockDisponible"]
    missing = [col for col in required if col not in df.columns]
    if missing:
        raise ValueError(
            "El archivo de Stock General no contiene las columnas requeridas: "
            + ", ".join(missing)
        )

    out = pd.DataFrame(index=df.index)

    out["Código"] = df["Producto"].fillna("").astype(str).str.strip()
    out["Producto"] = df["Descripción"].fillna("").astype(str).str.strip()
    out["Bodega"] = df["Bodega"].fillna("").astype(str).str.strip()

    for source, target in [
        ("Stock", "Stock físico"),
        ("StockDisponible", "Disponible"),
        ("StockPorLlegar", "Por llegar"),
        ("StockPorDespachar", "Por despachar"),
        ("Precio Venta", "Precio"),
    ]:
        if source in df.columns:
            out[target] = df[source].apply(parse_number).round().astype(int)
        else:
            out[target] = 0

    out["Familia"] = (
        df["DescFamilia"].fillna("").astype(str).str.strip()
        if "DescFamilia" in df.columns else ""
    )
    out["Subfamilia"] = (
        df["DescSubFamilia"].fillna("").astype(str).str.strip()
        if "DescSubFamilia" in df.columns else ""
    )
    out["Tipo"] = (
        df["DescTipo"].fillna("").astype(str).str.strip()
        if "DescTipo" in df.columns else ""
    )

    out = out[out["Código"] != ""].copy()
    return out.reset_index(drop=True)


def consolidate_inventory_by_product(df: pd.DataFrame) -> pd.DataFrame:
    """Consolida todas las bodegas por código de producto."""
    if df.empty:
        return df.copy()

    numeric_cols = [
        "Stock físico",
        "Disponible",
        "Por llegar",
        "Por despachar",
    ]

    grouped = (
        df.groupby("Código", as_index=False)
        .agg(
            {
                "Producto": "first",
                "Familia": "first",
                "Subfamilia": "first",
                "Tipo": "first",
                "Precio": "max",
                **{col: "sum" for col in numeric_cols},
                "Bodega": lambda s: s.replace("", pd.NA).dropna().nunique(),
            }
        )
        .rename(columns={"Bodega": "Bodegas"})
    )

    return grouped


def inventory_status_mask(df: pd.DataFrame, status: str) -> pd.Series:
    """Filtro operacional para Stock General."""
    available = pd.to_numeric(df["Disponible"], errors="coerce").fillna(0)
    incoming = pd.to_numeric(df["Por llegar"], errors="coerce").fillna(0)
    outgoing = pd.to_numeric(df["Por despachar"], errors="coerce").fillna(0)

    if status == "Con stock":
        return available > 0
    if status == "Stock 0":
        return available == 0
    if status == "Stock negativo":
        return available < 0
    if status == "Por llegar":
        return incoming > 0
    if status == "Por despachar":
        return outgoing > 0

    return pd.Series(True, index=df.index)



def classify_inventory_status(row) -> str:
    """
    Semáforo operativo.
    Usa Disponible / Por llegar / Por despachar.
    """
    available = parse_number(row.get("Disponible", 0))
    incoming = parse_number(row.get("Por llegar", 0))
    outgoing = parse_number(row.get("Por despachar", 0))

    if available < 0:
        return "🔴 Negativo"
    if available == 0 and incoming > 0:
        return "🔵 Por llegar"
    if available == 0:
        return "🔴 Sin stock"
    if available > 0 and outgoing > available:
        return "🟠 Riesgo despacho"
    if available <= 5:
        return "🟡 Stock bajo"
    return "🟢 Disponible"


def add_inventory_status(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out["Estado"] = out.apply(classify_inventory_status, axis=1)
    return out


def dataframe_to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Stock") -> bytes:
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    buffer.seek(0)
    return buffer.getvalue()


# ============================================================
# LECTURA Y CONSOLIDACIÓN FLEXLINE (con caché)
# ============================================================
#
# IMPORTANTE: Streamlit re-ejecuta todo el script en cada interacción
# (por ejemplo, cada letra tecleada en "SKU de prueba"). Sin caché, eso
# significaba releer y reconsolidar el archivo Flexline completo en
# cada tecla presionada. @st.cache_data evita ese trabajo repetido:
# solo se vuelve a ejecutar si cambian los bytes del archivo o los
# parámetros de la función.

@st.cache_data(show_spinner="Leyendo archivo Flexline...")
def read_flexline(file_bytes: bytes, file_name: str) -> pd.DataFrame:
    name = file_name.lower()

    if name.endswith((".xlsx", ".xls")):
        return read_excel_flexline(file_bytes, file_name)

    if name.endswith(".csv"):
        last_error = None

        # Primero intenta encontrar una combinación separador/encoding
        # que además contenga la columna esperada, para evitar quedarse
        # con una lectura "exitosa" pero mal delimitada.
        for encoding in ("utf-8-sig", "cp1252", "latin1"):
            try:
                df = pd.read_csv(
                    io.BytesIO(file_bytes),
                    sep=None,
                    engine="python",
                    encoding=encoding,
                    dtype=str,
                )
                if len(df.columns) > 1 and "Producto" in df.columns:
                    return df
            except Exception as exc:
                last_error = exc

        for encoding in ("utf-8-sig", "cp1252", "latin1"):
            try:
                df = pd.read_csv(
                    io.BytesIO(file_bytes),
                    sep=None,
                    engine="python",
                    encoding=encoding,
                    dtype=str,
                )
                if len(df.columns) > 1:
                    return df
            except Exception as exc:
                last_error = exc

        raise ValueError(f"No fue posible leer el CSV de Flexline: {last_error}")

    raise ValueError("El archivo Flexline debe ser CSV, XLS o XLSX.")


@st.cache_data(show_spinner="Consolidando stock Flexline...")
def consolidate_flexline(
    df: pd.DataFrame,
    code_column: str = "Producto",
    stock_column: str = "StockDisponible",
    warehouse_column: Optional[str] = None,
    warehouses: Optional[tuple] = None,
):
    if code_column not in df.columns:
        raise ValueError(f"No existe la columna '{code_column}' en Flexline.")

    if stock_column not in df.columns:
        raise ValueError(f"No existe la columna '{stock_column}' en Flexline.")

    work = df.copy()

    if warehouse_column and warehouse_column in work.columns and warehouses is not None:
        work = work[work[warehouse_column].astype(str).isin(warehouses)].copy()

    work["Código ERP original"] = work[code_column].fillna("").astype(str).str.strip()
    work["Código cruce"] = work[code_column].apply(normalize_code)
    work["StockDisponible ERP"] = work[stock_column].apply(parse_number)

    work = work[work["Código cruce"] != ""].copy()

    grouped_stock = work.groupby("Código cruce", as_index=False)["StockDisponible ERP"].sum()

    grouped_codes = (
        work.groupby("Código cruce")["Código ERP original"]
        .apply(lambda values: " | ".join(sorted({str(v).strip() for v in values if str(v).strip()})))
        .reset_index()
    )

    grouped = grouped_stock.merge(grouped_codes, on="Código cruce", how="left")

    stock_lookup = dict(zip(grouped["Código cruce"], grouped["StockDisponible ERP"]))
    original_lookup = dict(zip(grouped["Código cruce"], grouped["Código ERP original"]))

    description_candidates = [
        "Descripcion",
        "Descripción",
        "DescripcionProducto",
        "Descripción Producto",
        "NombreProducto",
        "Nombre Producto",
        "Nombre",
        "Detalle",
        "Glosa",
    ]

    description_column = next((col for col in description_candidates if col in work.columns), None)

    product_name_lookup = {}

    if description_column:
        product_names = work[["Código cruce", description_column]].copy()
        product_names[description_column] = (
            product_names[description_column].fillna("").astype(str).str.strip()
        )
        product_names = product_names[product_names[description_column] != ""]

        if not product_names.empty:
            product_name_lookup = (
                product_names.drop_duplicates(subset=["Código cruce"], keep="first")
                .set_index("Código cruce")[description_column]
                .to_dict()
            )

    return grouped, stock_lookup, original_lookup, product_name_lookup, description_column


# ============================================================
# VALIDACIÓN DE PLANTILLAS MARKETPLACE
# ============================================================

def validate_marketplace_template(wb, marketplace: str) -> None:
    """Valida la estructura de PLANTILLA MELI.xlsx."""
    if marketplace != "Mercado Libre":
        return

    if "Publicaciones" not in wb.sheetnames:
        raise ValueError(
            "La plantilla de Mercado Libre debe contener la hoja 'Publicaciones'."
        )

    ws = wb["Publicaciones"]

    expected_headers = {
        "A": "FAMILY_ID",
        "B": "ITEM_ID",
        "C": "PRODUCT_NUMBER",
        "D": "VARIATION_ID",
        "E": "SKU",
        "F": "TITLE",
        "G": "VARIATIONS",
        "H": "QUANTITY",
        "I": "DESCRIPTION",
        "J": "CATEGORY",
    }

    header_values = next(
        ws.iter_rows(
            min_row=1,
            max_row=1,
            min_col=1,
            max_col=10,
            values_only=True,
        ),
        tuple(),
    )

    invalid = []
    for index, (col, expected) in enumerate(expected_headers.items()):
        current = header_values[index] if index < len(header_values) else None
        current = "" if current is None else str(current).strip()
        if current != expected:
            invalid.append(
                f"{col}1: esperado '{expected}', recibido '{current}'"
            )

    if invalid:
        raise ValueError(
            "La estructura del archivo no corresponde a PLANTILLA MELI.xlsx. "
            + " | ".join(invalid)
        )


def get_template_codes(
    template_bytes: bytes,
    sheet_name: str,
    sku_column: str,
    first_data_row: int,
) -> set[str]:
    """Lee los SKU de la plantilla marketplace de forma eficiente."""
    wb = load_workbook(
        io.BytesIO(template_bytes),
        read_only=True,
        data_only=False,
    )

    try:
        if sheet_name not in wb.sheetnames:
            return set()

        ws = wb[sheet_name]
        sku_col = excel_column_number(sku_column)
        codes = set()

        for row in ws.iter_rows(
            min_row=int(first_data_row),
            min_col=sku_col,
            max_col=sku_col,
            values_only=True,
        ):
            value = row[0] if row else None
            code = normalize_code(value)
            if code:
                codes.add(code)

        return codes
    finally:
        wb.close()



@st.cache_data(show_spinner=False)
def prepare_marketplace_erp(erp_bytes: bytes, erp_name: str):
    erp = read_flexline(erp_bytes, erp_name)

    if "Bodega" not in erp.columns:
        raise ValueError(
            "El archivo Flexline debe contener la columna 'Bodega'. "
            "Marketplace utiliza exclusivamente Casa Matriz."
        )

    def _normalize_warehouse_name(value: str) -> str:
        value = str(value).strip().lower()
        value = (
            value.replace("á", "a")
            .replace("é", "e")
            .replace("í", "i")
            .replace("ó", "o")
            .replace("ú", "u")
        )
        return re.sub(r"[^a-z0-9]", "", value)

    warehouse_values = sorted(
        erp["Bodega"].dropna().astype(str).str.strip().unique().tolist()
    )
    casa_matriz_candidates = tuple(
        value
        for value in warehouse_values
        if "casamatriz" in _normalize_warehouse_name(value)
    )

    if not casa_matriz_candidates:
        raise ValueError("No se encontró Casa Matriz en el archivo Flexline.")

    (
        grouped,
        stock_lookup,
        original_lookup,
        product_name_lookup,
        description_column,
    ) = consolidate_flexline(
        erp,
        code_column="Producto",
        stock_column="StockDisponible",
        warehouse_column="Bodega",
        warehouses=casa_matriz_candidates,
    )

    return (
        erp,
        casa_matriz_candidates,
        grouped,
        stock_lookup,
        original_lookup,
        product_name_lookup,
        description_column,
    )


@st.cache_data(show_spinner=False)
def generate_saved_marketplace_output(
    marketplace_name: str,
    template_bytes: bytes,
    erp_bytes: bytes,
    erp_name: str,
    reserve_stock: int,
    max_stock_value,
):
    (
        erp,
        selected_warehouses,
        erp_grouped,
        stock_lookup,
        original_lookup,
        product_name_lookup,
        description_column,
    ) = prepare_marketplace_erp(erp_bytes, erp_name)

    config = MARKETPLACE_CONFIGS[marketplace_name]

    output_bytes, processed_result = process_template(
        template_bytes=template_bytes,
        sheet_name=config.sheet_name,
        sku_column=config.sku_column,
        destination_stock_column=config.destination_stock_column,
        first_data_row=config.first_data_row,
        stock_lookup=stock_lookup,
        original_lookup=original_lookup,
        current_stock_column=config.current_stock_column,
        product_column=config.product_column,
        reserve_stock=reserve_stock,
        max_stock=max_stock_value,
        marketplace=marketplace_name,
    )

    return output_bytes, processed_result, selected_warehouses


# ============================================================
# PROCESAMIENTO DE PLANTILLA
# ============================================================

def process_template(
    template_bytes: bytes,
    sheet_name: str,
    sku_column: str,
    destination_stock_column: str,
    first_data_row: int,
    stock_lookup: dict,
    original_lookup: dict,
    current_stock_column: Optional[str] = None,
    product_column: Optional[str] = None,
    reserve_stock: float = 0,
    max_stock: Optional[float] = None,
    marketplace: Optional[str] = None,
):
    wb = load_workbook(io.BytesIO(template_bytes))

    if marketplace:
        validate_marketplace_template(wb, marketplace)

    if sheet_name not in wb.sheetnames:
        raise ValueError(f"No existe la hoja '{sheet_name}'.")

    ws = wb[sheet_name]

    sku_col = excel_column_number(sku_column)
    stock_col = excel_column_number(destination_stock_column)
    current_col = excel_column_number(current_stock_column) if current_stock_column else None
    product_col = excel_column_number(product_column) if product_column else None

    records = []
    consecutive_empty = 0

    for row_number in range(int(first_data_row), ws.max_row + 1):
        sku_original = ws.cell(row=row_number, column=sku_col).value
        code = normalize_code(sku_original)

        if not code:
            # IMPORTANTE para Mercado Libre:
            # las filas padre de una publicación no tienen PRODUCT_NUMBER.
            # No se debe escribir 0 ni modificar QUANTITY en esas filas.
            consecutive_empty += 1
            if consecutive_empty >= MAX_CONSECUTIVE_EMPTY_ROWS:
                break
            continue

        consecutive_empty = 0

        erp_stock = stock_lookup.get(code)
        erp_code = original_lookup.get(code)

        current_stock = ws.cell(row=row_number, column=current_col).value if current_col else None
        product = ws.cell(row=row_number, column=product_col).value if product_col else None

        if erp_stock is None:
            new_stock = 0
            status = "🔴 No encontrado"
        else:
            new_stock = max(0, float(erp_stock) - float(reserve_stock))

            if max_stock is not None:
                new_stock = min(new_stock, float(max_stock))

            new_stock = int(new_stock)
            status = "🟡 Stock 0" if new_stock == 0 else "🟢 Encontrado"

        ws.cell(row=row_number, column=stock_col).value = new_stock

        current_numeric = parse_number(current_stock) if current_stock is not None else None
        current_numeric = int(round(current_numeric)) if current_numeric is not None else None
        difference = int(new_stock - current_numeric) if current_numeric is not None else None

        records.append(
            {
                "SKU Marketplace": sku_original,
                "Código cruce": code,
                "Código encontrado ERP": erp_code,
                "Producto": product,
                "Stock actual": current_numeric,
                "StockDisponible ERP": int(round(erp_stock)) if erp_stock is not None else None,
                "Nuevo stock": int(new_stock),
                "Diferencia": difference,
                "Estado": status,
            }
        )

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer.getvalue(), pd.DataFrame(records)



# ============================================================
# ESTILOS V61 · CARGADOS DESDE styles.css
# ============================================================


# ============================================================
# SESSION STATE
# ============================================================

defaults = {
    "result": None,
    "output_file": None,
    "processed_filename": None,
    "last_processed": None,
    "processing_message": None,
    "duplicate_count": 0,
    # Stock ERP compartido entre Stock General y Marketplaces.
    # Guardamos bytes/nombre fuera del estado del file_uploader para que
    # el archivo siga disponible al cambiar de pestaña.
    "shared_stock_bytes": None,
    "shared_stock_name": None,
    "shared_stock_loaded_at": None,
    "active_page": "stock_general",
    "marketplace_value": "Paris Marketplace",
    "reserve_stock_value": 0,
    "use_max_stock_value": False,
    "max_stock_value": 50,
    "shared_template_bytes": None,
    "shared_template_name": None,
    "shared_template_marketplace": None,
    "test_sku_value": "",
}

for key, value in defaults.items():
    if key not in st.session_state:
        st.session_state[key] = value


# ============================================================
# SIDEBAR V30 · NAVEGACIÓN PERSISTENTE
# ============================================================
valid_pages = {"stock_general", "marketplace", "metricas", "vendedores", "ventas", "configuracion"}

# Navegación persistente: usar botones de Streamlit evita recargar el navegador
# y conserva la misma sesión (archivos, filtros, resultados y configuraciones).
query_page = st.query_params.get("page")
if query_page in valid_pages and st.session_state.get("_query_page_initialized") is None:
    st.session_state.active_page = query_page
    st.session_state._query_page_initialized = True

page = st.session_state.get("active_page", "stock_general")
if page not in valid_pages:
    page = "stock_general"
    st.session_state.active_page = page

def change_page(target: str) -> None:
    st.session_state.active_page = target
    st.query_params["page"] = target

def nav_button(label: str, target: str, tool: bool = False) -> None:
    """
    Navegación lateral V51.
    - Sin iconos.
    - Botones largos tipo módulo administrativo.
    - El estado activo se resuelve con el tipo primary.
    - Plantillas conserva un chevron discreto como acción secundaria.
    """
    active = page == target
    display_label = f"{label}  ›" if tool else label
    st.sidebar.button(
        display_label,
        key=f"nav_{target}",
        use_container_width=True,
        type="primary" if active else "secondary",
        on_click=change_page,
        args=(target,),
    )

if LOGO_BASE64:
    render_sidebar_html(f"""
    <div class="sidebar-shell sidebar-shell-top">
      <div class="grupo-maritex-logo">
        <img src="data:image/png;base64,{LOGO_BASE64}" alt="Grupo Maritex">
      </div>
      <div class="sidebar-menu-label">Menú principal</div>
    </div>
    """)
else:
    render_sidebar_html("""
    <div class="sidebar-shell sidebar-shell-top">
      <div class="grupo-maritex-logo-fallback">
        <span class="gm-x">X</span><span>Grupo Maritex</span>
      </div>
      <div class="sidebar-menu-label">Menú principal</div>
    </div>
    """)

render_sidebar_html(
    '<div class="vtex-nav-section"><span>OPERACIÓN</span><i></i></div>'
)
nav_button("Stock General", "stock_general")
nav_button("Marketplaces", "marketplace")

render_sidebar_html(
    '<div class="vtex-nav-section vtex-nav-section-spaced"><span>ANÁLISIS</span><i></i></div>'
)
nav_button("Métricas de Stock", "metricas")
nav_button("Métricas Vendedores", "vendedores")
nav_button("Resumen Ejecutivo", "ventas")

render_sidebar_html(
    '<div class="vtex-nav-section vtex-nav-section-spaced"><span>HERRAMIENTAS</span><i></i></div>'
)
nav_button("Plantillas", "configuracion", tool=True)

render_sidebar_html("""
<div class="sidebar-footer-clean sidebar-footer-fixed">
  <div class="sidebar-sync">
    <span class="sidebar-sync-dot"></span>
    <span>Sistema sincronizado</span>
  </div>
  <div class="sidebar-footer-rule"></div>
  <div class="sidebar-product">Maritex Inventory Control</div>
  <div class="sidebar-meta">Versión 61.1<br>ERP Flexline + Marketplaces</div>
  <div class="sidebar-footer-rule sidebar-footer-rule-bottom"></div>
  <div class="sidebar-copy">© 2026 Maritex</div>
</div>
""")

marketplace = st.session_state.get("marketplace_value", "Paris Marketplace")
reserve_stock = int(st.session_state.get("reserve_stock_value", 0))
use_max_stock = bool(st.session_state.get("use_max_stock_value", False))
max_stock = int(st.session_state.get("max_stock_value", 50)) if use_max_stock else None


# ============================================================
# VISTAS PRINCIPALES
# ============================================================

if page == "metricas":
    # ========================================================
    # MÉTRICAS DE STOCK · V52 VTEX ADMIN
    # Fuente única: Stock General
    # ========================================================
    ensure_stock_source_loaded()
    shared_stock_bytes = st.session_state.get("shared_stock_bytes")
    shared_stock_name = st.session_state.get("shared_stock_name")
    shared_stock_loaded_at = st.session_state.get("shared_stock_loaded_at")

    if not shared_stock_bytes:
        render_html("""
        <div class="v52-page-head">
            <div class="v52-breadcrumb">Análisis <span>/</span> <strong>Métricas de Stock</strong></div>
            <div class="v52-title">Métricas de Stock</div>
            <div class="v52-subtitle">Análisis y estado actual del inventario en base al Stock General.</div>
        </div>
        <div class="v52-empty-state">
            <div class="v52-empty-title">Aún no existe Stock General cargado</div>
            <div class="v52-empty-text">
                Esta página se actualiza automáticamente con el archivo cargado en
                <strong>Stock General</strong>. Carga la exportación de Flexline y vuelve aquí.
            </div>
        </div>
        """)
        st.stop()

    try:
        inv_raw = read_flexline(
            shared_stock_bytes,
            shared_stock_name or "stock_general.csv",
        )
        inv = add_inventory_status(build_inventory_source(inv_raw))
        inv_cons = add_inventory_status(consolidate_inventory_by_product(inv))
    except Exception as exc:
        show_error(exc, "Error leyendo Stock General para Métricas de Stock")
        st.stop()

    loaded_text = ""
    if shared_stock_loaded_at:
        try:
            loaded_text = shared_stock_loaded_at.strftime("%d %b. %Y %H:%M")
        except Exception:
            loaded_text = str(shared_stock_loaded_at)

    # ========================================================
    # NORMALIZACIÓN
    # ========================================================
    stock_view = inv_cons.copy()

    for col in [
        "Disponible",
        "Por llegar",
        "Por despachar",
        "Stock físico",
        "Precio",
        "Capacidad",
    ]:
        if col in stock_view.columns:
            stock_view[f"{col}_num"] = pd.to_numeric(
                stock_view[col], errors="coerce"
            ).fillna(0)

    for col in ["Disponible", "Por llegar", "Por despachar", "Precio"]:
        num_col = f"{col}_num"
        if num_col not in stock_view.columns:
            stock_view[num_col] = 0.0

    # ========================================================
    # HEADER
    # ========================================================
    render_html(f"""
    <div class="v52-topline">
        <div class="v52-breadcrumb">Análisis <span>/</span> <strong>Métricas de Stock</strong></div>
        <div class="v52-updated">
            Última actualización: <strong>{loaded_text or "sesión actual"}</strong>
        </div>
    </div>
    <div class="v52-page-head">
        <div class="v52-title">Métricas de Stock</div>
        <div class="v52-subtitle">
            Análisis y estado actual del inventario en base al Stock General.
        </div>
    </div>
    """)

    # ========================================================
    # FILTROS
    # ========================================================
    filter_wrap = st.container()
    with filter_wrap:
        f1, f2, f3, f4 = st.columns([2.0, 1, 1, 1], gap="small")

        with f1:
            stock_search = st.text_input(
                "Buscar SKU o producto",
                placeholder="Buscar SKU o producto",
                key="v52_metrics_search",
                label_visibility="collapsed",
            )

        with f2:
            family_options = (
                sorted(
                    stock_view["Familia"]
                    .fillna("Sin familia")
                    .astype(str)
                    .str.strip()
                    .unique()
                    .tolist()
                )
                if "Familia" in stock_view.columns else []
            )
            family_filter = st.multiselect(
                "Familia",
                family_options,
                placeholder="Familia",
                key="v52_family_filter",
                label_visibility="collapsed",
            )

        with f3:
            status_options = (
                sorted(
                    stock_view["Estado"]
                    .fillna("Sin estado")
                    .astype(str)
                    .unique()
                    .tolist()
                )
                if "Estado" in stock_view.columns else []
            )
            status_filter = st.multiselect(
                "Estado",
                status_options,
                placeholder="Estado",
                key="v52_status_filter",
                label_visibility="collapsed",
            )

        with f4:
            warehouse_options = (
                sorted(
                    stock_view["Bodega"]
                    .fillna("Sin bodega")
                    .astype(str)
                    .str.strip()
                    .unique()
                    .tolist()
                )
                if "Bodega" in stock_view.columns else []
            )
            warehouse_filter = st.multiselect(
                "Bodega",
                warehouse_options,
                placeholder="Bodega",
                key="v52_warehouse_filter",
                label_visibility="collapsed",
                disabled=not bool(warehouse_options),
            )

    filtered_stock = stock_view.copy()

    if stock_search:
        term = stock_search.strip().lower()
        mask = pd.Series(False, index=filtered_stock.index)
        for col in ["Código", "Producto"]:
            if col in filtered_stock.columns:
                mask = mask | (
                    filtered_stock[col]
                    .fillna("")
                    .astype(str)
                    .str.lower()
                    .str.contains(term, regex=False)
                )
        filtered_stock = filtered_stock[mask]

    if family_filter and "Familia" in filtered_stock.columns:
        filtered_stock = filtered_stock[
            filtered_stock["Familia"]
            .fillna("Sin familia")
            .astype(str)
            .str.strip()
            .isin(family_filter)
        ]

    if status_filter and "Estado" in filtered_stock.columns:
        filtered_stock = filtered_stock[
            filtered_stock["Estado"]
            .fillna("Sin estado")
            .astype(str)
            .isin(status_filter)
        ]

    if warehouse_filter and "Bodega" in filtered_stock.columns:
        filtered_stock = filtered_stock[
            filtered_stock["Bodega"]
            .fillna("Sin bodega")
            .astype(str)
            .str.strip()
            .isin(warehouse_filter)
        ]

    # ========================================================
    # KPI
    # ========================================================
    sku_total = (
        int(filtered_stock["Código"].nunique())
        if "Código" in filtered_stock.columns
        else int(len(filtered_stock))
    )
    units_available = int(filtered_stock["Disponible_num"].sum())
    units_incoming = int(filtered_stock["Por llegar_num"].sum())

    states = (
        filtered_stock["Estado"].fillna("").astype(str)
        if "Estado" in filtered_stock.columns
        else pd.Series("", index=filtered_stock.index)
    )

    low_mask = states.str.contains("Stock bajo", case=False, regex=False)
    zero_mask = (
        states.str.contains("Sin stock", case=False, regex=False)
        | states.str.contains("Negativo", case=False, regex=False)
    )
    available_mask = states.str.contains("Disponible", case=False, regex=False)

    inv_low = int(low_mask.sum())
    inv_zero = int(zero_mask.sum())
    inv_ok = int(available_mask.sum())

    render_html(f"""
    <div class="v52-kpi-grid">
        <div class="v52-kpi-card">
            <div class="v52-kpi-icon v52-green">◇</div>
            <div class="v52-kpi-label">Unidades disponibles</div>
            <div class="v52-kpi-value">{units_available:,}</div>
            <div class="v52-kpi-unit">unidades</div>
        </div>
        <div class="v52-kpi-card">
            <div class="v52-kpi-icon v52-blue">◇</div>
            <div class="v52-kpi-label">SKU totales</div>
            <div class="v52-kpi-value">{sku_total:,}</div>
            <div class="v52-kpi-unit">productos</div>
        </div>
        <div class="v52-kpi-card">
            <div class="v52-kpi-icon v52-orange">!</div>
            <div class="v52-kpi-label">Stock bajo</div>
            <div class="v52-kpi-value">{inv_low:,}</div>
            <div class="v52-kpi-unit">sku</div>
        </div>
        <div class="v52-kpi-card">
            <div class="v52-kpi-icon v52-red">−</div>
            <div class="v52-kpi-label">Sin stock</div>
            <div class="v52-kpi-value">{inv_zero:,}</div>
            <div class="v52-kpi-unit">sku</div>
        </div>
        <div class="v52-kpi-card">
            <div class="v52-kpi-icon v52-purple">▣</div>
            <div class="v52-kpi-label">Por llegar</div>
            <div class="v52-kpi-value">{units_incoming:,}</div>
            <div class="v52-kpi-unit">unidades</div>
        </div>
    </div>
    """)

    # mover filtros visualmente debajo de KPI mediante CSS
    # (Streamlit ya los renderizó arriba en DOM; el wrapper general mantiene jerarquía funcional)

    # ========================================================
    # GRÁFICOS
    # ========================================================
    render_html('<div class="v52-section-gap"></div>')
    c1, c2 = st.columns(2, gap="medium")

    with c1:
        render_html("""
        <div class="v52-card-head">
            <div class="v52-card-title">Distribución por estado de stock</div>
            <div class="v52-card-subtitle">SKU disponibles por estado</div>
        </div>
        """)

        status_groups = pd.DataFrame(
            {
                "Estado": ["Disponible", "Stock bajo", "Sin stock"],
                "SKU": [inv_ok, inv_low, inv_zero],
            }
        )
        status_groups = status_groups[status_groups["SKU"] > 0]

        if not status_groups.empty:
            donut = (
                alt.Chart(status_groups)
                .mark_arc(innerRadius=58, outerRadius=88)
                .encode(
                    theta=alt.Theta("SKU:Q"),
                    color=alt.Color(
                        "Estado:N",
                        scale=alt.Scale(
                            domain=["Disponible", "Stock bajo", "Sin stock"],
                            range=["#2DBE78", "#FFB020", "#FF4D4F"],
                        ),
                        legend=alt.Legend(
                            title=None,
                            orient="right",
                            labelFontSize=11,
                            symbolSize=120,
                        ),
                    ),
                    tooltip=[
                        alt.Tooltip("Estado:N", title="Estado"),
                        alt.Tooltip("SKU:Q", title="SKU", format=","),
                    ],
                )
                .properties(height=260)
            )
            st.altair_chart(donut, use_container_width=True)
        else:
            st.info("No existen estados disponibles para graficar.")

        render_html(
            f'<div class="v52-chart-total">Total: <strong>{sku_total:,} SKU</strong></div>'
        )

    with c2:
        render_html("""
        <div class="v52-card-head">
            <div class="v52-card-title">Stock por llegar</div>
            <div class="v52-card-subtitle">Unidades pendientes de ingreso</div>
        </div>
        """)

        incoming_chart_data = pd.DataFrame()

        if "Bodega" in filtered_stock.columns:
            incoming_chart_data = (
                filtered_stock.assign(
                    Grupo=filtered_stock["Bodega"]
                    .fillna("Sin bodega")
                    .astype(str)
                    .str.strip()
                )
                .groupby("Grupo", as_index=False)["Por llegar_num"]
                .sum()
                .rename(columns={"Por llegar_num": "Unidades"})
            )
        elif "Familia" in filtered_stock.columns:
            incoming_chart_data = (
                filtered_stock.assign(
                    Grupo=filtered_stock["Familia"]
                    .fillna("Sin familia")
                    .astype(str)
                    .str.strip()
                )
                .groupby("Grupo", as_index=False)["Por llegar_num"]
                .sum()
                .rename(columns={"Por llegar_num": "Unidades"})
            )

        incoming_chart_data = (
            incoming_chart_data[incoming_chart_data["Unidades"] > 0]
            .sort_values("Unidades", ascending=False)
            .head(8)
            if not incoming_chart_data.empty
            else incoming_chart_data
        )

        if not incoming_chart_data.empty:
            incoming_chart = (
                alt.Chart(incoming_chart_data)
                .mark_bar(cornerRadiusEnd=4, color="#4F7CD7")
                .encode(
                    y=alt.Y(
                        "Grupo:N",
                        sort="-x",
                        title=None,
                        axis=alt.Axis(labelLimit=130),
                    ),
                    x=alt.X("Unidades:Q", title=None, axis=None),
                    tooltip=[
                        alt.Tooltip("Grupo:N", title="Grupo"),
                        alt.Tooltip("Unidades:Q", format=","),
                    ],
                )
                .properties(height=260)
            )
            labels = (
                alt.Chart(incoming_chart_data)
                .mark_text(align="left", baseline="middle", dx=6, color="#4b5563")
                .encode(
                    y=alt.Y("Grupo:N", sort="-x"),
                    x="Unidades:Q",
                    text=alt.Text("Unidades:Q", format=","),
                )
            )
            st.altair_chart(incoming_chart + labels, use_container_width=True)
        else:
            st.info("No existen unidades por llegar en la selección actual.")

        render_html(
            f'<div class="v52-chart-total">Total: <strong>{units_incoming:,} unidades</strong></div>'
        )

    # ========================================================
    # PRIORIZACIÓN
    # ========================================================
    attention = filtered_stock.copy()

    def v52_priority(row):
        state = str(row.get("Estado", ""))
        available = float(row.get("Disponible_num", 0))
        incoming = float(row.get("Por llegar_num", 0))
        outgoing = float(row.get("Por despachar_num", 0))

        if "Negativo" in state:
            return 100, "Alta"
        if "Sin stock" in state and incoming <= 0:
            return 98, "Alta"
        if "Sin stock" in state:
            return 94, "Alta"
        if "Riesgo despacho" in state:
            return 92, "Alta"
        if outgoing > available and outgoing > 0:
            return 90, "Alta"
        if "Stock bajo" in state:
            return 75, "Media"
        return 0, "Baja"

    attention[["Prioridad_score", "Prioridad"]] = attention.apply(
        lambda r: pd.Series(v52_priority(r)),
        axis=1,
    )

    critical = (
        attention[attention["Prioridad_score"] > 0]
        .sort_values(
            ["Prioridad_score", "Disponible_num"],
            ascending=[False, True],
        )
        .head(10)
        .copy()
    )

    critical_count = int((attention["Prioridad_score"] > 0).sum())

    render_html(f"""
    <div class="v52-table-title-row">
        <div>
            <div class="v52-card-title">
                Productos que requieren atención
                <span class="v52-count-badge">{critical_count:,}</span>
            </div>
            <div class="v52-card-subtitle">Lista de SKU con situación crítica o de riesgo</div>
        </div>
    </div>
    """)

    if not critical.empty:
        status_display = critical.get("Estado", pd.Series("", index=critical.index)).astype(str)
        status_display = (
            status_display
            .str.replace("🟢 ", "", regex=False)
            .str.replace("🟡 ", "", regex=False)
            .str.replace("🔴 ", "", regex=False)
            .str.replace("🟠 ", "", regex=False)
            .str.replace("🔵 ", "", regex=False)
        )

        table = pd.DataFrame({
            "Producto": critical["Producto"] if "Producto" in critical.columns else "",
            "SKU": critical["Código"] if "Código" in critical.columns else "",
            "Disponible": critical["Disponible_num"].round().astype(int),
            "Por llegar": critical["Por llegar_num"].round().astype(int),
            "Comprometido": critical["Por despachar_num"].round().astype(int),
            "Estado": status_display,
            "Prioridad": critical["Prioridad"],
        })

        st.dataframe(
            table,
            hide_index=True,
            use_container_width=True,
            height=min(440, 40 * len(table) + 44),
            column_config={
                "Disponible": st.column_config.NumberColumn(format="%d"),
                "Por llegar": st.column_config.NumberColumn(format="%d"),
                "Comprometido": st.column_config.NumberColumn(format="%d"),
            },
        )
    else:
        st.success("No se detectaron SKU críticos con los filtros actuales.")

    # ========================================================
    # INSIGHTS
    # ========================================================
    zero_pct = ((inv_zero / sku_total) * 100) if sku_total else 0
    critical_pct = ((critical_count / sku_total) * 100) if sku_total else 0

    render_html(f"""
    <div class="v52-insights">
        <div class="v52-insights-head">
            <span class="v52-info-dot">i</span>
            <strong>Insights automáticos</strong>
        </div>
        <div class="v52-insights-grid">
            <div class="v52-insight-item">
                <div class="v52-insight-icon v52-blue">▥</div>
                <div>
                    <strong>{zero_pct:.1f}% de los SKU</strong> están sin stock o negativos.
                    Revisa primero los SKU críticos.
                </div>
            </div>
            <div class="v52-insight-item">
                <div class="v52-insight-icon v52-orange">!</div>
                <div>
                    <strong>{critical_count:,} SKU</strong> requieren atención prioritaria
                    por riesgo operacional.
                </div>
            </div>
            <div class="v52-insight-item">
                <div class="v52-insight-icon v52-green">▣</div>
                <div>
                    Hay <strong>{units_incoming:,} unidades por llegar</strong>
                    en la selección actual.
                </div>
            </div>
        </div>
    </div>
    """)

    # ========================================================
    # DETALLE EXPANDIBLE
    # ========================================================
    with st.expander("Ver detalle completo del inventario", expanded=False):
        detail_cols = [
            c for c in [
                "Estado",
                "Código",
                "Producto",
                "Familia",
                "Stock físico",
                "Disponible",
                "Por llegar",
                "Por despachar",
                "Precio",
                "Bodega",
            ]
            if c in filtered_stock.columns
        ]

        st.dataframe(
            filtered_stock[detail_cols],
            hide_index=True,
            use_container_width=True,
            height=520,
        )

    st.stop()


if page == "vendedores":
    render_html("""
    <div class="seller-head">
        <div>
            <div class="seller-title">Métricas Vendedores</div>
            <div class="seller-subtitle">
                Analiza ventas, cumplimiento y desempeño comercial por vendedor.
            </div>
        </div>
        <div class="seller-rule">● Venta final = Facturas + Boletas − NC</div>
    </div>
    """)

    # Fuente centralizada: ERP Ventas se carga únicamente desde Plantillas.
    ensure_sales_source_loaded()
    metrics_raw = st.session_state.get("metrics_bytes")
    metrics_name = st.session_state.get("metrics_name")

    if metrics_raw is None:
        st.info(
            "No existe una fuente ERP Ventas activa. "
            "Cárgala desde Plantillas → Fuentes ERP → ERP Ventas."
        )
        st.stop()

    try:
        base_df = read_metrics_data(
            metrics_raw,
            metrics_name or "metricas.csv",
        )
        base_df = ensure_sales_amount_column(base_df)
    except Exception as exc:
        show_error(exc, "Error leyendo métricas para Vendedores")
        st.stop()

    commercial_all = filter_commercial_documents(base_df)

    if commercial_all.empty:
        st.warning("No se encontraron Facturas, Boletas o Notas de crédito.")
        st.stop()

    # --------------------------------------------------------
    # FILTROS
    # --------------------------------------------------------
    date_min = (
        commercial_all["Fecha_dt"].min().date()
        if "Fecha_dt" in commercial_all.columns
        and commercial_all["Fecha_dt"].notna().any()
        else None
    )
    date_max = (
        commercial_all["Fecha_dt"].max().date()
        if "Fecha_dt" in commercial_all.columns
        and commercial_all["Fecha_dt"].notna().any()
        else None
    )

    # V56: el filtro puede llegar hasta hoy aunque la exportación ERP
    # todavía termine ayer. El cálculo usa únicamente registros existentes.
    today_date = datetime.now().date()
    selectable_date_max = (
        max(date_max, today_date)
        if date_max is not None
        else today_date
    )

    seller_months = available_sales_months(commercial_all)
    if not seller_months:
        st.warning("No existen meses válidos en ERP Ventas.")
        st.stop()

    seller_month_labels = [month_label_es(m) for m in seller_months]
    seller_month_map = dict(zip(seller_month_labels, seller_months))

    f1, f2, f3, f4 = st.columns([1.15, 1, 1.2, 1.2])

    with f1:
        selected_seller_month_label = st.selectbox(
            "Mes",
            seller_month_labels,
            index=0,
            key="seller_month_filter_v59",
            help="Selecciona el mes base para el análisis comercial.",
        )
        selected_seller_month = seller_month_map[selected_seller_month_label]
        seller_month_start, seller_month_end = month_bounds(selected_seller_month)

        # V59: rango de días independiente por cada mes para evitar
        # que Streamlit conserve fechas inválidas al cambiar de mes.
        seller_month_rows_for_default = commercial_all[
            (commercial_all["Fecha_dt"].dt.date >= seller_month_start)
            & (commercial_all["Fecha_dt"].dt.date <= seller_month_end)
        ]
        seller_default_end = seller_month_end
        if (
            not seller_month_rows_for_default.empty
            and seller_month_rows_for_default["Fecha_dt"].notna().any()
        ):
            seller_default_end = min(
                seller_month_end,
                seller_month_rows_for_default["Fecha_dt"].max().date(),
            )

        seller_day_range = st.date_input(
            "Días",
            value=(seller_month_start, seller_default_end),
            min_value=seller_month_start,
            max_value=seller_month_end,
            key=f"seller_day_range_v59_{selected_seller_month}",
            help="Filtra un rango de días dentro del mes seleccionado.",
        )

    with f2:
        vat_percent = st.number_input(
            "IVA %",
            min_value=0.0,
            max_value=100.0,
            value=19.0,
            step=1.0,
            key="seller_vat_rate",
        )
        vat_rate = float(vat_percent) / 100.0

    with f3:
        type_options = sorted(
            commercial_all["TipoDocto"]
            .dropna()
            .astype(str)
            .unique()
            .tolist()
        )
        type_filter = st.multiselect(
            "Tipo de documento",
            type_options,
            placeholder="Facturas + Boletas + NC",
            key="seller_document_type_filter",
        )

    with f4:
        seller_options = sorted(
            commercial_all["Vendedor"]
            .fillna("Sin vendedor")
            .astype(str)
            .str.strip()
            .unique()
            .tolist()
        )
        seller_filter = st.multiselect(
            "Vendedor",
            seller_options,
            placeholder="Todos",
            key="seller_filter_v45",
        )

    g1, g2 = st.columns([1, 1.4])

    with g1:
        warehouse_options = (
            sorted(
                commercial_all["Bodega"]
                .fillna("Sin bodega")
                .astype(str)
                .str.strip()
                .unique()
                .tolist()
            )
            if "Bodega" in commercial_all.columns
            else []
        )
        warehouse_filter = st.multiselect(
            "Bodega",
            warehouse_options,
            placeholder="Todas",
            key="seller_warehouse_filter_v45",
        )

    with g2:
        client_filter = st.text_input(
            "Cliente",
            placeholder="Buscar cliente y mostrar sus documentos…",
            key="seller_client_filter_v45",
        )

    view = commercial_all.copy()

    current_start = seller_month_start
    current_end = seller_default_end

    if (
        seller_day_range
        and isinstance(seller_day_range, (tuple, list))
        and len(seller_day_range) == 2
    ):
        current_start, current_end = seller_day_range

    # V59: mes + rango de días.
    view = view[
        (view["Fecha_dt"].dt.date >= current_start)
        & (view["Fecha_dt"].dt.date <= current_end)
    ]

    if type_filter:
        view = view[view["TipoDocto"].isin(type_filter)]

    if seller_filter:
        view = view[
            view["Vendedor"]
            .fillna("Sin vendedor")
            .astype(str)
            .str.strip()
            .isin(seller_filter)
        ]

    if warehouse_filter and "Bodega" in view.columns:
        view = view[
            view["Bodega"]
            .fillna("Sin bodega")
            .astype(str)
            .str.strip()
            .isin(warehouse_filter)
        ]

    if client_filter and "RazonSocial" in view.columns:
        term = client_filter.strip().lower()
        view = view[
            view["RazonSocial"]
            .fillna("")
            .astype(str)
            .str.lower()
            .str.contains(term, regex=False)
        ]

    # V57: estado de cobertura del mes seleccionado.
    month_rows = commercial_all[
        (commercial_all["Fecha_dt"].dt.date >= current_start)
        & (commercial_all["Fecha_dt"].dt.date <= current_end)
    ]
    month_data_max = (
        month_rows["Fecha_dt"].max().date()
        if not month_rows.empty and month_rows["Fecha_dt"].notna().any()
        else None
    )

    if month_data_max is not None:
        data_status_class = (
            "sales-data-status-warn"
            if month_data_max < current_end
            else "sales-data-status-ok"
        )

        render_html(f"""
        <div class="sales-data-status {data_status_class}">
            <div>
                <span>Mes seleccionado</span>
                <strong>{selected_seller_month_label}</strong>
            </div>
            <div>
                <span>Días analizados</span>
                <strong>{pd.Timestamp(current_start).strftime("%d/%m")} – {pd.Timestamp(current_end).strftime("%d/%m/%Y")}</strong>
            </div>
            <div>
                <span>Datos ERP disponibles hasta</span>
                <strong>{pd.Timestamp(month_data_max).strftime("%d/%m/%Y")}</strong>
            </div>
            <div class="sales-data-status-note">
                {
                    "El total utiliza todas las ventas disponibles dentro del rango seleccionado."
                    if month_data_max < current_end
                    else "El ERP cubre el rango de días seleccionado."
                }
            </div>
        </div>
        """)

    # --------------------------------------------------------
    # TOTAL FINAL
    # --------------------------------------------------------
    totals = calculate_commercial_totals(
        view,
        vat_rate=vat_rate,
    )

    render_html(f"""
    <div class="sales-rule-strip">
        <strong>Regla aplicada:</strong>
        Facturas y Boletas suman al total. Las Notas de crédito se restan.
        El valor sin IVA se calcula usando una tasa de {vat_percent:.0f}%.
        <strong>Haz clic en cualquier indicador para ver qué documentos lo componen.</strong>
    </div>
    """)

    # --------------------------------------------------------
    # KPIs clickeables
    # --------------------------------------------------------
    if "selected_sales_metric" not in st.session_state:
        st.session_state.selected_sales_metric = None

    k1, k2, k3, k4 = st.columns(4, gap="small")

    with k1:
        render_html('<div class="sales-kpi-button-wrap">')
        if st.button(
            f"Ventas brutas con IVA\n\n{format_clp(totals['ventas_brutas_con_iva'])}\n\nFacturas + Boletas antes de NC",
            key="sales_metric_gross",
            use_container_width=True,
        ):
            st.session_state.selected_sales_metric = "gross"
        render_html("</div>")

    with k2:
        render_html('<div class="sales-kpi-button-wrap credit">')
        if st.button(
            f"Notas de crédito\n\n− {format_clp(totals['notas_credito_con_iva'])}\n\nMonto descontado",
            key="sales_metric_credits",
            use_container_width=True,
        ):
            st.session_state.selected_sales_metric = "credits"
        render_html("</div>")

    with k3:
        render_html('<div class="sales-kpi-button-wrap main">')
        if st.button(
            f"Venta final con IVA\n\n{format_clp(totals['venta_neta_con_iva'])}\n\nFacturas + Boletas − NC",
            key="sales_metric_net_vat",
            use_container_width=True,
        ):
            st.session_state.selected_sales_metric = "net_vat"
        render_html("</div>")

    with k4:
        render_html('<div class="sales-kpi-button-wrap main">')
        if st.button(
            f"Venta final sin IVA\n\n{format_clp(totals['venta_neta_sin_iva'])}\n\nIVA neto: {format_clp(totals['iva_neto'])}",
            key="sales_metric_net_no_vat",
            use_container_width=True,
        ):
            st.session_state.selected_sales_metric = "net_no_vat"
        render_html("</div>")

    # --------------------------------------------------------
    # DETALLE DE LO QUE SE ESTÁ CONTABILIZANDO
    # --------------------------------------------------------
    # --------------------------------------------------------
    # META COMERCIAL + FALTANTE + PROGRESO
    # --------------------------------------------------------
    goal_col1, goal_col2 = st.columns([1, 1])

    with goal_col1:
        sales_goal = st.number_input(
            "Meta de ventas",
            min_value=0,
            value=int(st.session_state.get("seller_sales_goal", 20000000)),
            step=100000,
            key="seller_sales_goal_input",
            help="Meta del período filtrado, expresada con IVA.",
        )
        st.session_state["seller_sales_goal"] = int(sales_goal)

    with goal_col2:
        goal_basis = st.selectbox(
            "Base para medir la meta",
            ["Venta final con IVA", "Venta final sin IVA"],
            index=0,
            key="seller_goal_basis",
        )

    if goal_basis == "Venta final con IVA":
        goal_current = float(totals["venta_neta_con_iva"])
    else:
        goal_current = float(totals["venta_neta_sin_iva"])

    goal_value = float(sales_goal)
    goal_missing = max(goal_value - goal_current, 0)
    goal_over = max(goal_current - goal_value, 0)

    if goal_value > 0:
        goal_progress = max(min(goal_current / goal_value, 1.0), 0.0)
        goal_progress_pct = goal_current / goal_value * 100
    else:
        goal_progress = 0.0
        goal_progress_pct = 0.0

    render_html(f"""
    <div class="seller-goal-wrap">
        <div class="seller-goal-head">
            <div>
                <div class="seller-goal-title">Cumplimiento de meta</div>
                <div class="seller-goal-sub">
                    Base: {goal_basis} · período y filtros actualmente seleccionados.
                </div>
            </div>

            <div class="seller-goal-values">
                <div class="seller-goal-item">
                    <span>Meta</span>
                    <strong>{format_clp(goal_value)}</strong>
                </div>

                <div class="seller-goal-item">
                    <span>Venta actual</span>
                    <strong>{format_clp(goal_current)}</strong>
                </div>

                <div class="seller-goal-item">
                    <span>Faltante</span>
                    <strong>{format_clp(goal_missing)}</strong>
                </div>

                <div class="seller-goal-item">
                    <span>Cumplimiento</span>
                    <strong>{goal_progress_pct:.1f}%</strong>
                </div>
            </div>
        </div>

        <div class="seller-goal-bar">
            <div class="seller-goal-fill" style="width:{goal_progress*100:.2f}%"></div>
        </div>

        <div class="seller-goal-foot">
            <span>Avance: <strong>{goal_progress_pct:.1f}%</strong></span>
            <span>
                {"Meta superada por " + format_clp(goal_over) if goal_over > 0 else "Faltan " + format_clp(goal_missing) + " para alcanzar la meta"}
            </span>
        </div>
    </div>
    """)

    selected_metric = st.session_state.get("selected_sales_metric")

    if selected_metric:
        detail_view = view.copy()

        if selected_metric == "gross":
            detail_view = detail_view[
                detail_view["Grupo comercial"].isin(["Factura", "Boleta"])
            ].copy()
            detail_title = "Ventas brutas con IVA"
            detail_subtitle = (
                "Se contabilizan únicamente Facturas y Boletas. "
                "Las Notas de crédito todavía no se descuentan en este indicador."
            )

        elif selected_metric == "credits":
            detail_view = detail_view[
                detail_view["Grupo comercial"].eq("Nota de crédito")
            ].copy()
            detail_title = "Notas de crédito"
            detail_subtitle = (
                "Estos documentos se restan del total de Facturas + Boletas."
            )

        else:
            detail_view = detail_view[
                detail_view["Grupo comercial"].isin(
                    ["Factura", "Boleta", "Nota de crédito"]
                )
            ].copy()

            if selected_metric == "net_vat":
                detail_title = "Venta final con IVA"
                detail_subtitle = (
                    "Facturas y Boletas suman; las Notas de crédito restan. "
                    "El cálculo conserva IVA."
                )
            else:
                detail_title = "Venta final sin IVA"
                detail_subtitle = (
                    f"Facturas y Boletas suman; las Notas de crédito restan. "
                    f"Luego se elimina el IVA usando una tasa de {vat_percent:.0f}%."
                )

        # Impacto firmado del documento
        detail_view["Impacto con IVA"] = detail_view.apply(
            lambda r: (
                -abs(float(r["VentaMonto_num"]))
                if r["Grupo comercial"] == "Nota de crédito"
                else float(r["VentaMonto_num"])
            ),
            axis=1,
        )
        detail_view["Impacto sin IVA"] = (
            detail_view["Impacto con IVA"] / (1 + vat_rate)
        )

        detail_count = (
            detail_view["Numero"].nunique()
            if "Numero" in detail_view.columns
            else len(detail_view)
        )

        signed_total_vat = float(detail_view["Impacto con IVA"].sum())
        signed_total_no_vat = float(detail_view["Impacto sin IVA"].sum())

        render_html(f"""
        <div class="sales-detail-box">
            <div class="sales-detail-title">{detail_title}</div>
            <div class="sales-detail-subtitle">
                {detail_subtitle}<br>
                <strong>{detail_count:,}</strong> documentos ·
                Total con IVA: <strong>{format_clp(signed_total_vat)}</strong> ·
                Total sin IVA: <strong>{format_clp(signed_total_no_vat)}</strong>
            </div>
        </div>
        """)

        detail_columns = [
            c for c in [
                "Fecha",
                "TipoDocto",
                "Numero",
                "RazonSocial",
                "Vendedor",
                "Bodega",
                "Grupo comercial",
                "Total",
                "ReferenciaExterna",
            ]
            if c in detail_view.columns
        ]

        detail_display = detail_view[detail_columns].copy()

        detail_display["Efecto"] = detail_view["Grupo comercial"].map(
            lambda x: "− Resta" if x == "Nota de crédito" else "+ Suma"
        ).values

        detail_display["Impacto con IVA"] = detail_view["Impacto con IVA"].round().astype("Int64").values
        detail_display["Impacto sin IVA"] = detail_view["Impacto sin IVA"].round().astype("Int64").values

        if {"TipoDocto", "Numero"}.issubset(detail_display.columns):
            detail_display = detail_display.drop_duplicates(
                subset=["TipoDocto", "Numero"],
                keep="first",
            )

        st.dataframe(
            detail_display,
            hide_index=True,
            use_container_width=True,
            height=min(470, max(190, 45 + len(detail_display) * 34)),
            column_config={
                "Impacto con IVA": st.column_config.NumberColumn(
                    "Impacto con IVA",
                    format="$%d",
                ),
                "Impacto sin IVA": st.column_config.NumberColumn(
                    "Impacto sin IVA",
                    format="$%d",
                ),
                "Total": st.column_config.TextColumn("Total documento"),
            },
        )

        export_detail = dataframe_to_excel_bytes(
            detail_display,
            sheet_name="Detalle_Metrica",
        )

        e1, e2 = st.columns([1, 4])
        with e1:
            st.download_button(
                "⬇ Exportar detalle",
                data=export_detail,
                file_name=f"Detalle_{selected_metric}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"download_metric_detail_{selected_metric}",
                use_container_width=True,
            )
        with e2:
            if st.button(
                "Cerrar detalle",
                key="close_sales_metric_detail",
            ):
                st.session_state.selected_sales_metric = None
                st.rerun()

    # --------------------------------------------------------
    # DOCUMENTOS DEL CLIENTE
    # --------------------------------------------------------
    if client_filter:
        if not view.empty:
            client_names = (
                view["RazonSocial"]
                .fillna("Sin cliente")
                .astype(str)
                .str.strip()
                .drop_duplicates()
                .tolist()
            )

            client_label = (
                client_names[0]
                if len(client_names) == 1
                else f"{len(client_names)} clientes encontrados"
            )

            document_count = (
                view["Numero"].nunique()
                if "Numero" in view.columns else len(view)
            )

            render_html(f"""
            <div class="client-search-summary">
                <div>
                    <div class="client-search-title">
                        Documentos · {client_label}
                    </div>
                    <div class="client-search-meta">
                        Incluye Facturas, Boletas y Notas de crédito según los filtros activos.
                    </div>
                </div>
                <div class="client-search-kpis">
                    <div class="client-search-kpi">
                        <span>Documentos</span>
                        <strong>{document_count:,}</strong>
                    </div>
                    <div class="client-search-kpi">
                        <span>Venta final con IVA</span>
                        <strong>{format_clp(totals["venta_neta_con_iva"])}</strong>
                    </div>
                    <div class="client-search-kpi">
                        <span>Venta final sin IVA</span>
                        <strong>{format_clp(totals["venta_neta_sin_iva"])}</strong>
                    </div>
                </div>
            </div>
            """)

            document_columns = [
                c for c in [
                    "Fecha",
                    "TipoDocto",
                    "Numero",
                    "RazonSocial",
                    "Vendedor",
                    "Bodega",
                    "Total",
                    "TotalIngreso",
                    "Vigencia",
                    "Emitido",
                    "ReferenciaExterna",
                ]
                if c in view.columns
            ]

            client_docs = view[document_columns].copy()

            if {"TipoDocto", "Numero"}.issubset(client_docs.columns):
                client_docs = client_docs.drop_duplicates(
                    subset=["TipoDocto", "Numero"],
                    keep="first",
                )

            # Agregar signo comercial para hacer explícita la NC.
            type_map = (
                view[["TipoDocto", "Grupo comercial"]]
                .drop_duplicates("TipoDocto")
                .set_index("TipoDocto")["Grupo comercial"]
                if "TipoDocto" in view.columns
                else pd.Series(dtype=str)
            )

            client_docs.insert(
                0,
                "Efecto",
                client_docs["TipoDocto"].map(type_map).map(
                    lambda x: "− Resta" if x == "Nota de crédito" else "+ Suma"
                ),
            )

            st.dataframe(
                client_docs,
                hide_index=True,
                use_container_width=True,
                height=min(470, max(190, 45 + len(client_docs) * 35)),
            )

            client_export = dataframe_to_excel_bytes(
                client_docs,
                sheet_name="Documentos_Cliente",
            )

            st.download_button(
                "⬇ Exportar documentos del cliente",
                data=client_export,
                file_name=(
                    "Documentos_Cliente_"
                    + re.sub(r"[^A-Za-z0-9_-]+", "_", client_filter.strip())
                    + ".xlsx"
                ),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_client_documents_v45",
            )
        else:
            st.warning("No existen documentos para el cliente y filtros seleccionados.")

    # --------------------------------------------------------
    # RESUMEN POR VENDEDOR
    # --------------------------------------------------------
    render_html('<div class="seller-section-title">Ventas por vendedor</div>')

    work = view.copy()
    work["Vendedor"] = (
        work["Vendedor"]
        .fillna("Sin vendedor")
        .astype(str)
        .str.strip()
    )

    # Impacto firmado: Factura/Boleta suma, NC resta.
    work["Impacto con IVA"] = work.apply(
        lambda r: (
            -float(r["VentaMonto_num"])
            if r["Grupo comercial"] == "Nota de crédito"
            else float(r["VentaMonto_num"])
        ),
        axis=1,
    )
    work["Impacto sin IVA"] = work["Impacto con IVA"] / (1 + vat_rate)

    seller_summary = (
        work.groupby("Vendedor", as_index=False)
        .agg(
            Venta_con_IVA=("Impacto con IVA", "sum"),
            Venta_sin_IVA=("Impacto sin IVA", "sum"),
            Documentos=("Numero", "nunique"),
            Clientes=("RazonSocial", "nunique"),
        )
        .sort_values("Venta_con_IVA", ascending=False)
    )

    total_net = seller_summary["Venta_con_IVA"].sum()
    seller_summary["Participación %"] = (
        seller_summary["Venta_con_IVA"] / total_net * 100
        if total_net else 0
    )
    seller_summary["Ticket promedio"] = seller_summary.apply(
        lambda r: (
            r["Venta_con_IVA"] / r["Documentos"]
            if r["Documentos"] else 0
        ),
        axis=1,
    )

    # Meta de referencia por vendedor:
    # se distribuye proporcionalmente según participación actual.
    # Si luego se cargan metas individuales, este bloque puede reemplazarse.
    basis_col = (
        "Venta_con_IVA"
        if goal_basis == "Venta final con IVA"
        else "Venta_sin_IVA"
    )

    basis_total = float(seller_summary[basis_col].sum())

    if basis_total > 0 and goal_value > 0:
        seller_summary["Meta referencia"] = (
            seller_summary[basis_col] / basis_total * goal_value
        )
    else:
        seller_summary["Meta referencia"] = 0.0

    seller_summary["Faltante meta"] = (
        seller_summary["Meta referencia"] - seller_summary[basis_col]
    ).clip(lower=0)

    seller_summary["Cumplimiento meta %"] = seller_summary.apply(
        lambda r: (
            (r[basis_col] / r["Meta referencia"] * 100)
            if r["Meta referencia"] > 0
            else 0
        ),
        axis=1,
    )

    # --------------------------------------------------------
    # COMPARACIÓN CONTRA PERÍODO ANTERIOR + INSIGHTS
    # --------------------------------------------------------
    previous_summary = pd.DataFrame()
    overall_previous = 0.0
    overall_variation = 0.0

    if current_start is not None and current_end is not None and "Fecha_dt" in commercial_all.columns:
        period_days = max((pd.Timestamp(current_end) - pd.Timestamp(current_start)).days + 1, 1)
        previous_end = pd.Timestamp(current_start) - pd.Timedelta(seconds=1)
        previous_start = pd.Timestamp(current_start) - pd.Timedelta(days=period_days)

        prev_view = commercial_all[
            (commercial_all["Fecha_dt"] >= previous_start)
            & (commercial_all["Fecha_dt"] <= previous_end)
        ].copy()

        if type_filter:
            prev_view = prev_view[prev_view["TipoDocto"].isin(type_filter)]
        if seller_filter:
            prev_view = prev_view[
                prev_view["Vendedor"].fillna("Sin vendedor").astype(str).str.strip().isin(seller_filter)
            ]
        if warehouse_filter and "Bodega" in prev_view.columns:
            prev_view = prev_view[
                prev_view["Bodega"].fillna("Sin bodega").astype(str).str.strip().isin(warehouse_filter)
            ]
        if client_filter and "RazonSocial" in prev_view.columns:
            term = client_filter.strip().lower()
            prev_view = prev_view[
                prev_view["RazonSocial"].fillna("").astype(str).str.lower().str.contains(term, regex=False)
            ]

        if not prev_view.empty:
            prev_view["Vendedor"] = prev_view["Vendedor"].fillna("Sin vendedor").astype(str).str.strip()
            prev_view["Impacto con IVA"] = prev_view.apply(
                lambda r: -abs(float(r["VentaMonto_num"])) if r["Grupo comercial"] == "Nota de crédito" else float(r["VentaMonto_num"]),
                axis=1,
            )
            previous_summary = (
                prev_view.groupby("Vendedor", as_index=False)
                .agg(Venta_anterior=("Impacto con IVA", "sum"))
            )
            overall_previous = float(previous_summary["Venta_anterior"].sum())
            current_total_cmp = float(seller_summary["Venta_con_IVA"].sum())
            overall_variation = (
                ((current_total_cmp - overall_previous) / overall_previous) * 100
                if overall_previous != 0 else (100.0 if current_total_cmp > 0 else 0.0)
            )

            seller_summary = seller_summary.merge(previous_summary, on="Vendedor", how="left")
            seller_summary["Venta_anterior"] = seller_summary["Venta_anterior"].fillna(0)
            seller_summary["Variación vs anterior %"] = seller_summary.apply(
                lambda r: (
                    ((r["Venta_con_IVA"] - r["Venta_anterior"]) / r["Venta_anterior"] * 100)
                    if r["Venta_anterior"] != 0
                    else (100.0 if r["Venta_con_IVA"] > 0 else 0.0)
                ),
                axis=1,
            )
        else:
            seller_summary["Venta_anterior"] = 0.0
            seller_summary["Variación vs anterior %"] = 0.0
    else:
        seller_summary["Venta_anterior"] = 0.0
        seller_summary["Variación vs anterior %"] = 0.0

    # Tasa de notas de crédito sobre ventas brutas.
    gross_amount = float(work.loc[work["Grupo comercial"].isin(["Factura", "Boleta"]), "VentaMonto_num"].sum())
    credit_amount = float(work.loc[work["Grupo comercial"].eq("Nota de crédito"), "VentaMonto_num"].sum())
    credit_rate = (credit_amount / gross_amount * 100) if gross_amount > 0 else 0.0

    # KPIs ejecutivos de vendedores.
    top_seller = seller_summary.iloc[0]["Vendedor"] if not seller_summary.empty else "—"
    top_seller_value = float(seller_summary.iloc[0]["Venta_con_IVA"]) if not seller_summary.empty else 0.0
    avg_ticket = (
        float(work.loc[work["Grupo comercial"].isin(["Factura", "Boleta"]), "VentaMonto_num"].sum())
        / max(int(work.loc[work["Grupo comercial"].isin(["Factura", "Boleta"]), "Numero"].nunique()), 1)
        if "Numero" in work.columns else 0.0
    )

    exec1, exec2, exec3, exec4 = st.columns(4, gap="small")
    exec1.metric("Mejor vendedor", str(top_seller), format_clp(top_seller_value))
    exec2.metric("Ticket promedio", format_clp(avg_ticket))
    exec3.metric("Variación vs período anterior", f"{overall_variation:+.1f}%")
    exec4.metric("Tasa notas de crédito", f"{credit_rate:.1f}%")

    # Insights automáticos, calculados únicamente desde los datos filtrados.
    insights = []
    if not seller_summary.empty:
        leader = seller_summary.iloc[0]
        insights.append(
            f"🏆 {leader['Vendedor']} lidera el período con {format_clp(leader['Venta_con_IVA'])} "
            f"y {leader['Participación %']:.1f}% de participación."
        )

        if len(seller_summary) > 1:
            movers = seller_summary.sort_values("Variación vs anterior %", ascending=False)
            best_growth = movers.iloc[0]
            worst_growth = movers.iloc[-1]
            if best_growth["Venta_anterior"] > 0:
                insights.append(
                    f"📈 Mayor crecimiento: {best_growth['Vendedor']} ({best_growth['Variación vs anterior %']:+.1f}% vs período anterior)."
                )
            if worst_growth["Venta_anterior"] > 0 and worst_growth["Variación vs anterior %"] < 0:
                insights.append(
                    f"⚠️ Mayor caída: {worst_growth['Vendedor']} ({worst_growth['Variación vs anterior %']:+.1f}% vs período anterior)."
                )

        if credit_rate >= 5:
            insights.append(f"🔴 Las notas de crédito equivalen al {credit_rate:.1f}% de las ventas brutas del período.")
        elif credit_rate > 0:
            insights.append(f"↩️ Las notas de crédito equivalen al {credit_rate:.1f}% de las ventas brutas.")

        below_goal = int((seller_summary["Cumplimiento meta %"] < 80).sum())
        if below_goal:
            insights.append(f"🎯 {below_goal} vendedor(es) están bajo 80% de su meta de referencia.")

    if insights:
        render_html('<div class="seller-section-title">Insights automáticos</div>')
        for insight in insights[:5]:
            st.info(insight)

    rank_col, mix_col = st.columns([1.3, 1], gap="medium")

    with rank_col:
        render_html(
            '<div class="seller-card"><div class="seller-card-title">'
            'Ranking venta final con IVA</div>'
        )

        chart_data = seller_summary.head(15)

        chart = (
            alt.Chart(chart_data)
            .mark_bar(cornerRadiusEnd=5)
            .encode(
                y=alt.Y("Vendedor:N", sort="-x", title=None),
                x=alt.X("Venta_con_IVA:Q", title=None),
                tooltip=[
                    alt.Tooltip("Vendedor:N"),
                    alt.Tooltip("Venta_con_IVA:Q", title="Con IVA", format=","),
                    alt.Tooltip("Venta_sin_IVA:Q", title="Sin IVA", format=","),
                    alt.Tooltip("Documentos:Q", format=","),
                ],
            )
            .properties(height=330)
        )
        st.altair_chart(chart, use_container_width=True)
        render_html("</div>")

    with mix_col:
        render_html(
            '<div class="seller-card"><div class="seller-card-title">'
            'Composición documental</div>'
        )

        mix = (
            work.groupby("Grupo comercial", as_index=False)
            .agg(
                Monto=("VentaMonto_num", "sum"),
                Documentos=("Numero", "nunique"),
            )
        )

        mix_chart = (
            alt.Chart(mix)
            .mark_arc(innerRadius=65)
            .encode(
                theta="Monto:Q",
                color=alt.Color(
                    "Grupo comercial:N",
                    legend=alt.Legend(title=None),
                ),
                tooltip=[
                    alt.Tooltip("Grupo comercial:N"),
                    alt.Tooltip("Monto:Q", format=","),
                    alt.Tooltip("Documentos:Q", format=","),
                ],
            )
            .properties(height=330)
        )

        st.altair_chart(mix_chart, use_container_width=True)
        render_html("</div>")

    # --------------------------------------------------------
    # EVOLUCIÓN NETA
    # --------------------------------------------------------
    render_html('<div class="seller-section-title">Evolución de venta neta</div>')

    if "Fecha_dt" in work.columns and work["Fecha_dt"].notna().any():
        daily = (
            work.dropna(subset=["Fecha_dt"])
            .assign(Dia=lambda d: d["Fecha_dt"].dt.floor("D"))
            .groupby("Dia", as_index=False)
            .agg(
                Venta_con_IVA=("Impacto con IVA", "sum"),
                Venta_sin_IVA=("Impacto sin IVA", "sum"),
            )
            .sort_values("Dia")
        )

        evo = (
            alt.Chart(daily)
            .transform_fold(
                ["Venta_con_IVA", "Venta_sin_IVA"],
                as_=["Serie", "Monto"],
            )
            .mark_line(point=True, strokeWidth=2)
            .encode(
                x=alt.X("Dia:T", title=None, axis=alt.Axis(format="%d %b")),
                y=alt.Y("Monto:Q", title=None),
                color=alt.Color(
                    "Serie:N",
                    legend=alt.Legend(
                        title=None,
                        labelExpr=(
                            "datum.label == 'Venta_con_IVA' "
                            "? 'Con IVA' : 'Sin IVA'"
                        ),
                    ),
                ),
                tooltip=[
                    alt.Tooltip("Dia:T", format="%d/%m/%Y"),
                    alt.Tooltip("Serie:N"),
                    alt.Tooltip("Monto:Q", format=","),
                ],
            )
            .properties(height=300)
        )

        st.altair_chart(evo, use_container_width=True)

    # --------------------------------------------------------
    # TABLA
    # --------------------------------------------------------
    render_html('<div class="seller-section-title">Ranking detallado</div>')

    seller_table = seller_summary.rename(
        columns={
            "Venta_con_IVA": "Venta final con IVA",
            "Venta_sin_IVA": "Venta final sin IVA",
        }
    ).copy()

    for col in [
        "Venta final con IVA",
        "Venta final sin IVA",
        "Ticket promedio",
        "Meta referencia",
        "Faltante meta",
    ]:
        seller_table[col] = seller_table[col].round().astype("Int64")

    seller_table["Participación %"] = seller_table["Participación %"].round(1)
    seller_table["Cumplimiento meta %"] = seller_table["Cumplimiento meta %"].round(1)
    if "Venta_anterior" in seller_table.columns:
        seller_table["Venta_anterior"] = seller_table["Venta_anterior"].round().astype("Int64")
    if "Variación vs anterior %" in seller_table.columns:
        seller_table["Variación vs anterior %"] = seller_table["Variación vs anterior %"].round(1)

    st.dataframe(
        seller_table,
        hide_index=True,
        use_container_width=True,
        height=570,
        column_config={
            "Venta final con IVA": st.column_config.NumberColumn(
                "Venta final con IVA",
                format="$%d",
            ),
            "Venta final sin IVA": st.column_config.NumberColumn(
                "Venta final sin IVA",
                format="$%d",
            ),
            "Ticket promedio": st.column_config.NumberColumn(
                "Ticket promedio",
                format="$%d",
            ),
            "Participación %": st.column_config.NumberColumn(
                "Participación",
                format="%.1f%%",
            ),
            "Meta referencia": st.column_config.NumberColumn(
                "Meta referencia",
                format="$%d",
            ),
            "Faltante meta": st.column_config.NumberColumn(
                "Faltante meta",
                format="$%d",
            ),
            "Cumplimiento meta %": st.column_config.NumberColumn(
                "Cumplimiento meta",
                format="%.1f%%",
            ),
            "Venta_anterior": st.column_config.NumberColumn(
                "Venta período anterior",
                format="$%d",
            ),
            "Variación vs anterior %": st.column_config.NumberColumn(
                "Variación vs anterior",
                format="%.1f%%",
            ),
        },
    )

    st.caption(
        "Facturas y Boletas suman. Notas de crédito restan. "
        f"Montos sin IVA calculados con tasa {vat_percent:.0f}%."
    )

    st.stop()



if page == "configuracion":
    render_html("""
    <div class="inventory-head template-compact-wrap template-compact-head">
        <div>
            <div class="inventory-head-title">Plantillas y Fuentes ERP</div>
            <div class="inventory-head-subtitle template-compact-subtitle">
                Centro único para cargar las fuentes de Stock, Ventas y las plantillas de Marketplaces.
            </div>
        </div>
    </div>
    """)

    render_html("""
    <div class="template-info-strip">
        <strong>Flujo de datos:</strong> ERP Stock alimenta Stock General, Métricas de Stock y Marketplaces.
        ERP Ventas alimenta Métricas Vendedores. Los archivos quedan guardados para reutilizarlos automáticamente.
    </div>
    """)

    render_html('<div class="metrics-section-title">Fuentes ERP</div>')
    source_stock, source_sales = st.columns(2, gap="medium")

    with source_stock:
        with st.container(border=True):
            render_html('<div class="template-card-title">ERP Stock</div>')
            render_html('<div class="template-card-meta">Fuente maestra de inventario para Stock General, Métricas de Stock y Marketplaces.</div>')

            saved_stock_raw, saved_stock_name, saved_stock_time = load_erp_source("stock")
            if saved_stock_raw is not None:
                st.success(f"✓ Fuente activa: {saved_stock_name}")
                render_html(
                    f'<div class="template-card-meta">Actualizada: {saved_stock_time.strftime("%d/%m/%Y %H:%M")}</div>'
                )
            else:
                st.warning("Aún no existe una fuente ERP Stock guardada.")

            stock_upload = st.file_uploader(
                "Cargar / reemplazar ERP Stock",
                type=["csv", "xls", "xlsx"],
                key="config_erp_stock_source_v54",
                help="Formatos permitidos: CSV, XLS y XLSX. Esta será la única fuente de inventario del dashboard.",
            )

            if stock_upload is not None:
                try:
                    stock_bytes = bytes(stock_upload.getvalue())
                    # Validar antes de guardar.
                    test_inventory = read_flexline(stock_bytes, stock_upload.name)
                    build_inventory_source(test_inventory)

                    saved = save_erp_source("stock", stock_upload)
                    st.session_state.shared_stock_bytes = stock_bytes
                    st.session_state.shared_stock_name = stock_upload.name
                    st.session_state.shared_stock_loaded_at = datetime.now()
                    st.success(f"✓ ERP Stock actualizado: {saved.name}")
                except Exception as exc:
                    show_error(exc, "Error cargando ERP Stock")

    with source_sales:
        with st.container(border=True):
            render_html('<div class="template-card-title">ERP Ventas</div>')
            render_html('<div class="template-card-meta">Fuente comercial para Métricas Vendedores: facturas, boletas y notas de crédito.</div>')

            saved_sales_raw, saved_sales_name, saved_sales_time = load_erp_source("ventas")
            if saved_sales_raw is not None:
                st.success(f"✓ Fuente activa: {saved_sales_name}")
                render_html(
                    f'<div class="template-card-meta">Actualizada: {saved_sales_time.strftime("%d/%m/%Y %H:%M")}</div>'
                )
                try:
                    active_sales_check = read_metrics_data(saved_sales_raw, saved_sales_name)
                    active_sales_check = ensure_sales_amount_column(active_sales_check)
                    active_amount_col, active_amount_label, active_amount_sum = resolve_sales_amount_column(active_sales_check)
                    active_commercial = filter_commercial_documents(active_sales_check)
                    active_totals = calculate_commercial_totals(active_commercial)

                    active_dates = (
                        active_commercial["Fecha_dt"].dropna()
                        if "Fecha_dt" in active_commercial.columns
                        else pd.Series(dtype="datetime64[ns]")
                    )
                    active_first_date = (
                        active_dates.min().strftime("%d/%m/%Y")
                        if not active_dates.empty else "Sin fecha"
                    )
                    active_last_date = (
                        active_dates.max().strftime("%d/%m/%Y")
                        if not active_dates.empty else "Sin fecha"
                    )

                    active_docs = (
                        int(active_commercial["Numero"].nunique())
                        if "Numero" in active_commercial.columns
                        else int(len(active_commercial))
                    )
                    active_sellers = (
                        int(active_commercial["Vendedor"].fillna("Sin vendedor").astype(str).str.strip().nunique())
                        if "Vendedor" in active_commercial.columns
                        else 0
                    )

                    render_html(f"""
                    <div class="template-card-meta" style="margin-top:.35rem;">
                        Campo monetario: <strong>{active_amount_label or "No detectado"}</strong> ·
                        {active_docs:,} documentos · {active_sellers:,} vendedores ·
                        Fechas: <strong>{active_first_date} → {active_last_date}</strong> ·
                        Venta neta: <strong>{format_clp(active_totals["venta_neta_con_iva"])}</strong>
                    </div>
                    """)
                except Exception as active_exc:
                    st.warning(f"No fue posible validar la fuente activa: {active_exc}")
            else:
                st.warning("Aún no existe una fuente ERP Ventas guardada.")

            sales_upload = st.file_uploader(
                "Cargar / reemplazar ERP Ventas",
                type=["csv", "xls", "xlsx"],
                key="config_erp_sales_source_v54",
                help="Formatos permitidos: CSV, XLS y XLSX. Fuente usada por Métricas Vendedores.",
            )

            if sales_upload is not None:
                try:
                    sales_bytes = bytes(sales_upload.getvalue())
                    test_sales = read_metrics_data(sales_bytes, sales_upload.name)
                    if "TipoDocto" not in test_sales.columns:
                        raise ValueError("El archivo ERP Ventas no contiene la columna TipoDocto.")

                    test_sales = ensure_sales_amount_column(test_sales)
                    amount_col, amount_label, amount_abs_sum = resolve_sales_amount_column(test_sales)

                    if amount_col is None:
                        raise ValueError(
                            "No se encontró una columna monetaria compatible. "
                            "Se esperaba Total, TotalIngreso, Monto, Importe o Valor."
                        )

                    commercial_test = filter_commercial_documents(test_sales)
                    if commercial_test.empty:
                        raise ValueError(
                            "El archivo no contiene Facturas, Boletas o Notas de crédito reconocibles."
                        )

                    commercial_test = ensure_sales_amount_column(commercial_test)
                    validation_totals = calculate_commercial_totals(commercial_test)

                    docs_detected = (
                        int(commercial_test["Numero"].nunique())
                        if "Numero" in commercial_test.columns
                        else int(len(commercial_test))
                    )
                    sellers_detected = (
                        int(
                            commercial_test["Vendedor"]
                            .fillna("Sin vendedor")
                            .astype(str)
                            .str.strip()
                            .nunique()
                        )
                        if "Vendedor" in commercial_test.columns
                        else 0
                    )
                    net_detected = validation_totals["venta_neta_con_iva"]

                    valid_dates_detected = (
                        commercial_test["Fecha_dt"].dropna()
                        if "Fecha_dt" in commercial_test.columns
                        else pd.Series(dtype="datetime64[ns]")
                    )
                    first_date_detected = (
                        valid_dates_detected.min().strftime("%d/%m/%Y")
                        if not valid_dates_detected.empty
                        else "Sin fecha"
                    )
                    last_date_detected = (
                        valid_dates_detected.max().strftime("%d/%m/%Y")
                        if not valid_dates_detected.empty
                        else "Sin fecha"
                    )

                    if amount_abs_sum <= 0:
                        raise ValueError(
                            f"Se detectó el campo monetario '{amount_label}', "
                            "pero todos sus valores son $0. Revisa la exportación de Flexline."
                        )

                    saved = save_erp_source("ventas", sales_upload)
                    st.session_state["metrics_bytes"] = sales_bytes
                    st.session_state["metrics_name"] = sales_upload.name
                    st.session_state["metrics_loaded_at"] = datetime.now()

                    st.success(f"✓ ERP Ventas actualizado: {saved.name}")
                    render_html(f"""
                    <div class="metrics-source" style="margin-top:.55rem;">
                        <div>
                            <span class="metrics-dot"></span>
                            <strong>Validación ERP Ventas correcta</strong>
                        </div>
                        <div>
                            Campo monetario: <strong>{amount_label}</strong> ·
                            {docs_detected:,} documentos ·
                            {sellers_detected:,} vendedores ·
                            Fechas: <strong>{first_date_detected} → {last_date_detected}</strong> ·
                            Venta neta detectada: <strong>{format_clp(net_detected)}</strong>
                        </div>
                    </div>
                    """)
                except Exception as exc:
                    show_error(exc, "Error cargando ERP Ventas")

    render_html('<div class="metrics-section-title">Plantillas Marketplaces</div>')
    c_paris, c_meli = st.columns(2, gap="medium")

    with c_paris:
        with st.container(border=True):
            render_html('<div class="template-card-title">Paris Marketplace</div>')
            paris_path = template_path_for("Paris Marketplace")

            if paris_path.exists():
                st.success(f"✓ {paris_path.name}")
                render_html(
                    f'<div class="template-card-meta">Actualizada: {datetime.fromtimestamp(paris_path.stat().st_mtime).strftime("%d/%m/%Y %H:%M")}</div>'
                )
            else:
                st.warning("Plantilla Paris no guardada.")

            paris_upload = st.file_uploader(
                "Reemplazar plantilla Paris",
                type=["xlsx"],
                key="config_paris_template",
            )

            if paris_upload is not None:
                try:
                    wb = load_workbook(io.BytesIO(paris_upload.getvalue()), read_only=True, data_only=False)
                    expected_sheet = MARKETPLACE_CONFIGS["Paris Marketplace"].sheet_name
                    if expected_sheet not in wb.sheetnames:
                        wb.close()
                        st.error(f"La plantilla Paris debe contener la hoja '{expected_sheet}'.")
                    else:
                        wb.close()
                        saved = save_marketplace_template("Paris Marketplace", paris_upload)
                        generate_saved_marketplace_output.clear()
                        st.success(f"✓ Guardada como {saved.name}")
                except Exception as exc:
                    show_error(exc, "Error guardando plantilla Paris")

    with c_meli:
        with st.container(border=True):
            render_html('<div class="template-card-title">Mercado Libre</div>')
            meli_path = template_path_for("Mercado Libre")

            if meli_path.exists():
                st.success(f"✓ {meli_path.name}")
                render_html(
                    f'<div class="template-card-meta">Actualizada: {datetime.fromtimestamp(meli_path.stat().st_mtime).strftime("%d/%m/%Y %H:%M")}</div>'
                )
            else:
                st.warning("Plantilla Mercado Libre no guardada.")

            meli_upload = st.file_uploader(
                "Reemplazar plantilla Mercado Libre",
                type=["xlsx"],
                key="config_meli_template",
            )

            if meli_upload is not None:
                try:
                    wb = load_workbook(io.BytesIO(meli_upload.getvalue()), read_only=True, data_only=False)
                    validate_marketplace_template(wb, "Mercado Libre")
                    wb.close()
                    saved = save_marketplace_template("Mercado Libre", meli_upload)
                    generate_saved_marketplace_output.clear()
                    st.success(f"✓ Guardada como {saved.name}")
                except Exception as exc:
                    show_error(exc, "Error guardando plantilla Mercado Libre")

    render_html("""
    <div class="template-ops-bar">
        <div class="template-ops-item"><span class="template-ops-label">ERP Stock</span><span class="template-ops-value">Stock + Marketplaces</span></div>
        <div class="template-ops-item"><span class="template-ops-label">ERP Ventas</span><span class="template-ops-value">Métricas Vendedores</span></div>
        <div class="template-ops-item"><span class="template-ops-label">Bodega Stock MP</span><span class="template-ops-value">Casa Matriz</span></div>
        <div class="template-ops-status">● Fuentes centralizadas</div>
    </div>
    """)

    st.stop()

if page == "stock_general":
    # --------------------------------------------------------
    # HEADER
    # --------------------------------------------------------
    head_left, head_right = st.columns([5, 1.15])
    with head_left:
        render_html("""
        <div class="v39-stock-page v39-page-head">
            <div>
                <div class="v39-page-title">Stock General</div>
                <div class="v39-page-subtitle">
                    Consulta y administra el inventario consolidado proveniente de Flexline
                </div>
            </div>
        </div>
        """)
    with head_right:
        st.write("")
        if st.button("↻  Actualizar datos", use_container_width=True, key="refresh_stock_v39"):
            read_flexline.clear()
            st.rerun()

    # Fuente centralizada: se carga únicamente desde Plantillas.
    ensure_stock_source_loaded()
    shared_stock_bytes = st.session_state.get("shared_stock_bytes")
    shared_stock_name = st.session_state.get("shared_stock_name")

    if shared_stock_bytes is None or not shared_stock_name:
        render_html("""
        <div class="page-empty">
            No existe una fuente <strong>ERP Stock</strong> activa.
            Cárgala desde <strong>Plantillas → Fuentes ERP → ERP Stock</strong>.
            Stock General, Métricas de Stock y Marketplaces usarán automáticamente esa misma fuente.
        </div>
        """)
        st.stop()

    try:
        raw_inventory = read_flexline(shared_stock_bytes, shared_stock_name)
        inventory = add_inventory_status(build_inventory_source(raw_inventory))

        loaded_time = st.session_state.get("shared_stock_loaded_at")
        loaded_label = loaded_time.strftime("%d/%m/%Y %H:%M") if loaded_time else "Sesión actual"

        render_html(f"""
        <div class="v39-file-card">
            <div class="v39-file-icon">▤</div>
            <div style="flex:1">
                <div class="v39-file-name">{shared_stock_name} &nbsp; <span style="color:#26b66f">●</span></div>
                <div class="v39-file-meta">
                    {len(inventory):,} registros cargados · Compartido con Marketplaces · {loaded_label}
                </div>
            </div>
        </div>
        """)

        # ----------------------------------------------------
        # KPI + STATUS
        # ----------------------------------------------------
        consolidated_full = add_inventory_status(consolidate_inventory_by_product(inventory))

        unique_products = int(consolidated_full["Código"].nunique())
        total_available = int(pd.to_numeric(consolidated_full["Disponible"], errors="coerce").fillna(0).sum())
        warehouse_count = int(inventory["Bodega"].replace("", pd.NA).dropna().nunique())

        status_counts = consolidated_full["Estado"].value_counts().to_dict()
        count_available = int(status_counts.get("🟢 Disponible", 0))
        products_low = int(status_counts.get("🟡 Stock bajo", 0))
        risk_dispatch = int(status_counts.get("🟠 Riesgo despacho", 0))
        products_zero = int(status_counts.get("🔴 Sin stock", 0))
        products_negative = int(status_counts.get("🔴 Negativo", 0))
        count_incoming = int(status_counts.get("🔵 Por llegar", 0))
        unavailable = products_zero + products_negative
        attention_total = products_negative + products_zero + products_low + risk_dispatch

        render_html(f"""
        <div class="v39-kpi-grid">
          <div class="v39-kpi-card">
            <div class="v39-kpi-top">
              <div class="v39-kpi-icon v39-icon-purple">◇</div>
              <div><div class="v39-kpi-label">Stock Totales</div><div class="v39-kpi-value">{unique_products:,}</div></div>
            </div>
            <div class="v39-kpi-foot">▧ &nbsp; SKU únicos consolidados</div>
          </div>
          <div class="v39-kpi-card">
            <div class="v39-kpi-top">
              <div class="v39-kpi-icon v39-icon-green">▤</div>
              <div><div class="v39-kpi-label">Unidades Totales</div><div class="v39-kpi-value">{total_available:,}</div></div>
            </div>
            <div class="v39-kpi-foot">◉ &nbsp; En todas las bodegas</div>
          </div>
          <div class="v39-kpi-card">
            <div class="v39-kpi-top">
              <div class="v39-kpi-icon v39-icon-blue">⌂</div>
              <div><div class="v39-kpi-label">Bodegas Activas</div><div class="v39-kpi-value">{warehouse_count:,}</div></div>
            </div>
            <div class="v39-kpi-foot">⌘ &nbsp; Distribución operativa</div>
          </div>
          <div class="v39-kpi-card">
            <div class="v39-kpi-top">
              <div class="v39-kpi-icon v39-icon-orange">□</div>
              <div><div class="v39-kpi-label">Sin Stock</div><div class="v39-kpi-value">{unavailable:,}</div></div>
            </div>
            <div class="v39-kpi-foot">◷ &nbsp; Sin disponibilidad</div>
          </div>
        </div>
        """)

        incoming_without_stock = int(
            (
                (pd.to_numeric(consolidated_full["Disponible"], errors="coerce").fillna(0) <= 0)
                & (pd.to_numeric(consolidated_full["Por llegar"], errors="coerce").fillna(0) > 0)
            ).sum()
        )

        render_html(f"""
        <div class="v39-alert">
          <div class="v39-alert-left">
            <div class="v39-alert-icon">!</div>
            <div>
              <div class="v39-alert-title">Requiere atención · {attention_total:,} incidencias operativas</div>
              <div class="v39-alert-items">
                <span><strong>{products_negative:,}</strong> negativos</span>
                <span><strong>{products_zero:,}</strong> sin stock</span>
                <span><strong>{products_low:,}</strong> stock bajo</span>
                <span><strong>{risk_dispatch:,}</strong> riesgo por despacho</span>
                <span><strong>{incoming_without_stock:,}</strong> con reposición en camino</span>
              </div>
            </div>
          </div>
        </div>
        """)

        # ----------------------------------------------------
        # ANALYTICS: DONUT + HISTORY
        # ----------------------------------------------------
        total_status = max(unique_products, 1)
        p_av = count_available / total_status * 100
        p_low = products_low / total_status * 100
        p_risk = risk_dispatch / total_status * 100
        p_bad = unavailable / total_status * 100

        a1 = p_av
        a2 = a1 + p_low
        a3 = a2 + p_risk
        a4 = a3 + p_bad

        analytics_left, analytics_right = st.columns([1, 1.22], gap="medium")

        with analytics_left:
            render_html(f"""
            <div class="v39-analytics-card">
              <div class="v39-card-title">Semáforo de disponibilidad <small>ⓘ</small></div>
              <div class="v39-donut-wrap">
                <div class="v39-donut" style="background:conic-gradient(
                  #2dbf72 0% {a1:.2f}%,
                  #f5b81b {a1:.2f}% {a2:.2f}%,
                  #ff7a45 {a2:.2f}% {a3:.2f}%,
                  #ef4444 {a3:.2f}% {a4:.2f}%,
                  #3b82f6 {a4:.2f}% 100%
                )">
                  <div class="v39-donut-center">
                    <strong>{unique_products:,}</strong><span>SKU totales</span>
                  </div>
                </div>
                <div class="v39-legend">
                  <div class="v39-legend-row"><span class="v39-legend-name"><i class="v39-dot" style="background:#2dbf72"></i>Disponible</span><strong>{count_available:,} ({count_available/total_status*100:.1f}%)</strong></div>
                  <div class="v39-legend-row"><span class="v39-legend-name"><i class="v39-dot" style="background:#f5b81b"></i>Stock bajo</span><strong>{products_low:,} ({products_low/total_status*100:.1f}%)</strong></div>
                  <div class="v39-legend-row"><span class="v39-legend-name"><i class="v39-dot" style="background:#ff7a45"></i>Riesgo despacho</span><strong>{risk_dispatch:,} ({risk_dispatch/total_status*100:.1f}%)</strong></div>
                  <div class="v39-legend-row"><span class="v39-legend-name"><i class="v39-dot" style="background:#ef4444"></i>Sin stock / negativo</span><strong>{unavailable:,} ({unavailable/total_status*100:.1f}%)</strong></div>
                  <div class="v39-legend-row"><span class="v39-legend-name"><i class="v39-dot" style="background:#3b82f6"></i>Por llegar</span><strong>{count_incoming:,} ({count_incoming/total_status*100:.1f}%)</strong></div>
                </div>
              </div>
            </div>
            """)

        # Historial real de cargas dentro de la sesión
        history = st.session_state.get("inventory_history_v39", [])
        snapshot_key = f"{shared_stock_name}|{loaded_label}|{total_available}"
        if not history or history[-1].get("key") != snapshot_key:
            history.append({
                "key": snapshot_key,
                "Fecha": loaded_time or datetime.now(),
                "Unidades": total_available,
            })
            history = history[-7:]
            st.session_state.inventory_history_v39 = history

        history_df = pd.DataFrame(history)
        if not history_df.empty:
            history_df["Fecha"] = pd.to_datetime(history_df["Fecha"])

        with analytics_right:
            render_html('<div class="v39-analytics-card"><div class="v39-card-title">Evolución de unidades <small>(historial de cargas)</small></div>')
            if len(history_df) >= 2:
                chart = (
                    alt.Chart(history_df)
                    .mark_area(
                        line={"color": "#6536f3", "strokeWidth": 2.5},
                        color=alt.Gradient(
                            gradient="linear",
                            stops=[
                                alt.GradientStop(color="#6536f3", offset=0),
                                alt.GradientStop(color="#ffffff", offset=1),
                            ],
                            x1=1, x2=1, y1=1, y2=0,
                        ),
                        opacity=0.28,
                        point={"filled": True, "size": 55, "color": "#6536f3"},
                    )
                    .encode(
                        x=alt.X("Fecha:T", title=None, axis=alt.Axis(format="%d %b", labelColor="#747e89", grid=False)),
                        y=alt.Y("Unidades:Q", title=None, scale=alt.Scale(zero=False), axis=alt.Axis(labelColor="#747e89", gridColor="#eef0f3")),
                        tooltip=[alt.Tooltip("Fecha:T", format="%d/%m/%Y %H:%M"), alt.Tooltip("Unidades:Q", format=",")],
                    )
                    .properties(height=195)
                )
                st.altair_chart(chart, use_container_width=True)
            else:
                st.metric("Unidades actuales", f"{total_available:,}")
                st.caption("La evolución se irá construyendo con las siguientes cargas de Stock General.")
            render_html("</div>")

        # ----------------------------------------------------
        # FILTER CARD
        # ----------------------------------------------------
        render_html('<div class="v39-filter-card"><div class="v39-section-line"><div class="v39-view-label">Vista y filtros</div></div>')

        mode = st.radio(
            "Vista",
            ["Consolidado por producto", "Por bodega"],
            horizontal=True,
            index=0,
            key="inventory_view_mode",
        )

        search_col, warehouse_col, family_col = st.columns([1.55, 1, 1])

        with search_col:
            inventory_search = st.text_input(
                "Buscar",
                placeholder="Buscar producto, código, SKU…",
                key="inventory_search",
            )

        with warehouse_col:
            all_warehouses = sorted(
                inventory["Bodega"].replace("", pd.NA).dropna().astype(str).unique().tolist()
            )
            selected_warehouses = st.multiselect(
                "Bodega",
                options=all_warehouses,
                placeholder="Todas las bodegas",
                key="inventory_warehouse_filter",
            )

        with family_col:
            all_families = sorted(
                inventory["Familia"].replace("", pd.NA).dropna().astype(str).unique().tolist()
            )
            selected_families = st.multiselect(
                "Familia",
                options=all_families,
                placeholder="Todas las familias",
                key="inventory_family_filter",
            )

        subfamily_col, status_col = st.columns([1, 1.65])

        with subfamily_col:
            all_subfamilies = sorted(
                inventory["Subfamilia"].replace("", pd.NA).dropna().astype(str).unique().tolist()
            )
            selected_subfamilies = st.multiselect(
                "Subfamilia",
                options=all_subfamilies,
                placeholder="Todas las subfamilias",
                key="inventory_subfamily_filter",
            )

        with status_col:
            stock_status = st.radio(
                "Estado",
                ["Todos", "Disponible", "Stock bajo", "Sin stock", "Negativo", "Riesgo despacho", "Por llegar"],
                horizontal=True,
                key="inventory_status_filter",
            )

        render_html("</div>")

        # ----------------------------------------------------
        # FILTER DATA
        # ----------------------------------------------------
        filtered = inventory.copy()

        if inventory_search:
            term = inventory_search.strip().lower()
            filtered = filtered[
                filtered["Código"].fillna("").astype(str).str.lower().str.contains(term, regex=False)
                | filtered["Producto"].fillna("").astype(str).str.lower().str.contains(term, regex=False)
            ]

        if selected_warehouses:
            filtered = filtered[filtered["Bodega"].isin(selected_warehouses)]
        if selected_families:
            filtered = filtered[filtered["Familia"].isin(selected_families)]
        if selected_subfamilies:
            filtered = filtered[filtered["Subfamilia"].isin(selected_subfamilies)]

        status_map = {
            "Disponible": "🟢 Disponible",
            "Stock bajo": "🟡 Stock bajo",
            "Sin stock": "🔴 Sin stock",
            "Negativo": "🔴 Negativo",
            "Riesgo despacho": "🟠 Riesgo despacho",
            "Por llegar": "🔵 Por llegar",
        }

        if mode == "Consolidado por producto":
            display_df = add_inventory_status(consolidate_inventory_by_product(filtered))
            column_order = [
                "Estado", "Código", "Producto", "Familia", "Subfamilia",
                "Bodegas", "Disponible", "Por llegar", "Por despachar",
                "Stock físico", "Precio",
            ]
            column_config = {
                "Estado": st.column_config.TextColumn("Estado", width="medium"),
                "Código": st.column_config.TextColumn("Código / Producto", width="medium"),
                "Producto": st.column_config.TextColumn("Descripción", width="large"),
                "Familia": st.column_config.TextColumn("Familia", width="medium"),
                "Subfamilia": st.column_config.TextColumn("Subfamilia", width="medium"),
                "Bodegas": st.column_config.NumberColumn("Bodegas", format="%d"),
                "Disponible": st.column_config.NumberColumn("Stock Disponible", format="%d"),
                "Por llegar": st.column_config.NumberColumn("Por llegar", format="%d"),
                "Por despachar": st.column_config.NumberColumn("Por despachar", format="%d"),
                "Stock físico": st.column_config.NumberColumn("Unidades", format="%d"),
                "Precio": st.column_config.NumberColumn("Precio", format="$%d"),
            }
        else:
            display_df = add_inventory_status(filtered)
            column_order = [
                "Estado", "Código", "Producto", "Bodega", "Familia", "Subfamilia",
                "Disponible", "Por llegar", "Por despachar", "Stock físico", "Precio",
            ]
            column_config = {
                "Estado": st.column_config.TextColumn("Estado", width="medium"),
                "Código": st.column_config.TextColumn("Código / Producto", width="medium"),
                "Producto": st.column_config.TextColumn("Descripción", width="large"),
                "Bodega": st.column_config.TextColumn("Bodega", width="medium"),
                "Familia": st.column_config.TextColumn("Familia", width="medium"),
                "Subfamilia": st.column_config.TextColumn("Subfamilia", width="medium"),
                "Disponible": st.column_config.NumberColumn("Stock Disponible", format="%d"),
                "Por llegar": st.column_config.NumberColumn("Por llegar", format="%d"),
                "Por despachar": st.column_config.NumberColumn("Por despachar", format="%d"),
                "Stock físico": st.column_config.NumberColumn("Unidades", format="%d"),
                "Precio": st.column_config.NumberColumn("Precio", format="$%d"),
            }

        if stock_status != "Todos":
            display_df = display_df[display_df["Estado"] == status_map[stock_status]]

        for col in ["Stock físico", "Disponible", "Por llegar", "Por despachar", "Precio", "Bodegas"]:
            if col in display_df.columns:
                display_df[col] = pd.to_numeric(display_df[col], errors="coerce").astype("Int64")

        result_rows = len(display_df)
        result_products = display_df["Código"].nunique() if not display_df.empty else 0
        result_available = int(pd.to_numeric(display_df["Disponible"], errors="coerce").fillna(0).sum()) if "Disponible" in display_df.columns else 0

        export_bytes = dataframe_to_excel_bytes(
            display_df[column_order].copy(),
            sheet_name="Stock_Filtrado",
        )

        summary_col, export_col = st.columns([4.6, 1])
        with summary_col:
            render_html(f"""
            <div class="inventory-result-bar" style="margin-top:14px">
                <span>Mostrando <strong>{result_rows:,}</strong> registros · <strong>{result_products:,}</strong> productos</span>
                <span>Disponible filtrado: <strong>{result_available:,}</strong> uds</span>
            </div>
            """)
        with export_col:
            st.write("")
            st.download_button(
                "⬇ Exportar",
                data=export_bytes,
                file_name="Stock_General_Filtrado.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        st.dataframe(
            display_df,
            column_order=column_order,
            column_config=column_config,
            hide_index=True,
            use_container_width=True,
            height=620,
        )

        # ----------------------------------------------------
        # PRODUCT DETAIL
        # ----------------------------------------------------
        available_codes = display_df["Código"].dropna().astype(str).drop_duplicates().tolist()

        if available_codes:
            st.markdown("### Detalle por producto")
            selected_code = st.selectbox(
                "Seleccionar SKU",
                options=available_codes,
                key="inventory_detail_sku",
            )

            sku_rows = inventory[inventory["Código"] == selected_code].copy()

            if not sku_rows.empty:
                sku_name = sku_rows["Producto"].iloc[0]
                sku_available = int(sku_rows["Disponible"].sum())
                sku_physical = int(sku_rows["Stock físico"].sum())
                sku_incoming = int(sku_rows["Por llegar"].sum())
                sku_outgoing = int(sku_rows["Por despachar"].sum())

                render_html(f"""
                <div class="product-detail">
                    <div class="product-detail-title">{sku_name}</div>
                    <div class="product-detail-code">SKU {selected_code}</div>
                    <div class="product-detail-grid">
                        <div class="product-mini-kpi"><div class="product-mini-label">Disponible</div><div class="product-mini-value">{sku_available:,}</div></div>
                        <div class="product-mini-kpi"><div class="product-mini-label">Stock físico</div><div class="product-mini-value">{sku_physical:,}</div></div>
                        <div class="product-mini-kpi"><div class="product-mini-label">Por llegar</div><div class="product-mini-value">{sku_incoming:,}</div></div>
                        <div class="product-mini-kpi"><div class="product-mini-label">Por despachar</div><div class="product-mini-value">{sku_outgoing:,}</div></div>
                    </div>
                </div>
                """)

                sku_detail = sku_rows[["Bodega", "Stock físico", "Disponible", "Por llegar", "Por despachar"]].copy()
                for col in ["Stock físico", "Disponible", "Por llegar", "Por despachar"]:
                    sku_detail[col] = pd.to_numeric(sku_detail[col], errors="coerce").astype("Int64")

                st.dataframe(
                    sku_detail,
                    hide_index=True,
                    use_container_width=True,
                    height=300,
                )

    except Exception as exc:
        show_error(exc, "Error cargando Stock General")

    st.stop()


if page == "ventas":
    # ========================================================
    # RESUMEN EJECUTIVO · V56
    # Performance comercial + meta + proyección
    # ========================================================
    render_html("""
    <div class="exec-head">
        <div>
            <div class="exec-eyebrow">MARITEX ADMIN / RESUMEN</div>
            <div class="exec-title">Resumen Ejecutivo</div>
            <div class="exec-subtitle">
                Performance comercial, cumplimiento de meta y proyección de cierre.
            </div>
        </div>
        <div class="exec-rule">Selecciona análisis con IVA o sin IVA</div>
    </div>
    """)

    ensure_sales_source_loaded()
    metrics_raw = st.session_state.get("metrics_bytes")
    metrics_name = st.session_state.get("metrics_name")

    if metrics_raw is None:
        st.info(
            "No existe una fuente ERP Ventas activa. "
            "Cárgala desde Plantillas → Fuentes ERP → ERP Ventas."
        )
        st.stop()

    try:
        exec_df = read_metrics_data(
            metrics_raw,
            metrics_name or "erp_ventas.csv",
        )
        exec_df = ensure_sales_amount_column(exec_df)
        exec_all = filter_commercial_documents(exec_df)
    except Exception as exc:
        show_error(exc, "Error leyendo ERP Ventas para Resumen Ejecutivo")
        st.stop()

    if exec_all.empty:
        st.warning("No se encontraron Facturas, Boletas o Notas de crédito.")
        st.stop()

    if (
        "Fecha_dt" not in exec_all.columns
        or exec_all["Fecha_dt"].isna().all()
    ):
        st.warning("ERP Ventas no contiene fechas válidas para proyectar.")
        st.stop()

    data_min = exec_all["Fecha_dt"].min().date()
    data_max = exec_all["Fecha_dt"].max().date()
    today_date = datetime.now().date()

    # V57: meses reales presentes en ERP Ventas.
    exec_months = available_sales_months(exec_all)
    if not exec_months:
        st.warning("ERP Ventas no contiene meses válidos.")
        st.stop()

    exec_month_labels = [month_label_es(m) for m in exec_months]
    exec_month_map = dict(zip(exec_month_labels, exec_months))

    seller_options = sorted(
        exec_all["Vendedor"]
        .fillna("Sin vendedor")
        .astype(str)
        .str.strip()
        .unique()
        .tolist()
    ) if "Vendedor" in exec_all.columns else []

    warehouse_options = sorted(
        exec_all["Bodega"]
        .fillna("Sin bodega")
        .astype(str)
        .str.strip()
        .unique()
        .tolist()
    ) if "Bodega" in exec_all.columns else []

    type_options = sorted(
        exec_all["TipoDocto"]
        .dropna()
        .astype(str)
        .unique()
        .tolist()
    ) if "TipoDocto" in exec_all.columns else []

    # --------------------------------------------------------
    # FILTROS
    # --------------------------------------------------------
    render_html("""
    <div class="exec-section-label">Filtros de performance</div>
    """)

    ef1, ef2, ef3, ef4, ef5 = st.columns([1.05, 1.1, .9, 1.05, 1.05], gap="small")

    with ef1:
        selected_exec_month_label = st.selectbox(
            "Mes",
            exec_month_labels,
            index=0,
            key="exec_month_filter_v59",
            help="Selecciona el mes base para analizar y proyectar.",
        )
        selected_exec_month = exec_month_map[selected_exec_month_label]

        exec_month_start, exec_month_end = month_bounds(selected_exec_month)

        # Para el mes actual, por defecto llegar hasta hoy.
        # Para meses históricos, por defecto usar el mes completo.
        selected_period_obj = pd.Period(selected_exec_month, freq="M")
        today_period_obj = pd.Period(pd.Timestamp(today_date), freq="M")

        if selected_period_obj == today_period_obj:
            exec_default_end = min(exec_month_end, today_date)
        else:
            exec_default_end = exec_month_end

        exec_day_range = st.date_input(
            "Días",
            value=(exec_month_start, exec_default_end),
            min_value=exec_month_start,
            max_value=exec_month_end,
            key=f"exec_day_range_v59_{selected_exec_month}",
            help=(
                "Selecciona los días que deseas analizar. "
                "Si el ERP termina antes del último día seleccionado, "
                "la diferencia se usa para proyectar."
            ),
        )

    with ef2:
        exec_sellers = st.multiselect(
            "Vendedor",
            seller_options,
            placeholder="Todos los vendedores",
            key="exec_seller_filter_v56",
        )

    with ef3:
        exec_warehouses = st.multiselect(
            "Bodega",
            warehouse_options,
            placeholder="Todas",
            key="exec_warehouse_filter_v56",
            disabled=not bool(warehouse_options),
        )

    with ef4:
        exec_types = st.multiselect(
            "Tipo de documento",
            type_options,
            placeholder="Facturas + Boletas + NC",
            key="exec_type_filter_v56",
        )

    with ef5:
        analysis_base = st.selectbox(
            "Base de análisis",
            ["Venta final con IVA", "Venta final sin IVA"],
            index=0,
            key="exec_analysis_base_v60",
            help=(
                "Define la base monetaria usada por KPIs, meta, proyección, "
                "cumplimiento, evolución mensual y performance por vendedor."
            ),
        )

    em1, em2 = st.columns([1, 1], gap="small")

    with em1:
        exec_goal = st.number_input(
            "Meta de venta (según base seleccionada)",
            min_value=0,
            value=int(st.session_state.get("exec_sales_goal", 100000000)),
            step=100000,
            key="exec_sales_goal_v56",
            help="Meta del período seleccionado, expresada con IVA.",
        )
        st.session_state["exec_sales_goal"] = int(exec_goal)

    with em2:
        projection_mode = st.selectbox(
            "Método de proyección",
            [
                "Ritmo diario del período",
                "Ritmo por días hábiles (lun-vie)",
            ],
            key="exec_projection_mode_v56",
        )

    selected_start = exec_month_start
    selected_end = exec_default_end

    if (
        exec_day_range
        and isinstance(exec_day_range, (tuple, list))
        and len(exec_day_range) == 2
    ):
        selected_start, selected_end = exec_day_range

    # --------------------------------------------------------
    # APLICAR FILTROS NO TEMPORALES PRIMERO
    # --------------------------------------------------------
    exec_filtered = exec_all.copy()

    if exec_sellers:
        exec_filtered = exec_filtered[
            exec_filtered["Vendedor"]
            .fillna("Sin vendedor")
            .astype(str)
            .str.strip()
            .isin(exec_sellers)
        ]

    if exec_warehouses and "Bodega" in exec_filtered.columns:
        exec_filtered = exec_filtered[
            exec_filtered["Bodega"]
            .fillna("Sin bodega")
            .astype(str)
            .str.strip()
            .isin(exec_warehouses)
        ]

    if exec_types:
        exec_filtered = exec_filtered[
            exec_filtered["TipoDocto"].isin(exec_types)
        ]

    # Datos reales del mes seleccionado.
    month_available = exec_filtered[
        (exec_filtered["Fecha_dt"].dt.date >= selected_start)
        & (exec_filtered["Fecha_dt"].dt.date <= selected_end)
    ].copy()

    if not month_available.empty and month_available["Fecha_dt"].notna().any():
        actual_end = month_available["Fecha_dt"].max().date()
    else:
        actual_end = selected_start

    actual_view = month_available[
        month_available["Fecha_dt"].dt.date <= actual_end
    ].copy()

    totals = calculate_commercial_totals(actual_view, vat_rate=0.19)

    # V60: base global del Resumen Ejecutivo.
    use_no_vat = analysis_base == "Venta final sin IVA"

    actual_sales_with_vat = float(totals["venta_neta_con_iva"])
    actual_sales_without_vat = float(totals["venta_neta_sin_iva"])
    gross_sales_with_vat = float(totals["ventas_brutas_con_iva"])
    gross_sales_without_vat = float(totals["ventas_brutas_sin_iva"])
    credits_with_vat = float(totals["notas_credito_con_iva"])
    credits_without_vat = float(totals["notas_credito_sin_iva"])

    actual_sales = (
        actual_sales_without_vat if use_no_vat else actual_sales_with_vat
    )
    gross_sales = (
        gross_sales_without_vat if use_no_vat else gross_sales_with_vat
    )
    credits = (
        credits_without_vat if use_no_vat else credits_with_vat
    )

    opposite_sales = (
        actual_sales_with_vat if use_no_vat else actual_sales_without_vat
    )
    selected_base_short = "Sin IVA" if use_no_vat else "Con IVA"
    opposite_base_short = "Con IVA" if use_no_vat else "Sin IVA"

    # --------------------------------------------------------
    # PROYECCIÓN
    # --------------------------------------------------------
    def _calendar_days(start_date, end_date):
        if start_date is None or end_date is None or end_date < start_date:
            return 0
        return (pd.Timestamp(end_date) - pd.Timestamp(start_date)).days + 1

    def _business_days(start_date, end_date):
        if start_date is None or end_date is None or end_date < start_date:
            return 0
        return len(pd.bdate_range(start=start_date, end=end_date))

    if projection_mode == "Ritmo por días hábiles (lun-vie)":
        elapsed_units = _business_days(selected_start, actual_end)
        target_units = _business_days(selected_start, selected_end)
        projection_unit_label = "día hábil"
    else:
        elapsed_units = _calendar_days(selected_start, actual_end)
        target_units = _calendar_days(selected_start, selected_end)
        projection_unit_label = "día"

    daily_run_rate = (
        actual_sales / elapsed_units
        if elapsed_units > 0
        else 0.0
    )

    if actual_view.empty:
        projected_sales = 0.0
    elif selected_end <= actual_end:
        projected_sales = actual_sales
    else:
        projected_sales = daily_run_rate * target_units

    goal_value = float(exec_goal)
    current_compliance = (
        actual_sales / goal_value * 100
        if goal_value > 0
        else 0.0
    )
    projected_compliance = (
        projected_sales / goal_value * 100
        if goal_value > 0
        else 0.0
    )
    current_missing = max(goal_value - actual_sales, 0.0)
    projected_gap = goal_value - projected_sales

    # Ticket promedio.
    sales_only = actual_view[
        actual_view["Grupo comercial"].isin(["Factura", "Boleta"])
    ].copy()
    document_count = (
        int(sales_only["Numero"].nunique())
        if "Numero" in sales_only.columns
        else int(len(sales_only))
    )
    avg_ticket = gross_sales / document_count if document_count else 0.0

    # Comparación período anterior.
    comparison_source = exec_filtered[
        exec_filtered["Grupo comercial"].isin(["Factura", "Boleta", "Nota de crédito"])
    ].copy()

    period_length = _calendar_days(selected_start, actual_end)
    prev_end = pd.Timestamp(selected_start) - pd.Timedelta(days=1)
    prev_start = prev_end - pd.Timedelta(days=max(period_length - 1, 0))

    previous_view = comparison_source[
        (comparison_source["Fecha_dt"] >= prev_start.normalize())
        & (
            comparison_source["Fecha_dt"]
            <= prev_end.normalize() + pd.Timedelta(days=1) - pd.Timedelta(seconds=1)
        )
    ].copy()

    previous_totals = calculate_commercial_totals(previous_view, vat_rate=0.19)
    previous_sales = float(
        previous_totals["venta_neta_sin_iva"]
        if use_no_vat
        else previous_totals["venta_neta_con_iva"]
    )
    variation_pct = (
        ((actual_sales - previous_sales) / previous_sales) * 100
        if previous_sales != 0
        else (100.0 if actual_sales > 0 else 0.0)
    )

    # --------------------------------------------------------
    # ESTADO DE DATOS
    # --------------------------------------------------------
    data_coverage_pct = (
        min(elapsed_units / target_units * 100, 100.0)
        if target_units > 0
        else 0.0
    )

    render_html(f"""
    <div class="exec-data-strip">
        <div>
            <span>Datos ERP hasta</span>
            <strong>{pd.Timestamp(actual_end).strftime("%d/%m/%Y")}</strong>
        </div>
        <div>
            <span>Mes analizado</span>
            <strong>{selected_exec_month_label}</strong>
        </div>
        <div>
            <span>Días seleccionados</span>
            <strong>{pd.Timestamp(selected_start).strftime("%d/%m")} – {pd.Timestamp(selected_end).strftime("%d/%m/%Y")}</strong>
        </div>
        <div>
            <span>Cobertura temporal</span>
            <strong>{data_coverage_pct:.1f}%</strong>
        </div>
        <div class="exec-data-note">
            {
                "El ERP aún no cubre todo el período; la venta faltante se proyecta con el ritmo observado."
                if actual_end < selected_end
                else "El ERP cubre completamente el período seleccionado."
            }
        </div>
    </div>
    """)

    # --------------------------------------------------------
    # KPI
    # --------------------------------------------------------
    render_html(f"""
    <div class="exec-kpi-grid">
        <div class="exec-kpi">
            <span>Venta actual · {selected_base_short}</span>
            <strong>{format_clp(actual_sales)}</strong>
            <small>
                {opposite_base_short}: {format_clp(opposite_sales)} ·
                hasta {pd.Timestamp(actual_end).strftime("%d/%m/%Y")}
            </small>
        </div>
        <div class="exec-kpi">
            <span>Meta · {selected_base_short}</span>
            <strong>{format_clp(goal_value)}</strong>
            <small>{current_compliance:.1f}% cumplido</small>
        </div>
        <div class="exec-kpi exec-kpi-focus">
            <span>Proyección cierre · {selected_base_short}</span>
            <strong>{format_clp(projected_sales)}</strong>
            <small>{projected_compliance:.1f}% de la meta</small>
        </div>
        <div class="exec-kpi">
            <span>Variación</span>
            <strong>{variation_pct:+.1f}%</strong>
            <small>vs. período anterior equivalente</small>
        </div>
        <div class="exec-kpi">
            <span>Ticket promedio</span>
            <strong>{format_clp(avg_ticket)}</strong>
            <small>{document_count:,} documentos de venta</small>
        </div>
    </div>
    """)

    # --------------------------------------------------------
    # META Y PROYECCIÓN
    # --------------------------------------------------------
    goal_bar_current = (
        min(max(current_compliance, 0), 100)
        if goal_value > 0 else 0
    )
    goal_bar_projected = (
        min(max(projected_compliance, 0), 100)
        if goal_value > 0 else 0
    )

    projection_message = (
        f"Proyección supera la meta en {format_clp(abs(projected_gap))}"
        if projected_gap < 0
        else f"Proyección queda bajo la meta en {format_clp(projected_gap)}"
    )

    render_html(f"""
    <div class="exec-goal-card">
        <div class="exec-goal-head">
            <div>
                <div class="exec-card-title">Cumplimiento y proyección · {selected_base_short}</div>
                <div class="exec-card-subtitle">
                    Ritmo promedio: {format_clp(daily_run_rate)} por {projection_unit_label}.
                    Meta y proyección calculadas en la base seleccionada.
                </div>
            </div>
            <div class="exec-projection-state">
                {projection_message}
            </div>
        </div>

        <div class="exec-progress-label">
            <span>Venta actual</span><strong>{current_compliance:.1f}%</strong>
        </div>
        <div class="exec-progress-track">
            <div class="exec-progress-fill current" style="width:{goal_bar_current:.2f}%"></div>
        </div>

        <div class="exec-progress-label exec-progress-label-second">
            <span>Proyección al cierre</span><strong>{projected_compliance:.1f}%</strong>
        </div>
        <div class="exec-progress-track">
            <div class="exec-progress-fill projected" style="width:{goal_bar_projected:.2f}%"></div>
        </div>

        <div class="exec-goal-foot">
            <span>Faltante actual: <strong>{format_clp(current_missing)}</strong></span>
            <span>Venta bruta: <strong>{format_clp(gross_sales)}</strong></span>
            <span>NC: <strong>{format_clp(credits)}</strong></span>
        </div>
    </div>
    """)

    # --------------------------------------------------------
    # EVOLUCIÓN MENSUAL
    # --------------------------------------------------------
    monthly_source = exec_filtered.copy()
    monthly_source["_Mes"] = monthly_source["Fecha_dt"].dt.to_period("M")
    monthly_source["_VentaBase"] = pd.to_numeric(
        monthly_source["VentaMonto_num"],
        errors="coerce",
    ).fillna(0.0)

    if use_no_vat:
        monthly_source["_VentaBase"] = monthly_source["_VentaBase"] / 1.19

    monthly_source["_VentaFirmada"] = monthly_source.apply(
        lambda r: (
            -abs(float(r.get("_VentaBase", 0) or 0))
            if r.get("Grupo comercial") == "Nota de crédito"
            else float(r.get("_VentaBase", 0) or 0)
        ),
        axis=1,
    )

    monthly_sales = (
        monthly_source.dropna(subset=["Fecha_dt"])
        .groupby("_Mes", as_index=False)
        .agg(Venta=("_VentaFirmada", "sum"))
        .sort_values("_Mes")
    )

    if not monthly_sales.empty:
        monthly_sales["Mes"] = monthly_sales["_Mes"].apply(month_label_es)

        render_html("""
        <div class="exec-section-head">
            <div>
                <div class="exec-card-title">Evolución mensual · {selected_base_short}</div>
                <div class="exec-card-subtitle">
                    Venta neta mensual según la base y filtros comerciales seleccionados.
                </div>
            </div>
        </div>
        """)

        monthly_chart = (
            alt.Chart(monthly_sales)
            .mark_bar(cornerRadiusTopLeft=4, cornerRadiusTopRight=4, color="#4F7CD7")
            .encode(
                x=alt.X(
                    "Mes:N",
                    sort=list(monthly_sales["Mes"]),
                    title=None,
                    axis=alt.Axis(labelAngle=-25),
                ),
                y=alt.Y("Venta:Q", title="Venta neta"),
                tooltip=[
                    alt.Tooltip("Mes:N", title="Mes"),
                    alt.Tooltip("Venta:Q", title="Venta neta", format=","),
                ],
            )
            .properties(height=285)
        )
        st.altair_chart(monthly_chart, use_container_width=True)

    # --------------------------------------------------------
    # PERFORMANCE POR VENDEDOR
    # --------------------------------------------------------
    render_html("""
    <div class="exec-section-head">
        <div>
            <div class="exec-card-title">Performance por vendedor · {selected_base_short}</div>
            <div class="exec-card-subtitle">
                Venta neta, participación, ticket y comparación usando la base seleccionada.
            </div>
        </div>
    </div>
    """)

    seller_perf_source = actual_view.copy()
    if "Vendedor" in seller_perf_source.columns:
        seller_perf_source["Vendedor"] = (
            seller_perf_source["Vendedor"]
            .fillna("Sin vendedor")
            .astype(str)
            .str.strip()
        )
        seller_perf_source["_VentaBase"] = pd.to_numeric(
            seller_perf_source["VentaMonto_num"],
            errors="coerce",
        ).fillna(0.0)

        if use_no_vat:
            seller_perf_source["_VentaBase"] = seller_perf_source["_VentaBase"] / 1.19

        seller_perf_source["_VentaFirmada"] = seller_perf_source.apply(
            lambda r: (
                -abs(float(r.get("_VentaBase", 0) or 0))
                if r.get("Grupo comercial") == "Nota de crédito"
                else float(r.get("_VentaBase", 0) or 0)
            ),
            axis=1,
        )

        number_col = "Numero" if "Numero" in seller_perf_source.columns else "Vendedor"

        seller_perf = (
            seller_perf_source.groupby("Vendedor", as_index=False)
            .agg(
                Venta=("_VentaFirmada", "sum"),
                Documentos=(number_col, "nunique"),
            )
        )

        seller_perf["Participación %"] = (
            seller_perf["Venta"] / actual_sales * 100
            if actual_sales != 0
            else 0.0
        )
        seller_perf["Ticket promedio"] = seller_perf.apply(
            lambda r: (
                r["Venta"] / r["Documentos"]
                if r["Documentos"] else 0.0
            ),
            axis=1,
        )

        # Comparación individual contra período anterior.
        if not previous_view.empty and "Vendedor" in previous_view.columns:
            prev_perf = previous_view.copy()
            prev_perf["Vendedor"] = (
                prev_perf["Vendedor"]
                .fillna("Sin vendedor")
                .astype(str)
                .str.strip()
            )
            prev_perf["_VentaBase"] = pd.to_numeric(
                prev_perf["VentaMonto_num"],
                errors="coerce",
            ).fillna(0.0)

            if use_no_vat:
                prev_perf["_VentaBase"] = prev_perf["_VentaBase"] / 1.19

            prev_perf["_VentaFirmada"] = prev_perf.apply(
                lambda r: (
                    -abs(float(r.get("_VentaBase", 0) or 0))
                    if r.get("Grupo comercial") == "Nota de crédito"
                    else float(r.get("_VentaBase", 0) or 0)
                ),
                axis=1,
            )
            prev_by_seller = (
                prev_perf.groupby("Vendedor", as_index=False)
                .agg(Venta_anterior=("_VentaFirmada", "sum"))
            )
            seller_perf = seller_perf.merge(
                prev_by_seller,
                on="Vendedor",
                how="left",
            )
        else:
            seller_perf["Venta_anterior"] = 0.0

        seller_perf["Venta_anterior"] = seller_perf["Venta_anterior"].fillna(0.0)
        seller_perf["Variación %"] = seller_perf.apply(
            lambda r: (
                ((r["Venta"] - r["Venta_anterior"]) / r["Venta_anterior"]) * 100
                if r["Venta_anterior"] != 0
                else (100.0 if r["Venta"] > 0 else 0.0)
            ),
            axis=1,
        )

        seller_perf = seller_perf.sort_values("Venta", ascending=False)

        st.dataframe(
            seller_perf[
                [
                    "Vendedor",
                    "Venta",
                    "Participación %",
                    "Documentos",
                    "Ticket promedio",
                    "Variación %",
                ]
            ],
            hide_index=True,
            use_container_width=True,
            height=min(520, max(220, len(seller_perf) * 36 + 44)),
            column_config={
                "Venta": st.column_config.NumberColumn(
                    "Venta neta",
                    format="$%d",
                ),
                "Participación %": st.column_config.NumberColumn(
                    "Participación",
                    format="%.1f%%",
                ),
                "Ticket promedio": st.column_config.NumberColumn(
                    "Ticket promedio",
                    format="$%d",
                ),
                "Variación %": st.column_config.NumberColumn(
                    "vs. período anterior",
                    format="%+.1f%%",
                ),
            },
        )

        # Gráfico ranking.
        chart_source = seller_perf.head(12).copy()
        if not chart_source.empty:
            seller_chart = (
                alt.Chart(chart_source)
                .mark_bar(cornerRadiusEnd=4, color="#4F7CD7")
                .encode(
                    y=alt.Y("Vendedor:N", sort="-x", title=None),
                    x=alt.X("Venta:Q", title="Venta neta"),
                    tooltip=[
                        alt.Tooltip("Vendedor:N"),
                        alt.Tooltip("Venta:Q", format=","),
                        alt.Tooltip("Participación %:Q", format=".1f"),
                        alt.Tooltip("Variación %:Q", format=".1f"),
                    ],
                )
                .properties(height=max(260, min(420, len(chart_source) * 32)))
            )
            st.altair_chart(seller_chart, use_container_width=True)
    else:
        st.info("La fuente ERP Ventas no contiene columna Vendedor.")

    st.stop()


# ============================================================
# MARKETPLACES · PLANTILLAS GUARDADAS + DESCARGA DIRECTA
# ============================================================
if page == "marketplace":
    render_html("""
    <div class="inventory-head">
        <div>
            <div class="inventory-head-title">Marketplaces</div>
            <div class="inventory-head-subtitle">
                Publicación y actualización de stock para canales marketplace.
            </div>
        </div>
    </div>
    """)

    with st.container(border=True):
        st.markdown("#### Reglas de publicación")
        rc1, rc2, rc3 = st.columns([1, 1, 1])

        with rc1:
            reserve_stock = st.number_input(
                "Stock de seguridad",
                min_value=0,
                value=int(st.session_state.get("reserve_stock_value", 0)),
                step=1,
                key="reserve_stock_v37",
            )
            st.session_state.reserve_stock_value = int(reserve_stock)

        with rc2:
            use_max_stock = st.checkbox(
                "Limitar stock máximo",
                value=bool(st.session_state.get("use_max_stock_value", False)),
                key="use_max_stock_v37",
            )
            st.session_state.use_max_stock_value = bool(use_max_stock)

        with rc3:
            if use_max_stock:
                max_stock = st.number_input(
                    "Stock máximo",
                    min_value=0,
                    value=int(st.session_state.get("max_stock_value", 50)),
                    step=1,
                    key="max_stock_v37",
                )
                st.session_state.max_stock_value = int(max_stock)
            else:
                max_stock = None
                st.text_input("Stock máximo", value="Sin límite", disabled=True)

    ensure_stock_source_loaded()
    shared_stock_bytes = st.session_state.get("shared_stock_bytes")
    shared_stock_name = st.session_state.get("shared_stock_name")

    if not shared_stock_bytes:
        st.warning(
            "Primero carga el archivo ERP en Stock General. "
            "Después vuelve a Marketplaces."
        )
        st.stop()

    try:
        (
            erp,
            selected_warehouses,
            erp_grouped,
            stock_lookup,
            original_lookup,
            product_name_lookup,
            description_column,
        ) = prepare_marketplace_erp(
            shared_stock_bytes,
            shared_stock_name or "Stock_General.csv",
        )

        render_html(f"""
        <div class="upload-status">
            <strong>✓ Stock ERP listo: {shared_stock_name}</strong><br>
            <small>
                {len(stock_lookup):,} SKU únicos ·
                bodega: {", ".join(selected_warehouses)} ·
                fuente: StockDisponible
            </small>
        </div>
        """)

    except Exception as exc:
        show_error(exc, "Error preparando Stock General")
        st.stop()

    st.markdown("### Archivos listos para marketplace")
    st.caption(
        "Paris y Mercado Libre usan sus plantillas base guardadas automáticamente. "
        "Solo debes cargar Stock General y descargar el archivo que corresponda. "
        "Si el ERP y la plantilla no cambian, el resultado queda cacheado."
    )

    col_paris, col_meli = st.columns(2, gap="large")

    marketplace_cards = [
        ("Paris Marketplace", col_paris, "PARIS"),
        ("Mercado Libre", col_meli, "MERCADO LIBRE"),
    ]

    for marketplace_name, column, short_name in marketplace_cards:
        with column:
            with st.container(border=True):
                st.markdown(f"### {short_name}")

                template_bytes, template_name = get_saved_template(marketplace_name)

                if template_bytes is None:
                    st.warning("Plantilla no guardada.")
                    st.caption("Ve a Configuración y guarda la plantilla una sola vez.")
                    continue

                st.success(f"✓ Plantilla guardada: {template_name}")

                try:
                    with st.spinner(f"Preparando {short_name}…"):
                        output_bytes, processed_result, used_warehouses = (
                            generate_saved_marketplace_output(
                                marketplace_name=marketplace_name,
                                template_bytes=template_bytes,
                                erp_bytes=shared_stock_bytes,
                                erp_name=shared_stock_name or "Stock_General.csv",
                                reserve_stock=int(reserve_stock),
                                max_stock_value=int(max_stock) if max_stock is not None else None,
                            )
                        )

                    total = len(processed_result)
                    found = int(processed_result["StockDisponible ERP"].notna().sum())
                    missing = total - found
                    diff = pd.to_numeric(
                        processed_result["Diferencia"],
                        errors="coerce",
                    ).fillna(0)
                    changed = int((diff != 0).sum())
                    coverage = (found / total * 100) if total else 0

                    m1, m2, m3 = st.columns(3)
                    m1.metric("SKU", f"{total:,}")
                    m2.metric("Cobertura", f"{coverage:.1f}%")
                    m3.metric("Cambios", f"{changed:,}")

                    if missing:
                        st.warning(
                            f"{missing:,} SKU no encontrados en Flexline. "
                            "Se recomienda revisar antes de cargar."
                        )
                    else:
                        st.success("Todos los SKU fueron encontrados.")

                    generated_name = (
                        f"Stock_{marketplace_name.replace(' ', '_')}_"
                        f"{datetime.now().strftime('%Y-%m-%d')}.xlsx"
                    )

                    st.download_button(
                        f"⬇ DESCARGAR {short_name} ACTUALIZADO",
                        data=output_bytes,
                        file_name=generated_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",
                        use_container_width=True,
                        key=f"download_saved_{marketplace_name}",
                    )

                    with st.expander(f"Ver monitor {short_name}"):
                        monitor_cols = [
                            "SKU Marketplace",
                            "Producto",
                            "Código encontrado ERP",
                            "StockDisponible ERP",
                            "Stock actual",
                            "Nuevo stock",
                            "Diferencia",
                            "Estado",
                        ]
                        monitor_cols = [c for c in monitor_cols if c in processed_result.columns]
                        st.dataframe(
                            processed_result[monitor_cols],
                            hide_index=True,
                            use_container_width=True,
                            height=430,
                        )

                except Exception as exc:
                    show_error(exc, f"Error generando {marketplace_name}")

    st.markdown("---")
    st.caption(
        "Flujo: Stock General → Casa Matriz → SKU marketplace → "
        "StockDisponible ERP → archivo actualizado."
    )
    st.stop()
