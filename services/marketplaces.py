from __future__ import annotations

from io import BytesIO
from pathlib import Path
import re

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from analytics.stock_metrics import (
    stock_view,
    consolidate_inventory,
)
from config.settings import MARKETPLACE_TEMPLATES
from ui.components import render_html


# ============================================================
# HELPERS GENERALES
# ============================================================

def _safe_int(value) -> int:
    try:
        if pd.isna(value):
            return 0
        return int(round(float(value)))
    except Exception:
        return 0


def _normalize_sku(value) -> str:
    """
    Normaliza SKU para cruces entre Llegadas_OK y marketplaces.

    Ejemplos:
        13051205  -> 13051205
        1305120-5 -> 13051205
        200122.0  -> 200122
    """
    if value is None:
        return ""

    text = str(value).strip().upper()

    if not text:
        return ""

    text = re.sub(r"\.0$", "", text)
    text = re.sub(r"[^A-Z0-9]", "", text)

    return text


@st.cache_data(
    ttl=300,
    show_spinner=False,
)
def _prepare_house_stock(
    stock: pd.DataFrame,
) -> pd.DataFrame:
    """
    Prepara exclusivamente Casa Matriz desde el stock ya normalizado.

    IMPORTANTE:
    - CD, Patronato y Concepción NO participan.
    - Se cachea 5 minutos para no repetir filtros/groupby en cada rerun.
    """
    if stock is None or stock.empty:
        return pd.DataFrame()

    warehouse = (
        stock["Bodega"]
        .fillna("")
        .astype(str)
        .str.strip()
        .str.upper()
    )

    house = stock[
        warehouse.str.contains(
            "CASA MATRIZ",
            regex=False,
        )
    ].copy()

    if house.empty:
        return pd.DataFrame()

    house = consolidate_inventory(
        house
    )

    house["_sku_match"] = (
        house["Código"]
        .map(_normalize_sku)
    )

    house = house[
        house["_sku_match"].ne("")
    ].copy()

    if house["_sku_match"].duplicated().any():
        agg = {
            "Código": "first",
            "Producto": "first",
            "Disponible": "sum",
        }

        for col in [
            "Stock físico",
            "Por llegar",
            "Por despachar",
        ]:
            if col in house.columns:
                agg[col] = "sum"

        house = (
            house.groupby(
                "_sku_match",
                as_index=False,
            )
            .agg(agg)
        )

    house["_search_codigo"] = (
        house["Código"]
        .fillna("")
        .astype(str)
        .str.lower()
    )

    house["_search_producto"] = (
        house["Producto"]
        .fillna("")
        .astype(str)
        .str.lower()
    )

    return house.reset_index(
        drop=True
    )


def _stock_map(
    house: pd.DataFrame,
    reserve: int,
) -> dict:
    """
    Stock a publicar:
        max(Disponible Casa Matriz - reserva, 0)
    """
    available = pd.to_numeric(
        house["Disponible"],
        errors="coerce",
    ).fillna(0)

    publishable = (
        available
        - int(reserve)
    ).clip(
        lower=0
    ).round().astype(int)

    return dict(
        zip(
            house["_sku_match"],
            publishable,
        )
    )


@st.cache_data(
    show_spinner=False,
)
def _read_template_bytes(
    path_str: str,
    modified_time: float,
) -> bytes:
    """Lee la plantilla solo cuando el archivo cambia físicamente."""
    return Path(path_str).read_bytes()


def _find_header_column(
    ws,
    row_number: int,
    expected_names,
):
    expected = {
        re.sub(
            r"[^A-Z0-9]",
            "",
            str(name).upper(),
        )
        for name in expected_names
    }

    for cell in ws[row_number]:
        value = cell.value

        if value is None:
            continue

        normalized = re.sub(
            r"[^A-Z0-9]",
            "",
            str(value).strip().upper(),
        )

        if normalized in expected:
            return cell.column

    return None


def _change_status(
    current_stock,
    new_stock,
    found: bool,
) -> tuple[str, int]:
    current = _safe_int(current_stock)
    new = _safe_int(new_stock)
    delta = new - current

    if not found:
        return "Sin coincidencia", delta

    if new == current:
        return "Sin cambios", delta

    if new == 0 and current != 0:
        return "Queda en cero", delta

    if delta > 0:
        return "Sube stock", delta

    return "Baja stock", delta


def _stats_from_preview(
    preview: pd.DataFrame,
    base_stats: dict,
) -> dict:
    stats = dict(base_stats)

    if preview is None or preview.empty:
        stats.update(
            {
                "up": 0,
                "down": 0,
                "zero": 0,
                "same": 0,
                "match_pct": 0.0,
            }
        )
        return stats

    status = (
        preview["Cambio"]
        if "Cambio" in preview.columns
        else pd.Series("", index=preview.index)
    )

    stats["up"] = int(
        status.eq("Sube stock").sum()
    )
    stats["down"] = int(
        status.eq("Baja stock").sum()
    )
    stats["zero"] = int(
        status.eq("Queda en cero").sum()
    )
    stats["same"] = int(
        status.eq("Sin cambios").sum()
    )

    rows = max(
        _safe_int(
            stats.get("rows", 0)
        ),
        0,
    )

    matched = max(
        _safe_int(
            stats.get("matched_rows", 0)
        ),
        0,
    )

    stats["match_pct"] = (
        (matched / rows) * 100
        if rows > 0
        else 0.0
    )

    return stats


# ============================================================
# PARIS
# ============================================================

@st.cache_data(
    show_spinner=False,
)
def _build_paris_workbook(
    template_bytes: bytes,
    stock_items: tuple,
):
    """
    Paris:
        Hoja       -> stock
        Cruce      -> sku_seller
        Stock base -> stock
        Escribe    -> nuevo_stock
    """
    stock_lookup = dict(
        stock_items
    )

    wb = load_workbook(
        BytesIO(template_bytes)
    )

    if "stock" not in wb.sheetnames:
        raise ValueError(
            "La plantilla Paris no contiene la hoja 'stock'."
        )

    ws = wb["stock"]

    sku_col = _find_header_column(
        ws,
        1,
        [
            "sku_seller",
            "sku seller",
        ],
    )

    target_col = _find_header_column(
        ws,
        1,
        [
            "nuevo_stock",
            "nuevo stock",
        ],
    )

    current_stock_col = _find_header_column(
        ws,
        1,
        [
            "stock",
        ],
    )

    title_col = _find_header_column(
        ws,
        1,
        [
            "titulo",
            "título",
        ],
    )

    size_col = _find_header_column(
        ws,
        1,
        [
            "talla",
        ],
    )

    mkp_col = _find_header_column(
        ws,
        1,
        [
            "sku_mkp",
            "sku mkp",
        ],
    )

    if sku_col is None:
        raise ValueError(
            "No se encontró la columna 'sku_seller' "
            "en la plantilla Paris."
        )

    if target_col is None:
        raise ValueError(
            "No se encontró la columna 'nuevo_stock' "
            "en la plantilla Paris."
        )

    preview_rows = []
    matched_rows = 0
    unmatched_rows = 0
    publishable_units = 0
    processed_rows = 0

    matched_skus = set()
    unmatched_skus = set()

    for row in range(
        2,
        ws.max_row + 1,
    ):
        raw_sku = ws.cell(
            row=row,
            column=sku_col,
        ).value

        if raw_sku is None:
            continue

        sku_text = str(
            raw_sku
        ).strip()

        if not sku_text:
            continue

        processed_rows += 1

        sku_key = _normalize_sku(
            raw_sku
        )

        found = (
            sku_key in stock_lookup
        )

        new_stock = int(
            stock_lookup.get(
                sku_key,
                0,
            )
        )

        current_stock = (
            ws.cell(
                row=row,
                column=current_stock_col,
            ).value
            if current_stock_col
            else 0
        )

        change, delta = _change_status(
            current_stock,
            new_stock,
            found,
        )

        ws.cell(
            row=row,
            column=target_col,
        ).value = new_stock

        if found:
            matched_rows += 1
            matched_skus.add(
                sku_key
            )
        else:
            unmatched_rows += 1
            unmatched_skus.add(
                sku_key
            )

        publishable_units += new_stock

        preview_rows.append(
            {
                "SKU Marketplace": (
                    ws.cell(
                        row=row,
                        column=mkp_col,
                    ).value
                    if mkp_col
                    else ""
                ),
                "SKU Seller": sku_text,
                "Producto": (
                    ws.cell(
                        row=row,
                        column=title_col,
                    ).value
                    if title_col
                    else ""
                ),
                "Talla": (
                    ws.cell(
                        row=row,
                        column=size_col,
                    ).value
                    if size_col
                    else ""
                ),
                "Stock actual": _safe_int(
                    current_stock
                ),
                "Nuevo stock": new_stock,
                "Diferencia": delta,
                "Cambio": change,
                "Coincidencia Stock CM": (
                    "Encontrado"
                    if found
                    else "Sin coincidencia"
                ),
            }
        )

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    preview = pd.DataFrame(
        preview_rows
    )

    stats = _stats_from_preview(
        preview,
        {
            "rows": processed_rows,
            "matched_rows": matched_rows,
            "unmatched_rows": unmatched_rows,
            "matched_skus": len(matched_skus),
            "unmatched_skus": len(unmatched_skus),
            "units": publishable_units,
        },
    )

    return (
        output.getvalue(),
        preview,
        stats,
    )


# ============================================================
# MERCADO LIBRE
# ============================================================

@st.cache_data(
    show_spinner=False,
)
def _build_meli_workbook(
    template_bytes: bytes,
    stock_items: tuple,
):
    """
    Mercado Libre:
        Hoja       -> Publicaciones
        Cruce      -> SKU
        Stock base -> QUANTITY
        Escribe    -> QUANTITY

    Las filas de publicaciones comienzan en la fila 6.
    """
    stock_lookup = dict(
        stock_items
    )

    wb = load_workbook(
        BytesIO(template_bytes)
    )

    if "Publicaciones" not in wb.sheetnames:
        raise ValueError(
            "La plantilla Mercado Libre no contiene "
            "la hoja 'Publicaciones'."
        )

    ws = wb["Publicaciones"]

    sku_col = _find_header_column(
        ws,
        1,
        ["SKU"],
    )

    quantity_col = _find_header_column(
        ws,
        1,
        ["QUANTITY"],
    )

    title_col = _find_header_column(
        ws,
        1,
        ["TITLE"],
    )

    variation_col = _find_header_column(
        ws,
        1,
        ["VARIATIONS"],
    )

    item_col = _find_header_column(
        ws,
        1,
        ["ITEM_ID"],
    )

    if sku_col is None:
        raise ValueError(
            "No se encontró la columna 'SKU' "
            "en la plantilla Mercado Libre."
        )

    if quantity_col is None:
        raise ValueError(
            "No se encontró la columna 'QUANTITY' "
            "en la plantilla Mercado Libre."
        )

    preview_rows = []
    matched_rows = 0
    unmatched_rows = 0
    publishable_units = 0
    processed_rows = 0

    matched_skus = set()
    unmatched_skus = set()

    last_product_title = ""

    for row in range(
        6,
        ws.max_row + 1,
    ):
        raw_sku = ws.cell(
            row=row,
            column=sku_col,
        ).value

        if raw_sku is None:
            continue

        sku_text = str(
            raw_sku
        ).strip()

        if not sku_text:
            continue

        processed_rows += 1

        sku_key = _normalize_sku(
            raw_sku
        )

        found = (
            sku_key in stock_lookup
        )

        current_stock = ws.cell(
            row=row,
            column=quantity_col,
        ).value

        new_stock = int(
            stock_lookup.get(
                sku_key,
                0,
            )
        )

        change, delta = _change_status(
            current_stock,
            new_stock,
            found,
        )

        # Solo se actualiza QUANTITY.
        ws.cell(
            row=row,
            column=quantity_col,
        ).value = new_stock

        if found:
            matched_rows += 1
            matched_skus.add(
                sku_key
            )
        else:
            unmatched_rows += 1
            unmatched_skus.add(
                sku_key
            )

        publishable_units += new_stock

        raw_title = (
            ws.cell(
                row=row,
                column=title_col,
            ).value
            if title_col
            else ""
        )

        product_title = raw_title or ""

        # MELI puede repetir TITLE mediante fórmula. Guardamos el último
        # título real y evitamos recorrer filas anteriores una por una.
        if title_col:
            if (
                isinstance(raw_title, str)
                and raw_title.strip().startswith("=")
            ):
                product_title = last_product_title
            else:
                candidate = str(raw_title or "").strip()
                if candidate:
                    last_product_title = candidate
                    product_title = candidate

        preview_rows.append(
            {
                "Publicación": (
                    ws.cell(
                        row=row,
                        column=item_col,
                    ).value
                    if item_col
                    else ""
                ),
                "SKU": sku_text,
                "Producto": product_title,
                "Variación": (
                    ws.cell(
                        row=row,
                        column=variation_col,
                    ).value
                    if variation_col
                    else ""
                ),
                "Stock actual": _safe_int(
                    current_stock
                ),
                "Nuevo stock": new_stock,
                "Diferencia": delta,
                "Cambio": change,
                "Coincidencia Stock CM": (
                    "Encontrado"
                    if found
                    else "Sin coincidencia"
                ),
            }
        )

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    preview = pd.DataFrame(
        preview_rows
    )

    stats = _stats_from_preview(
        preview,
        {
            "rows": processed_rows,
            "matched_rows": matched_rows,
            "unmatched_rows": unmatched_rows,
            "matched_skus": len(matched_skus),
            "unmatched_skus": len(unmatched_skus),
            "units": publishable_units,
        },
    )

    return (
        output.getvalue(),
        preview,
        stats,
    )


# ============================================================
# COMPONENTES DE INTERFAZ
# ============================================================

def _marketplace_header(
    name: str,
    filename: str,
    stats: dict,
):
    pct = float(
        stats.get(
            "match_pct",
            0.0,
        )
    )

    badge_class = (
        "ok"
        if pct >= 98
        else "warn"
        if pct >= 90
        else "risk"
    )

    render_html(
        f"""
        <div class="mk2-platform-head">
            <div class="mk2-platform-title">
                <div class="mk2-platform-mark">
                    {"P" if "Paris" in name else "ML"}
                </div>
                <div>
                    <strong>{name}</strong>
                    <span>Plantilla oficial · {filename}</span>
                </div>
            </div>

            <div class="mk2-platform-badges">
                <span class="mk2-badge neutral">CASA MATRIZ</span>
                <span class="mk2-badge {badge_class}">
                    {pct:.1f}% MATCH
                </span>
            </div>
        </div>
        """
    )


def _render_match_health(
    stats: dict,
):
    pct = max(
        0.0,
        min(
            float(
                stats.get(
                    "match_pct",
                    0.0,
                )
            ),
            100.0,
        ),
    )

    render_html(
        f"""
        <div class="mk2-health">
            <div class="mk2-health-head">
                <div>
                    <strong>Coincidencia de SKU</strong>
                    <span>
                        Cruce entre la plantilla y Stock Casa Matriz
                    </span>
                </div>
                <b>{pct:.1f}%</b>
            </div>

            <div class="mk2-progress">
                <i style="width:{pct:.2f}%"></i>
            </div>

            <div class="mk2-health-foot">
                <span>
                    <b>{stats.get("matched_rows", 0):,}</b>
                    encontrados
                </span>
                <span>
                    <b>{stats.get("unmatched_rows", 0):,}</b>
                    sin coincidencia
                </span>
            </div>
        </div>
        """
    )


def _render_change_cards(
    stats: dict,
):
    render_html(
        f"""
        <div class="mk2-change-grid">
            <div class="mk2-change up">
                <span>SUBE STOCK</span>
                <strong>{stats.get("up", 0):,}</strong>
                <small>filas aumentan disponibilidad</small>
            </div>

            <div class="mk2-change down">
                <span>BAJA STOCK</span>
                <strong>{stats.get("down", 0):,}</strong>
                <small>filas reducen disponibilidad</small>
            </div>

            <div class="mk2-change zero">
                <span>QUEDA EN CERO</span>
                <strong>{stats.get("zero", 0):,}</strong>
                <small>requieren atención</small>
            </div>

            <div class="mk2-change same">
                <span>SIN CAMBIOS</span>
                <strong>{stats.get("same", 0):,}</strong>
                <small>mantienen stock publicado</small>
            </div>
        </div>
        """
    )


def _filter_preview(
    preview: pd.DataFrame,
    search: str,
    change_filter: str,
) -> pd.DataFrame:
    view = preview.copy()

    if search:
        term = search.strip().lower()

        text_columns = [
            col
            for col in view.columns
            if (
                view[col].dtype == object
                or pd.api.types.is_string_dtype(
                    view[col]
                )
            )
        ]

        mask = pd.Series(
            False,
            index=view.index,
        )

        for col in text_columns:
            mask = (
                mask
                |
                view[col]
                .fillna("")
                .astype(str)
                .str.lower()
                .str.contains(
                    term,
                    regex=False,
                )
            )

        view = view[
            mask
        ].copy()

    mapping = {
        "Suben": "Sube stock",
        "Bajan": "Baja stock",
        "Quedan en cero": "Queda en cero",
        "Sin cambios": "Sin cambios",
        "Sin coincidencia": "Sin coincidencia",
    }

    target = mapping.get(
        change_filter
    )

    if target == "Sin coincidencia":
        if "Coincidencia Stock CM" in view.columns:
            view = view[
                view[
                    "Coincidencia Stock CM"
                ].eq(
                    "Sin coincidencia"
                )
            ].copy()

    elif target and "Cambio" in view.columns:
        view = view[
            view["Cambio"].eq(
                target
            )
        ].copy()

    return view


# ============================================================
# PANEL MARKETPLACE
# ============================================================

def _render_marketplace_panel(
    name: str,
    key_name: str,
    house: pd.DataFrame,
):
    path = MARKETPLACE_TEMPLATES[
        name
    ]

    if not path.exists():
        render_html(
            f"""
            <div class="mk2-platform-head">
                <div class="mk2-platform-title">
                    <div class="mk2-platform-mark">
                        {"P" if "Paris" in name else "ML"}
                    </div>
                    <div>
                        <strong>{name}</strong>
                        <span>Plantilla pendiente</span>
                    </div>
                </div>
                <span class="mk2-badge risk">SIN PLANTILLA</span>
            </div>
            """
        )

        st.warning(
            f"No existe la plantilla oficial de {name}. "
            "Cárgala desde Plantillas."
        )
        return

    # --------------------------------------------------------
    # CONTROLES
    # --------------------------------------------------------
    c1, c2 = st.columns(
        [1.55, 0.75],
        gap="medium",
    )

    with c1:
        search = st.text_input(
            "Buscar SKU o producto",
            placeholder=(
                "Ej: 13051205, parka, softshell..."
            ),
            key=f"market_search_v2_{key_name}",
        )

    with c2:
        reserve = st.number_input(
            "Stock de seguridad",
            min_value=0,
            value=0,
            step=1,
            help=(
                "Stock a publicar = Stock Casa Matriz - reserva. "
                "Nunca se publican cantidades negativas."
            ),
            key=f"market_reserve_v2_{key_name}",
        )

    lookup = _stock_map(
        house,
        reserve,
    )

    stock_items = tuple(
        sorted(
            lookup.items()
        )
    )

    template_bytes = _read_template_bytes(
        str(path),
        path.stat().st_mtime,
    )

    try:
        if name == "Paris Marketplace":
            (
                output_bytes,
                preview,
                stats,
            ) = _build_paris_workbook(
                template_bytes,
                stock_items,
            )

            download_name = (
                "Paris_stock_actualizado.xlsx"
            )

            button_label = (
                "Descargar Paris actualizado"
            )

        else:
            (
                output_bytes,
                preview,
                stats,
            ) = _build_meli_workbook(
                template_bytes,
                stock_items,
            )

            download_name = (
                "Mercado_Libre_stock_actualizado.xlsx"
            )

            button_label = (
                "Descargar Mercado Libre actualizado"
            )

    except Exception as exc:
        st.error(
            f"No fue posible procesar la plantilla: {exc}"
        )
        return

    _marketplace_header(
        name,
        path.name,
        stats,
    )

    # --------------------------------------------------------
    # KPI CRUCE
    # --------------------------------------------------------
    render_html(
        f"""
        <div class="mk2-summary-grid">
            <div>
                <span>FILAS PLANTILLA</span>
                <strong>{stats["rows"]:,}</strong>
                <small>registros procesados</small>
            </div>

            <div>
                <span>SKU ENCONTRADOS</span>
                <strong>{stats["matched_rows"]:,}</strong>
                <small>coinciden con Casa Matriz</small>
            </div>

            <div>
                <span>SIN COINCIDENCIA</span>
                <strong>{stats["unmatched_rows"]:,}</strong>
                <small>se exportarán con stock 0</small>
            </div>

            <div>
                <span>STOCK A PUBLICAR</span>
                <strong>{stats["units"]:,}</strong>
                <small>unidades después de reserva</small>
            </div>
        </div>
        """
    )

    _render_match_health(
        stats
    )

    render_html(
        """
        <div class="mk2-section-head">
            <div>
                <strong>Cambios a publicar</strong>
                <span>
                    Comparación entre stock actual de la plantilla
                    y stock nuevo de Casa Matriz
                </span>
            </div>
        </div>
        """
    )

    _render_change_cards(
        stats
    )

    # --------------------------------------------------------
    # FILTRO
    # --------------------------------------------------------
    change_filter = st.radio(
        "Mostrar",
        [
            "Todos",
            "Suben",
            "Bajan",
            "Quedan en cero",
            "Sin cambios",
            "Sin coincidencia",
        ],
        horizontal=True,
        key=f"market_change_filter_v2_{key_name}",
    )

    preview_view = _filter_preview(
        preview,
        search,
        change_filter,
    )

    render_html(
        f"""
        <div class="mk2-table-head">
            <div>
                <strong>Detalle de publicaciones</strong>
                <span>
                    {len(preview_view):,} registros visibles
                </span>
            </div>
            <div>
                Stock de seguridad:
                <b>{int(reserve)}</b>
            </div>
        </div>
        """
    )

    column_config = {
        "Stock actual": st.column_config.NumberColumn(
            "Stock actual",
            format="%d",
        ),
        "Nuevo stock": st.column_config.NumberColumn(
            "Nuevo stock",
            format="%d",
        ),
        "Diferencia": st.column_config.NumberColumn(
            "Diferencia",
            format="%+d",
        ),
        "Cambio": st.column_config.TextColumn(
            "Cambio",
            width="medium",
        ),
        "Coincidencia Stock CM": st.column_config.TextColumn(
            "Coincidencia",
            width="medium",
        ),
    }

    st.dataframe(
        preview_view,
        hide_index=True,
        width="stretch",
        height=430,
        column_config=column_config,
    )

    if stats["unmatched_rows"] > 0:
        st.warning(
            f"{stats['unmatched_rows']:,} fila(s) no tienen "
            "coincidencia con el stock de Casa Matriz de Llegadas_OK. "
            "Esas filas se exportarán con stock 0."
        )

    st.download_button(
        button_label,
        data=output_bytes,
        file_name=download_name,
        mime=(
            "application/"
            "vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        width="stretch",
        type="primary",
        icon=":material/download:",
        key=f"market_download_v2_{key_name}",
    )


# ============================================================
# RENDER PRINCIPAL
# ============================================================

def render(ctx):
    df = ctx.get(
        "stock_df"
    )

    meta = (
        ctx.get(
            "stock_meta"
        )
        or {}
    )

    render_html(
        """
        <div class="mk2-page-head">
            <div>
                <div class="mk2-eyebrow">
                    MARITEX / OPERACIÓN
                </div>
                <div class="mk2-title">
                    Marketplace
                </div>
                <div class="mk2-subtitle">
                    Sincronización de stock para Paris y Mercado Libre
                    utilizando exclusivamente Casa Matriz.
                </div>
            </div>
            <div class="mk2-live">
                <i></i>
                Stock automático
            </div>
        </div>
        """
    )

    if df is None or df.empty:
        st.info(
            "No hay stock disponible desde Llegadas_OK."
        )
        return

    # app.py ya entrega stock_normalized. Reutilizarlo evita volver a
    # ejecutar stock_view() cada vez que se entra a Marketplace.
    normalized_stock = ctx.get(
        "stock_normalized"
    )

    if normalized_stock is None or normalized_stock.empty:
        normalized_stock = stock_view(df)

    house = _prepare_house_stock(
        normalized_stock
    )

    if house.empty:
        st.warning(
            "Llegadas_OK no contiene registros asociados "
            "a Casa Matriz."
        )
        return

    available = pd.to_numeric(
        house["Disponible"],
        errors="coerce",
    ).fillna(0)

    total_skus = int(
        house["_sku_match"]
        .nunique()
    )

    total_available = _safe_int(
        available.sum()
    )

    positive_skus = int(
        (available > 0).sum()
    )

    zero_skus = int(
        (available <= 0).sum()
    )

    source_name = (
        meta.get(
            "filename"
        )
        or "Stock automático · Llegadas_OK"
    )

    loaded_at = (
        meta.get(
            "loaded_at"
        )
        or meta.get(
            "generated_at"
        )
        or "actualización automática"
    )

    render_html(
        f"""
        <div class="mk2-source">
            <div class="mk2-source-icon">CM</div>
            <div class="mk2-source-main">
                <span>FUENTE DE STOCK</span>
                <strong>{source_name}</strong>
                <small>
                    Bodega utilizada:
                    <b>CASA MATRIZ</b>
                    · {loaded_at}
                </small>
            </div>
            <div class="mk2-source-rule">
                <span>REGLA</span>
                <strong>Solo Casa Matriz</strong>
                <small>CD, Patronato y Concepción excluidos</small>
            </div>
        </div>
        """
    )

    render_html(
        f"""
        <div class="mk2-kpi-grid">
            <div class="mk2-kpi">
                <span>SKU CASA MATRIZ</span>
                <strong>{total_skus:,}</strong>
                <small>SKU únicos consolidados</small>
            </div>

            <div class="mk2-kpi">
                <span>UNIDADES DISPONIBLES</span>
                <strong>{total_available:,}</strong>
                <small>stock actual Casa Matriz</small>
            </div>

            <div class="mk2-kpi">
                <span>SKU CON STOCK</span>
                <strong>{positive_skus:,}</strong>
                <small>disponibles para publicación</small>
            </div>

            <div class="mk2-kpi alert">
                <span>SIN DISPONIBILIDAD</span>
                <strong>{zero_skus:,}</strong>
                <small>SKU con stock cero o menor</small>
            </div>
        </div>
        """
    )

    render_html(
        """
        <div class="mk2-rule">
            <div class="mk2-rule-icon">i</div>
            <div>
                <strong>Regla de publicación</strong>
                <span>
                    Marketplace toma solamente el stock disponible de
                    <b>Casa Matriz</b> proveniente de
                    <b>Llegadas_OK</b>.
                    El stock de CD, Patronato y Concepción nunca se suma
                    a las plantillas.
                </span>
            </div>
        </div>
        """
    )

    # Streamlit ejecuta el contenido de todos los st.tabs, incluso los
    # que no están visibles. El selector permite procesar solo una
    # plantilla a la vez y reduce mucho el tiempo de entrada a la página.
    platform = st.radio(
        "Marketplace",
        [
            "Paris",
            "Mercado Libre",
        ],
        horizontal=True,
        label_visibility="collapsed",
        key="market_platform_v3",
    )

    if platform == "Paris":
        _render_marketplace_panel(
            name="Paris Marketplace",
            key_name="paris",
            house=house,
        )
    else:
        _render_marketplace_panel(
            name="Mercado Libre",
            key_name="meli",
            house=house,
        )
