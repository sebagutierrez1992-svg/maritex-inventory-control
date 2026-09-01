

from io import BytesIO
import re

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from analytics.stock_metrics import (
    stock_view,
    consolidate_inventory,
)
from config.settings import MARKETPLACE_TEMPLATES
from ui.components import render_html


# ============================================================
# HELPERS GENERALES
# ============================================================

def _safe_int(value) -> int:
    try:
        if pd.isna(value):
            return 0
        return int(round(float(value)))
    except Exception:
        return 0


def _normalize_sku(value) -> str:
    """
    Normaliza SKU para cruces entre Llegadas_OK y marketplaces.

    Ejemplos:
        13051205  -> 13051205
        1305120-5 -> 13051205
        200122.0  -> 200122
    """
    if value is None:
        return ""

    text = str(value).strip().upper()

    if not text:
        return ""

    text = re.sub(r"\.0$", "", text)
    text = re.sub(r"[^A-Z0-9]", "", text)

    return text


def _prepare_house_stock(
    df: pd.DataFrame,
) -> pd.DataFrame:
    """
    Fuente Marketplace:
        Llegadas_OK
        -> stock_view
        -> SOLO CASA MATRIZ
        -> consolidación por SKU

    CD, Patronato y Concepción NO participan.
    """
    stock = stock_view(df)

    if stock is None or stock.empty:
        return pd.DataFrame()

    warehouse = (
        stock["Bodega"]
        .fillna("")
        .astype(str)
        .str.strip()
        .str.upper()
    )

    house = stock[
        warehouse.str.contains(
            "CASA MATRIZ",
            regex=False,
        )
    ].copy()

    if house.empty:
        return pd.DataFrame()

    house = consolidate_inventory(
        house
    )

    house["_sku_match"] = (
        house["Código"]
        .map(_normalize_sku)
    )

    house = house[
        house["_sku_match"].ne("")
    ].copy()

    if house["_sku_match"].duplicated().any():
        agg = {
            "Código": "first",
            "Producto": "first",
            "Disponible": "sum",
        }

        for col in [
            "Stock físico",
            "Por llegar",
            "Por despachar",
        ]:
            if col in house.columns:
                agg[col] = "sum"

        house = (
            house.groupby(
                "_sku_match",
                as_index=False,
            )
            .agg(agg)
        )

    house["_search_codigo"] = (
        house["Código"]
        .fillna("")
        .astype(str)
        .str.lower()
    )

    house["_search_producto"] = (
        house["Producto"]
        .fillna("")
        .astype(str)
        .str.lower()
    )

    return house.reset_index(
        drop=True
    )


def _stock_map(
    house: pd.DataFrame,
    reserve: int,
) -> dict:
    """
    Stock a publicar:
        max(Disponible Casa Matriz - reserva, 0)
    """
    available = pd.to_numeric(
        house["Disponible"],
        errors="coerce",
    ).fillna(0)

    publishable = (
        available
        - int(reserve)
    ).clip(
        lower=0
    ).round().astype(int)

    return dict(
        zip(
            house["_sku_match"],
            publishable,
        )
    )


def _find_header_column(
    ws,
    row_number: int,
    expected_names,
):
    expected = {
        re.sub(
            r"[^A-Z0-9]",
            "",
            str(name).upper(),
        )
        for name in expected_names
    }

    for cell in ws[row_number]:
        value = cell.value

        if value is None:
            continue

        normalized = re.sub(
            r"[^A-Z0-9]",
            "",
            str(value).strip().upper(),
        )

        if normalized in expected:
            return cell.column

    return None


def _change_status(
    current_stock,
    new_stock,
    found: bool,
) -> tuple[str, int]:
    current = _safe_int(current_stock)
    new = _safe_int(new_stock)
    delta = new - current

    if not found:
        return "Sin coincidencia", delta

    if new == current:
        return "Sin cambios", delta

    if new == 0 and current != 0:
        return "Queda en cero", delta

    if delta > 0:
        return "Sube stock", delta

    return "Baja stock", delta


def _stats_from_preview(
    preview: pd.DataFrame,
    base_stats: dict,
) -> dict:
    stats = dict(base_stats)

    if preview is None or preview.empty:
        stats.update(
            {
                "up": 0,
                "down": 0,
                "zero": 0,
                "same": 0,
                "match_pct": 0.0,
            }
        )
        return stats

    status = (
        preview["Cambio"]
        if "Cambio" in preview.columns
        else pd.Series("", index=preview.index)
    )

    stats["up"] = int(
        status.eq("Sube stock").sum()
    )
    stats["down"] = int(
        status.eq("Baja stock").sum()
    )
    stats["zero"] = int(
        status.eq("Queda en cero").sum()
    )
    stats["same"] = int(
        status.eq("Sin cambios").sum()
    )

    rows = max(
        _safe_int(
            stats.get("rows", 0)
        ),
        0,
    )

    matched = max(
        _safe_int(
            stats.get("matched_rows", 0)
        ),
        0,
    )

    stats["match_pct"] = (
        (matched / rows) * 100
        if rows > 0
        else 0.0
    )

    return stats


# ============================================================
# PARIS
# ============================================================

@st.cache_data(
    show_spinner=False,
)
def _build_paris_workbook(
    template_bytes: bytes,
    stock_items: tuple,
):
    """
    Paris:
        Hoja       -> stock
        Cruce      -> sku_seller
        Stock base -> stock
        Escribe    -> nuevo_stock
    """
    stock_lookup = dict(
        stock_items
    )

    wb = load_workbook(
        BytesIO(template_bytes)
    )

    if "stock" not in wb.sheetnames:
        raise ValueError(
            "La plantilla Paris no contiene la hoja 'stock'."
        )

    ws = wb["stock"]

    sku_col = _find_header_column(
        ws,
        1,
        [
            "sku_seller",
            "sku seller",
        ],
    )

    target_col = _find_header_column(
        ws,
        1,
        [
            "nuevo_stock",
            "nuevo stock",
        ],
    )

    current_stock_col = _find_header_column(
        ws,
        1,
        [
            "stock",
        ],
    )

    title_col = _find_header_column(
        ws,
        1,
        [
            "titulo",
            "título",
        ],
    )

    size_col = _find_header_column(
        ws,
        1,
        [
            "talla",
        ],
    )

    mkp_col = _find_header_column(
        ws,
        1,
        [
            "sku_mkp",
            "sku mkp",
        ],
    )

    if sku_col is None:
        raise ValueError(
            "No se encontró la columna 'sku_seller' "
            "en la plantilla Paris."
        )

    if target_col is None:
        raise ValueError(
            "No se encontró la columna 'nuevo_stock' "
            "en la plantilla Paris."
        )

    preview_rows = []
    matched_rows = 0
    unmatched_rows = 0
    publishable_units = 0
    processed_rows = 0

    matched_skus = set()
    unmatched_skus = set()

    for row in range(
        2,
        ws.max_row + 1,
    ):
        raw_sku = ws.cell(
            row=row,
            column=sku_col,
        ).value

        if raw_sku is None:
            continue

        sku_text = str(
            raw_sku
        ).strip()

        if not sku_text:
            continue

        processed_rows += 1

        sku_key = _normalize_sku(
            raw_sku
        )

        found = (
            sku_key in stock_lookup
        )

        new_stock = int(
            stock_lookup.get(
                sku_key,
                0,
            )
        )

        current_stock = (
            ws.cell(
                row=row,
                column=current_stock_col,
            ).value
            if current_stock_col
            else 0
        )

        change, delta = _change_status(
            current_stock,
            new_stock,
            found,
        )

        ws.cell(
            row=row,
            column=target_col,
        ).value = new_stock

        if found:
            matched_rows += 1
            matched_skus.add(
                sku_key
            )
        else:
            unmatched_rows += 1
            unmatched_skus.add(
                sku_key
            )

        publishable_units += new_stock

        preview_rows.append(
            {
                "SKU Marketplace": (
                    ws.cell(
                        row=row,
                        column=mkp_col,
                    ).value
                    if mkp_col
                    else ""
                ),
                "SKU Seller": sku_text,
                "Producto": (
                    ws.cell(
                        row=row,
                        column=title_col,
                    ).value
                    if title_col
                    else ""
                ),
                "Talla": (
                    ws.cell(
                        row=row,
                        column=size_col,
                    ).value
                    if size_col
                    else ""
                ),
                "Stock actual": _safe_int(
                    current_stock
                ),
                "Nuevo stock": new_stock,
                "Diferencia": delta,
                "Cambio": change,
                "Coincidencia Stock CM": (
                    "Encontrado"
                    if found
                    else "Sin coincidencia"
                ),
            }
        )

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    preview = pd.DataFrame(
        preview_rows
    )

    stats = _stats_from_preview(
        preview,
        {
            "rows": processed_rows,
            "matched_rows": matched_rows,
            "unmatched_rows": unmatched_rows,
            "matched_skus": len(matched_skus),
            "unmatched_skus": len(unmatched_skus),
            "units": publishable_units,
        },
    )

    return (
        output.getvalue(),
        preview,
        stats,
    )


# ============================================================
# MERCADO LIBRE
# ============================================================

@st.cache_data(
    show_spinner=False,
)
def _build_meli_workbook(
    template_bytes: bytes,
    stock_items: tuple,
):
    """
    Mercado Libre:
        Hoja       -> Publicaciones
        Cruce      -> SKU
        Stock base -> QUANTITY
        Escribe    -> QUANTITY

    Las filas de publicaciones comienzan en la fila 6.
    """
    stock_lookup = dict(
        stock_items
    )

    wb = load_workbook(
        BytesIO(template_bytes)
    )

    if "Publicaciones" not in wb.sheetnames:
        raise ValueError(
            "La plantilla Mercado Libre no contiene "
            "la hoja 'Publicaciones'."
        )

    ws = wb["Publicaciones"]

    sku_col = _find_header_column(
        ws,
        1,
        ["SKU"],
    )

    quantity_col = _find_header_column(
        ws,
        1,
        ["QUANTITY"],
    )

    title_col = _find_header_column(
        ws,
        1,
        ["TITLE"],
    )

    variation_col = _find_header_column(
        ws,
        1,
        ["VARIATIONS"],
    )

    item_col = _find_header_column(
        ws,
        1,
        ["ITEM_ID"],
    )

    if sku_col is None:
        raise ValueError(
            "No se encontró la columna 'SKU' "
            "en la plantilla Mercado Libre."
        )

    if quantity_col is None:
        raise ValueError(
            "No se encontró la columna 'QUANTITY' "
            "en la plantilla Mercado Libre."
        )

    preview_rows = []
    matched_rows = 0
    unmatched_rows = 0
    publishable_units = 0
    processed_rows = 0

    matched_skus = set()
    unmatched_skus = set()

    for row in range(
        6,
        ws.max_row + 1,
    ):
        raw_sku = ws.cell(
            row=row,
            column=sku_col,
        ).value

        if raw_sku is None:
            continue

        sku_text = str(
            raw_sku
        ).strip()

        if not sku_text:
            continue

        processed_rows += 1

        sku_key = _normalize_sku(
            raw_sku
        )

        found = (
            sku_key in stock_lookup
        )

        current_stock = ws.cell(
            row=row,
            column=quantity_col,
        ).value

        new_stock = int(
            stock_lookup.get(
                sku_key,
                0,
            )
        )

        change, delta = _change_status(
            current_stock,
            new_stock,
            found,
        )

        # Solo se actualiza QUANTITY.
        ws.cell(
            row=row,
            column=quantity_col,
        ).value = new_stock

        if found:
            matched_rows += 1
            matched_skus.add(
                sku_key
            )
        else:
            unmatched_rows += 1
            unmatched_skus.add(
                sku_key
            )

        publishable_units += new_stock

        raw_title = (
            ws.cell(
                row=row,
                column=title_col,
            ).value
            if title_col
            else ""
        )

        product_title = raw_title or ""

        # MELI puede repetir el TITLE mediante fórmula.
        if (
            title_col
            and isinstance(
                raw_title,
                str,
            )
            and raw_title.strip().startswith("=")
        ):
            for previous_row in range(
                row - 1,
                5,
                -1,
            ):
                previous_title = ws.cell(
                    row=previous_row,
                    column=title_col,
                ).value

                if previous_title is None:
                    continue

                previous_title = str(
                    previous_title
                ).strip()

                if (
                    not previous_title
                    or previous_title.startswith("=")
                ):
                    continue

                product_title = previous_title
                break

        preview_rows.append(
            {
                "Publicación": (
                    ws.cell(
                        row=row,
                        column=item_col,
                    ).value
                    if item_col
                    else ""
                ),
                "SKU": sku_text,
                "Producto": product_title,
                "Variación": (
                    ws.cell(
                        row=row,
                        column=variation_col,
                    ).value
                    if variation_col
                    else ""
                ),
                "Stock actual": _safe_int(
                    current_stock
                ),
                "Nuevo stock": new_stock,
                "Diferencia": delta,
                "Cambio": change,
                "Coincidencia Stock CM": (
                    "Encontrado"
                    if found
                    else "Sin coincidencia"
                ),
            }
        )

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    preview = pd.DataFrame(
        preview_rows
    )

    stats = _stats_from_preview(
        preview,
        {
            "rows": processed_rows,
            "matched_rows": matched_rows,
            "unmatched_rows": unmatched_rows,
            "matched_skus": len(matched_skus),
            "unmatched_skus": len(unmatched_skus),
            "units": publishable_units,
        },
    )

    return (
        output.getvalue(),
        preview,
        stats,
    )


# ============================================================
# COMPONENTES DE INTERFAZ
# ============================================================

def _marketplace_header(
    name: str,
    filename: str,
    stats: dict,
):
    pct = float(
        stats.get(
            "match_pct",
            0.0,
        )
    )

    badge_class = (
        "ok"
        if pct >= 98
        else "warn"
        if pct >= 90
        else "risk"
    )

    render_html(
        f"""
        <div class="mk2-platform-head">
            <div class="mk2-platform-title">
                <div class="mk2-platform-mark">
                    {"P" if "Paris" in name else "ML"}
                </div>
                <div>
                    <strong>{name}</strong>
                    <span>Plantilla oficial · {filename}</span>
                </div>
            </div>

            <div class="mk2-platform-badges">
                <span class="mk2-badge neutral">CASA MATRIZ</span>
                <span class="mk2-badge {badge_class}">
                    {pct:.1f}% MATCH
                </span>
            </div>
        </div>
        """
    )


def _filter_preview(
    preview: pd.DataFrame,
    search: str,
    change_filter: str,
) -> pd.DataFrame:
    view = preview.copy()

    if search:
        term = search.strip().lower()

        text_columns = [
            col
            for col in view.columns
            if (
                view[col].dtype == object
                or pd.api.types.is_string_dtype(
                    view[col]
                )
            )
        ]

        mask = pd.Series(
            False,
            index=view.index,
        )

        for col in text_columns:
            mask = (
                mask
                |
                view[col]
                .fillna("")
                .astype(str)
                .str.lower()
                .str.contains(
                    term,
                    regex=False,
                )
            )

        view = view[
            mask
        ].copy()

    if change_filter == "Sin coincidencia":
        if "Coincidencia Stock CM" in view.columns:
            view = view[
                view["Coincidencia Stock CM"].eq("Sin coincidencia")
            ].copy()

    elif change_filter == "Con cambio":
        if "Cambio" in view.columns:
            view = view[
                ~view["Cambio"].eq("Sin cambios")
            ].copy()

    return view


# ============================================================
# PANEL MARKETPLACE
# ============================================================

def _render_marketplace_panel(
    name: str,
    key_name: str,
    house: pd.DataFrame,
):
    path = MARKETPLACE_TEMPLATES[
        name
    ]

    if not path.exists():
        render_html(
            f"""
            <div class="mk2-platform-head">
                <div class="mk2-platform-title">
                    <div class="mk2-platform-mark">
                        {"P" if "Paris" in name else "ML"}
                    </div>
                    <div>
                        <strong>{name}</strong>
                        <span>Plantilla pendiente</span>
                    </div>
                </div>
                <span class="mk2-badge risk">SIN PLANTILLA</span>
            </div>
            """
        )

        st.warning(
            f"No existe la plantilla oficial de {name}. "
            "Cárgala desde Plantillas."
        )
        return

    # --------------------------------------------------------
    # CONTROLES
    # --------------------------------------------------------
    c1, c2 = st.columns(
        [1.55, 0.75],
        gap="medium",
    )

    with c1:
        search = st.text_input(
            "Buscar SKU o producto",
            placeholder=(
                "Ej: 13051205, parka, softshell..."
            ),
            key=f"market_search_v2_{key_name}",
        )

    with c2:
        reserve = st.number_input(
            "Stock de seguridad",
            min_value=0,
            value=0,
            step=1,
            help=(
                "Stock a publicar = Stock Casa Matriz - reserva. "
                "Nunca se publican cantidades negativas."
            ),
            key=f"market_reserve_v2_{key_name}",
        )

    lookup = _stock_map(
        house,
        reserve,
    )

    stock_items = tuple(
        sorted(
            lookup.items()
        )
    )

    template_bytes = path.read_bytes()

    try:
        if name == "Paris Marketplace":
            (
                output_bytes,
                preview,
                stats,
            ) = _build_paris_workbook(
                template_bytes,
                stock_items,
            )

            download_name = (
                "Paris_stock_actualizado.xlsx"
            )

            button_label = (
                "Descargar Paris actualizado"
            )

        else:
            (
                output_bytes,
                preview,
                stats,
            ) = _build_meli_workbook(
                template_bytes,
                stock_items,
            )

            download_name = (
                "Mercado_Libre_stock_actualizado.xlsx"
            )

            button_label = (
                "Descargar Mercado Libre actualizado"
            )

    except Exception as exc:
        st.error(
            f"No fue posible procesar la plantilla: {exc}"
        )
        return

    _marketplace_header(
        name,
        path.name,
        stats,
    )

    # --------------------------------------------------------
    # KPI CRUCE
    # --------------------------------------------------------
    render_html(
        f"""
        <div class="mk2-summary-grid">
            <div>
                <span>FILAS PLANTILLA</span>
                <strong>{stats["rows"]:,}</strong>
                <small>registros procesados</small>
            </div>

            <div>
                <span>SKU ENCONTRADOS</span>
                <strong>{stats["matched_rows"]:,}</strong>
                <small>coinciden con Casa Matriz</small>
            </div>

            <div>
                <span>SIN COINCIDENCIA</span>
                <strong>{stats["unmatched_rows"]:,}</strong>
                <small>se exportarán con stock 0</small>
            </div>

            <div>
                <span>STOCK A PUBLICAR</span>
                <strong>{stats["units"]:,}</strong>
                <small>unidades después de reserva</small>
            </div>
        </div>
        """
    )

    # --------------------------------------------------------
    # FILTRO
    # --------------------------------------------------------
    change_filter = st.radio(
        "Mostrar",
        [
            "Todos",
            "Con cambio",
            "Sin coincidencia",
        ],
        horizontal=True,
        key=f"market_change_filter_v2_{key_name}",
    )

    preview_view = _filter_preview(
        preview,
        search,
        change_filter,
    )

    render_html(
        f"""
        <div class="mk2-table-head">
            <div>
                <strong>Stock a publicar</strong>
                <span>
                    {len(preview_view):,} registros visibles · Casa Matriz
                </span>
            </div>
            <div>
                Stock de seguridad:
                <b>{int(reserve)}</b>
            </div>
        </div>
        """
    )

    column_config = {
        "Stock actual": st.column_config.NumberColumn(
            "Stock actual",
            format="%d",
        ),
        "Nuevo stock": st.column_config.NumberColumn(
            "Nuevo stock",
            format="%d",
        ),
        "Diferencia": st.column_config.NumberColumn(
            "Diferencia",
            format="%+d",
        ),
        "Cambio": st.column_config.TextColumn(
            "Cambio",
            width="medium",
        ),
        "Coincidencia Stock CM": st.column_config.TextColumn(
            "Coincidencia",
            width="medium",
        ),
    }

    st.dataframe(
        preview_view,
        hide_index=True,
        use_container_width=True,
        height=430,
        column_config=column_config,
    )

    if stats["unmatched_rows"] > 0:
        st.warning(
            f"{stats['unmatched_rows']:,} fila(s) no tienen "
            "coincidencia con el stock de Casa Matriz de Llegadas_OK. "
            "Esas filas se exportarán con stock 0."
        )

    st.download_button(
        button_label,
        data=output_bytes,
        file_name=download_name,
        mime=(
            "application/"
            "vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        use_container_width=True,
        type="primary",
        icon=":material/download:",
        key=f"market_download_v2_{key_name}",
    )


# ============================================================
# RENDER PRINCIPAL
# ============================================================

def render(ctx):
    # Tema visual Marketplace · Dark corporativo Maritex.
    # Solo presentación: no altera cruces, stock, plantillas ni exportación.
    st.markdown(
        """
        <style>
        .mk2-page-head{
            display:flex;justify-content:space-between;align-items:flex-start;
            gap:18px;margin:2px 0 18px;
        }
        .mk2-eyebrow{
            color:#FFC400!important;font-size:10px!important;font-weight:800!important;
            letter-spacing:.8px!important;margin-bottom:5px!important;
        }
        .mk2-title{
            color:#F7F8FA!important;font-size:30px!important;font-weight:800!important;
            line-height:1.08!important;
        }
        .mk2-subtitle{color:#9FB0C0!important;margin-top:7px!important;}
        .mk2-live{
            color:#B8C5CF!important;background:#151F28!important;
            border:1px solid #34414D!important;border-radius:999px!important;
            padding:8px 12px!important;font-size:11px!important;white-space:nowrap!important;
        }
        .mk2-live i{
            display:inline-block;width:8px;height:8px;border-radius:50%;
            background:#22C55E;margin-right:7px;box-shadow:0 0 0 4px rgba(34,197,94,.10);
        }

        .mk2-source{
            display:grid!important;grid-template-columns:auto 1fr auto!important;
            align-items:center!important;gap:14px!important;
            background:linear-gradient(145deg,#18242E,#121B23)!important;
            border:1px solid #34414D!important;border-radius:12px!important;
            padding:15px 17px!important;margin:0 0 10px!important;
        }
        .mk2-source-icon{
            width:42px;height:42px;border-radius:10px;display:flex;align-items:center;
            justify-content:center;background:#3B3007!important;color:#FFC400!important;
            border:1px solid rgba(255,196,0,.35)!important;font-weight:900!important;
        }
        .mk2-source-main span,.mk2-source-rule span{
            color:#8FA2B3!important;font-size:9px!important;font-weight:800!important;
            letter-spacing:.6px!important;
        }
        .mk2-source-main strong,.mk2-source-rule strong{
            display:block;color:#F7F8FA!important;font-weight:800!important;margin-top:2px!important;
        }
        .mk2-source-main small,.mk2-source-rule small{color:#9FB0C0!important;}
        .mk2-source-main small b{color:#FFC400!important;}
        .mk2-source-rule{
            padding-left:18px!important;border-left:1px solid #34414D!important;
        }

        .mk3-compact-summary{
            display:flex!important;gap:10px!important;flex-wrap:wrap!important;
            background:#111B24!important;border:1px solid #34414D!important;
            border-radius:10px!important;padding:10px 13px!important;margin-bottom:14px!important;
            color:#9FB0C0!important;
        }
        .mk3-compact-summary span{
            padding-right:12px!important;border-right:1px solid #34414D!important;
        }
        .mk3-compact-summary span:last-child{border-right:0!important;}
        .mk3-compact-summary b{color:#F7F8FA!important;}

        .mk2-platform-head{
            display:flex!important;justify-content:space-between!important;align-items:center!important;
            gap:14px!important;background:linear-gradient(145deg,#18242E,#121B23)!important;
            border:1px solid #34414D!important;border-radius:12px!important;
            padding:14px 16px!important;margin:8px 0 12px!important;
        }
        .mk2-platform-title{display:flex!important;align-items:center!important;gap:11px!important;}
        .mk2-platform-mark{
            min-width:38px;height:38px;border-radius:9px;display:flex;align-items:center;
            justify-content:center;background:#3B3007!important;color:#FFC400!important;
            border:1px solid rgba(255,196,0,.32)!important;font-weight:900!important;
        }
        .mk2-platform-title strong{display:block;color:#F7F8FA!important;}
        .mk2-platform-title span{display:block;color:#9FB0C0!important;font-size:10px!important;margin-top:2px!important;}
        .mk2-platform-badges{display:flex!important;gap:7px!important;flex-wrap:wrap!important;}
        .mk2-badge{
            border-radius:999px!important;padding:5px 9px!important;font-size:9px!important;
            font-weight:800!important;letter-spacing:.35px!important;
        }
        .mk2-badge.neutral{background:#202C36!important;color:#C7D1D9!important;border:1px solid #40505D!important;}
        .mk2-badge.ok{background:#123824!important;color:#70DF96!important;border:1px solid #236A40!important;}
        .mk2-badge.warn{background:#453606!important;color:#FFD75A!important;border:1px solid #806407!important;}
        .mk2-badge.risk{background:#4A2020!important;color:#FF8181!important;border:1px solid #843737!important;}

        .mk2-summary-grid{
            display:grid!important;grid-template-columns:repeat(4,minmax(0,1fr))!important;
            gap:12px!important;margin:12px 0 14px!important;
        }
        .mk2-summary-grid>div{
            background:linear-gradient(145deg,#18242E,#121B23)!important;
            border:1px solid #34414D!important;border-radius:11px!important;
            padding:14px!important;min-height:94px!important;
        }
        .mk2-summary-grid span{
            display:block;color:#9FB0C0!important;font-size:9px!important;
            font-weight:800!important;letter-spacing:.45px!important;
        }
        .mk2-summary-grid strong{
            display:block;color:#F7F8FA!important;font-size:22px!important;
            line-height:1.15!important;margin:7px 0 4px!important;
        }
        .mk2-summary-grid small{color:#8FA2B3!important;}

        .mk2-table-head{
            display:flex!important;justify-content:space-between!important;align-items:end!important;
            gap:12px!important;margin:14px 0 8px!important;
        }
        .mk2-table-head strong{display:block;color:#F7F8FA!important;font-size:14px!important;}
        .mk2-table-head span,.mk2-table-head>div:last-child{color:#9FB0C0!important;font-size:10px!important;}
        .mk2-table-head b{color:#FFC400!important;}

        [data-testid="stTabs"] [data-baseweb="tab-list"]{
            gap:8px!important;border-bottom:1px solid #34414D!important;
        }
        [data-testid="stTabs"] button[role="tab"]{
            color:#9FB0C0!important;background:transparent!important;
            border-radius:8px 8px 0 0!important;
        }
        [data-testid="stTabs"] button[role="tab"][aria-selected="true"]{
            color:#FFC400!important;background:#18242E!important;
        }
        [data-testid="stTabs"] [data-baseweb="tab-highlight"]{background:#FFC400!important;}

        [data-testid="stTextInput"] label p,
        [data-testid="stNumberInput"] label p,
        [data-testid="stRadio"] label p{color:#B8C5CF!important;}
        [data-testid="stTextInput"] input,
        [data-testid="stNumberInput"] input{
            background:#141E27!important;color:#F7F8FA!important;
        }
        [data-testid="stRadio"] [role="radiogroup"]{
            background:#111B24!important;border:1px solid #34414D!important;
            border-radius:10px!important;padding:5px 9px!important;
        }

        [data-testid="stDataFrame"]{
            background:#111B24!important;border:1px solid #34414D!important;
            border-radius:10px!important;overflow:hidden!important;
        }

        [data-testid="stAlert"]{
            background:#2A240E!important;border-color:#6F5B0B!important;color:#F5DE87!important;
        }

        .main .stDownloadButton>button[kind="primary"],
        .main .stDownloadButton>button{
            background:#FFC400!important;color:#111820!important;
            border:1px solid #FFC400!important;font-weight:800!important;
        }
        .main .stDownloadButton>button:hover{
            background:#FFD02D!important;color:#111820!important;border-color:#FFD02D!important;
        }

        @media(max-width:900px){
            .mk2-source{grid-template-columns:auto 1fr!important;}
            .mk2-source-rule{grid-column:1/-1;border-left:0!important;border-top:1px solid #34414D!important;padding:10px 0 0!important;}
            .mk2-summary-grid{grid-template-columns:repeat(2,minmax(0,1fr))!important;}
        }

        /* ============================================================
           MARKETPLACE · CONTRASTE FINAL V2
           Solo mejora legibilidad; no modifica lógica ni cálculos.
           ============================================================ */

        .mk2-source-main strong,
        .mk2-source-rule strong,
        .mk2-platform-title strong,
        .mk2-summary-grid strong,
        .mk2-table-head strong {
            color:#F7F8FA !important;
            opacity:1 !important;
        }

        .mk2-source-main span,
        .mk2-source-rule span,
        .mk2-source-main small,
        .mk2-source-rule small,
        .mk2-platform-title span,
        .mk2-summary-grid span,
        .mk2-summary-grid small,
        .mk2-table-head span,
        .mk2-table-head > div:last-child {
            color:#AEBBC6 !important;
            opacity:1 !important;
        }

        .mk2-source-main small b,
        .mk2-table-head b {
            color:#FFC400 !important;
        }

        .mk2-summary-grid > div {
            background:linear-gradient(145deg,#1A2530,#141E27) !important;
        }

        .mk2-summary-grid > div strong {
            color:#FFFFFF !important;
            font-size:23px !important;
            font-weight:850 !important;
            text-shadow:none !important;
        }

        .mk2-platform-head {
            background:linear-gradient(145deg,#1A2530,#141E27) !important;
        }

        .mk2-platform-title strong {
            color:#FFFFFF !important;
            font-weight:800 !important;
        }

        .mk2-badge.neutral {
            background:#26333D !important;
            color:#E9EEF2 !important;
            border-color:#465764 !important;
        }

        .mk2-badge.risk {
            background:#4A2020 !important;
            color:#FF8A8A !important;
            border-color:#8F3A3A !important;
        }

        .mk2-live {
            background:#151F28 !important;
            color:#D9E1E7 !important;
            border-color:#3A4955 !important;
        }

        .mk3-compact-summary b {
            color:#FFFFFF !important;
        }

        [data-testid="stTabs"] button[role="tab"] {
            color:#CFD8DF !important;
            font-weight:700 !important;
        }

        [data-testid="stTabs"] button[role="tab"][aria-selected="true"] {
            color:#FFC400 !important;
        }

        </style>
        """,
        unsafe_allow_html=True,
    )
    df = ctx.get(
        "stock_df"
    )

    meta = (
        ctx.get(
            "stock_meta"
        )
        or {}
    )

    render_html(
        """
        <div class="mk2-page-head">
            <div>
                <div class="mk2-eyebrow">
                    MARITEX / OPERACIÓN
                </div>
                <div class="mk2-title">
                    Marketplace
                </div>
                <div class="mk2-subtitle">
                    Sincronización de stock para Paris y Mercado Libre
                    utilizando exclusivamente Casa Matriz.
                </div>
            </div>
            <div class="mk2-live">
                <i></i>
                Stock automático
            </div>
        </div>
        """
    )

    if df is None or df.empty:
        st.info(
            "No hay stock disponible desde Llegadas_OK."
        )
        return

    house = _prepare_house_stock(
        df
    )

    if house.empty:
        st.warning(
            "Llegadas_OK no contiene registros asociados "
            "a Casa Matriz."
        )
        return

    available = pd.to_numeric(
        house["Disponible"],
        errors="coerce",
    ).fillna(0)

    total_skus = int(
        house["_sku_match"]
        .nunique()
    )

    total_available = _safe_int(
        available.sum()
    )

    positive_skus = int(
        (available > 0).sum()
    )

    zero_skus = int(
        (available <= 0).sum()
    )

    source_name = (
        meta.get(
            "filename"
        )
        or "Stock automático · Llegadas_OK"
    )

    loaded_at = (
        meta.get(
            "loaded_at"
        )
        or meta.get(
            "generated_at"
        )
        or "actualización automática"
    )

    render_html(
        f"""
        <div class="mk2-source">
            <div class="mk2-source-icon">CM</div>
            <div class="mk2-source-main">
                <span>FUENTE DE STOCK</span>
                <strong>{source_name}</strong>
                <small>
                    Bodega utilizada:
                    <b>CASA MATRIZ</b>
                    · {loaded_at}
                </small>
            </div>
            <div class="mk2-source-rule">
                <span>REGLA</span>
                <strong>Solo Casa Matriz</strong>
                <small>CD, Patronato y Concepción excluidos</small>
            </div>
        </div>
        """
    )

    render_html(
        f"""
        <div class="mk3-compact-summary">
            <span><b>{total_skus:,}</b> SKU CM</span>
            <span><b>{total_available:,}</b> unidades disponibles</span>
            <span><b>{positive_skus:,}</b> con stock</span>
            <span><b>{zero_skus:,}</b> sin disponibilidad</span>
        </div>
        """
    )

    paris_tab, meli_tab = st.tabs(
        [
            "Paris",
            "Mercado Libre",
        ]
    )

    with paris_tab:
        _render_marketplace_panel(
            name="Paris Marketplace",
            key_name="paris",
            house=house,
        )

    with meli_tab:
        _render_marketplace_panel(
            name="Mercado Libre",
            key_name="meli",
            house=house,
        )
