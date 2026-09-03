

import hashlib
from datetime import datetime
from io import BytesIO
from pathlib import Path

import streamlit as st
from openpyxl import load_workbook

from config.settings import (
    ERP_SALES_FILE,
    ERP_SALES_META,
    MARKETPLACE_TEMPLATES,
)
from services.erp_sales import read_sales_source
from services.storage import save_source, load_source
from services.validation import validate_sales_source
from ui.components import render_html
from utils.numbers import format_clp


def _norm(value) -> str:
    if value is None:
        return ""

    return (
        str(value)
        .strip()
        .lower()
        .replace(" ", "")
        .replace("_", "")
        .replace("-", "")
    )


def _headers(ws, row: int = 1) -> set[str]:
    return {
        _norm(cell.value)
        for cell in ws[row]
        if cell.value is not None
    }


def _validate_template(
    name: str,
    raw: bytes,
) -> tuple[bool, str]:

    try:
        wb = load_workbook(
            BytesIO(raw),
            read_only=True,
            data_only=False,
        )

    except Exception as exc:
        return False, f"No se pudo abrir el archivo: {exc}"

    if name == "Paris Marketplace":

        if "stock" not in wb.sheetnames:
            return (
                False,
                "La plantilla Paris debe contener la hoja 'stock'.",
            )

        headers = _headers(
            wb["stock"],
            1,
        )

        missing = {
            "skuseller",
            "nuevostock",
        } - headers

        if missing:
            return (
                False,
                "Faltan columnas: "
                + ", ".join(
                    sorted(missing)
                ),
            )

        return True, "Plantilla Paris válida."

    if name == "Mercado Libre":

        if "Publicaciones" not in wb.sheetnames:
            return (
                False,
                "La plantilla Mercado Libre debe contener "
                "la hoja 'Publicaciones'.",
            )

        # Algunas planillas MELI contienen metadata
        # antes de la fila real de encabezados.
        ws = wb["Publicaciones"]

        found = False

        for row in range(
            1,
            min(ws.max_row, 10) + 1,
        ):
            headers = _headers(
                ws,
                row,
            )

            if {
                "sku",
                "quantity",
            }.issubset(headers):
                found = True
                break

        if not found:
            return (
                False,
                "No se encontraron las columnas "
                "SKU y QUANTITY.",
            )

        return True, "Plantilla Mercado Libre válida."

    return False, "Marketplace no reconocido."


def _short_name(
    name: str,
) -> str:
    return (
        "Paris"
        if name == "Paris Marketplace"
        else "Mercado Libre"
    )


def _template_uploader_key(
    name: str,
    mode: str,
) -> str:
    """
    Genera una key versionada para cada uploader de plantilla.

    Esto evita que React/Streamlit intente reutilizar un nodo
    file_uploader que ya fue desmontado después de st.rerun().
    """
    state_key = (
        f"tpl3_upload_version_"
        f"{mode}_"
        f"{_norm(name)}"
    )

    if state_key not in st.session_state:
        st.session_state[state_key] = 0

    return (
        f"tpl3_upload_"
        f"{mode}_"
        f"{name}_"
        f"{st.session_state[state_key]}"
    )


def _advance_template_uploader(
    name: str,
    mode: str,
) -> None:

    state_key = (
        f"tpl3_upload_version_"
        f"{mode}_"
        f"{_norm(name)}"
    )

    st.session_state[state_key] = (
        int(
            st.session_state.get(
                state_key,
                0,
            )
        )
        + 1
    )


def _render_template_card(
    name: str,
    path: Path,
):
    short = _short_name(
        name
    )

    exists = path.exists()

    status = (
        "ACTIVA"
        if exists
        else "NO CARGADA"
    )

    status_class = (
        "ok"
        if exists
        else "missing"
    )

    if exists:
        updated = datetime.fromtimestamp(
            path.stat().st_mtime
        ).strftime(
            "%d/%m/%Y %H:%M"
        )

        filename = path.name

    else:
        updated = "—"
        filename = "Sin archivo"

    render_html(
        f"""
        <div class="tpl3-card">
            <div class="tpl3-top">
                <div>
                    <div class="tpl3-name">
                        {short}
                    </div>
                    <div class="tpl3-file">
                        {filename}
                    </div>
                </div>

                <div class="tpl3-status {status_class}">
                    {status}
                </div>
            </div>

            <div class="tpl3-date">
                Actualizada: {updated}
            </div>
        </div>
        """
    )

    if exists:

        c1, c2 = st.columns(
            2,
            gap="small",
        )

        with c1:

            with st.popover(
                "Reemplazar",
                use_container_width=True,
                icon=":material/upload_file:",
            ):

                uploaded = st.file_uploader(
                    f"Nueva plantilla {short}",
                    type=["xlsx"],
                    key=_template_uploader_key(
                        name,
                        "replace",
                    ),
                    label_visibility="collapsed",
                )

                if uploaded is not None:

                    raw = uploaded.getvalue()

                    valid, message = (
                        _validate_template(
                            name,
                            raw,
                        )
                    )

                    if not valid:
                        st.error(
                            message
                        )

                    else:

                        try:
                            path.parent.mkdir(
                                parents=True,
                                exist_ok=True,
                            )

                            path.write_bytes(
                                raw
                            )

                            st.cache_data.clear()

                            st.success(
                                f"✓ {short} actualizada."
                            )

                            _advance_template_uploader(
                                name,
                                "replace",
                            )

                            st.rerun()

                        except Exception as exc:
                            st.error(
                                "No fue posible guardar "
                                f"la plantilla: {exc}"
                            )

        with c2:

            try:
                st.download_button(
                    "Descargar",
                    data=path.read_bytes(),
                    file_name=path.name,
                    mime=(
                        "application/vnd.openxmlformats-"
                        "officedocument.spreadsheetml.sheet"
                    ),
                    use_container_width=True,
                    icon=":material/download:",
                    key=f"tpl3_download_{name}",
                )

            except Exception as exc:
                st.warning(
                    "No fue posible leer "
                    f"la plantilla: {exc}"
                )

    else:

        uploaded = st.file_uploader(
            f"Cargar plantilla {short}",
            type=["xlsx"],
            key=_template_uploader_key(
                name,
                "missing",
            ),
        )

        if uploaded is not None:

            raw = uploaded.getvalue()

            valid, message = (
                _validate_template(
                    name,
                    raw,
                )
            )

            if not valid:
                st.error(
                    message
                )

            else:

                try:
                    path.parent.mkdir(
                        parents=True,
                        exist_ok=True,
                    )

                    path.write_bytes(
                        raw
                    )

                    st.cache_data.clear()

                    st.success(
                        f"✓ {short} asociada."
                    )

                    _advance_template_uploader(
                        name,
                        "missing",
                    )

                    st.rerun()

                except Exception as exc:
                    st.error(
                        "No fue posible guardar "
                        f"la plantilla: {exc}"
                    )


def _render_sales_source():
    """
    Fuente ERP Ventas.

    Se evita st.rerun() inmediatamente después del file_uploader porque,
    en algunas versiones de Streamlit/React, puede provocar:
    NotFoundError: removeChild ... el nodo no es un hijo de este nodo.

    Para impedir que el mismo archivo se procese nuevamente en cada rerun,
    se guarda su SHA256 en session_state.
    """

    _, meta = load_source(
        ERP_SALES_FILE,
        ERP_SALES_META,
    )

    if meta:
        st.success(
            "Fuente activa: "
            f"{meta.get('filename', 'ERP Ventas')} · "
            f"{meta.get('loaded_at', '')}"
        )
    else:
        st.info(
            "Aún no existe una fuente "
            "ERP Ventas guardada."
        )

    uploaded = st.file_uploader(
        "Cargar / reemplazar ERP Ventas",
        type=[
            "csv",
            "xls",
            "xlsx",
        ],
        key="tpl3_sales",
    )

    if uploaded is None:
        return

    try:
        raw = uploaded.getvalue()

        upload_hash = hashlib.sha256(
            raw
        ).hexdigest()

        last_hash = st.session_state.get(
            "tpl3_sales_last_processed_hash"
        )

        if upload_hash == last_hash:
            st.caption(
                "Este archivo ya fue procesado en esta sesión."
            )
            return

        df = read_sales_source(
            raw,
            uploaded.name,
        )

        info = validate_sales_source(
            df
        )

        save_source(
            raw,
            uploaded.name,
            ERP_SALES_FILE,
            ERP_SALES_META,
            info,
        )

        st.session_state[
            "tpl3_sales_last_processed_hash"
        ] = upload_hash

        st.cache_data.clear()

        st.success(
            "✓ ERP Ventas actualizado · "
            f"{info['commercial_rows']:,} documentos · "
            f"{info['min_date']} → {info['max_date']} · "
            f"{format_clp(info['net_sales_with_vat'])}"
        )

        st.caption(
            "La fuente quedó guardada. "
            "Puedes navegar a CRM, Métricas Vendedores "
            "o Resumen Ejecutivo sin volver a cargar el archivo."
        )

    except Exception as exc:
        st.error(
            f"Error cargando ERP Ventas: {exc}"
        )


def render(ctx):

    render_html(
        """
        <div class="tpl3-head">
            <div class="tpl3-title">
                Plantillas
            </div>

            <div class="tpl3-subtitle">
                Archivos base utilizados para generar
                las actualizaciones de Marketplace.
            </div>
        </div>

        <div class="tpl3-source">
            <strong>
                Stock Marketplace
            </strong>

            <span>
                Automático desde Llegadas_OK ·
                solo Casa Matriz.
                Aquí solo administras las plantillas.
            </span>
        </div>
        """
    )

    paris_path = (
        MARKETPLACE_TEMPLATES.get(
            "Paris Marketplace"
        )
    )

    meli_path = (
        MARKETPLACE_TEMPLATES.get(
            "Mercado Libre"
        )
    )

    c1, c2 = st.columns(
        2,
        gap="medium",
    )

    with c1:

        if paris_path is None:
            st.error(
                "Paris no está configurado."
            )

        else:
            _render_template_card(
                "Paris Marketplace",
                paris_path,
            )

    with c2:

        if meli_path is None:
            st.error(
                "Mercado Libre no está configurado."
            )

        else:
            _render_template_card(
                "Mercado Libre",
                meli_path,
            )

    st.markdown("")

    with st.expander(
        "Fuente ERP Ventas",
        expanded=False,
    ):

        st.caption(
            "Usada por Métricas Vendedores, "
            "Resumen Ejecutivo y CRM. "
            "No interviene en el stock de Marketplace."
        )

        _render_sales_source()
